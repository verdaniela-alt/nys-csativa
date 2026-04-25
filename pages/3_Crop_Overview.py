"""
3_Crop_Overview.py — Integrated Crop Data Tool
Merges Pre-Harvest, Post-Harvest, and Batch Dashboard into a single unified page.
"""

import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

st.set_page_config(
    page_title="Crop Overview | NYS Cannabis Tool",
    page_icon="🌿",
    layout="wide",
)

st.markdown("""
<style>
.disclaimer-box {
    background: #fff3cd; border: 2px solid #e0a800;
    border-radius: 8px; padding: 16px 20px; margin-bottom: 20px; font-size: 0.92rem;
}
.batch-badge-ph {
    display: inline-block; padding: 3px 12px; border-radius: 12px;
    background: #d4edda; color: #155724; font-weight: bold; font-size: 0.9rem;
}
.batch-badge-pos {
    display: inline-block; padding: 3px 12px; border-radius: 12px;
    background: #e8f4fd; color: #0c5460; font-weight: bold; font-size: 0.9rem;
}
.weight-card {
    background: #f8f9fa; border-left: 4px solid #2e7d32;
    padding: 10px 14px; border-radius: 6px; margin-bottom: 8px;
}
.status-complete { color: #155724; font-weight: bold; }
.status-partial  { color: #856404; font-weight: bold; }
.status-missing  { color: #721c24; font-weight: bold; }
.section-header  {
    font-size: 1.05rem; font-weight: 600; color: #1565C0;
    border-bottom: 2px solid #1565C0; padding-bottom: 4px; margin-bottom: 12px;
}
</style>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
if "preharvest_batches" not in st.session_state:
    st.session_state.preharvest_batches = {}
if "postharvest_batches" not in st.session_state:
    st.session_state.postharvest_batches = {}

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_preharvest_excel(batches_dict):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1_Batch_Aggregate"
    ws1.append([
        "Batch Number","Farm Name","Strain Name","Seed Source","Flowering Strategy",
        "Season / Cycle","Cultivated Area (sqM)","Growing Media","Soil Type",
        "Field Aspect","Field Steepness",
        "# Plants Outdoors","# Plants Hoop House","# Plants Greenhouse","# Plants Indoors",
        "Germination Rate (%)","First Flower Date","Days in Ground",
        "Disease Presence","Disease Name","Disease Severity","Disease Date First Seen",
        "% White Trichomes","% Clear Trichomes","% Amber Trichomes","Biomass Weight (kg)",
        "THC (%)","CBD (%)","Other Cannabinoids","B-Myrcene (%)","B-Caryophyllene (%)",
    ])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws1.append([
            bn, d.get("farm_name",""), d.get("strain_name",""), d.get("seed_source",""),
            d.get("flowering_strategy",""), d.get("season_cycle",""),
            d.get("cultivated_area_sqm") or "", d.get("growing_media",""), d.get("soil_type",""),
            d.get("field_aspect",""), d.get("field_steepness",""),
            d.get("n_plants_outdoor") or 0, d.get("n_plants_hoop") or 0,
            d.get("n_plants_greenhouse") or 0, d.get("n_plants_indoor") or 0,
            d.get("germination_rate") or "", d.get("first_flower_date",""),
            d.get("days_in_ground") or "",
            d.get("disease_presence",""), d.get("disease_name",""),
            d.get("disease_severity",""), d.get("disease_date",""),
            d.get("pct_white_trichomes") or "", d.get("pct_clear_trichomes") or "",
            d.get("pct_amber_trichomes") or "", d.get("biomass_weight_kg") or "",
            d.get("thc_pct") or "", d.get("cbd_pct") or "",
            d.get("other_cannabinoids",""),
            d.get("b_myrcene_pct") or "", d.get("b_caryophyllene_pct") or "",
        ])
    ws2 = wb.create_sheet("2_Individual_Plants")
    ws2.append(["Batch Number","Plant Number","Plant ID / Tag","Environment","Growing Media",
                "Seed Planting Date","Transplanting Date","Topped (Y/N)","Pruned (Y/N)",
                "Trellised (Y/N)","Height at Harvest (cm)","Stem Width (mm)","Node Count","Bud Size"])
    for bn, b in batches_dict.items():
        df = b.get("plants_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws2.append([bn] + list(row))
    ws3 = wb.create_sheet("3_Inputs_Survey")
    ws3.append(["Batch Number","Mineral Soil","Blood Meal","Green Sand","Feather Meal",
                "Poultry Manure","Livestock Manure","Compost/Worm Cast","Biological",
                "Regular Compost","Plasticulture","Mulch/Cover Crop"])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws3.append([bn,
            1 if d.get("input_mineral_soil") else 0, 1 if d.get("input_blood_meal") else 0,
            1 if d.get("input_greensand") else 0,    1 if d.get("input_feather_meal") else 0,
            1 if d.get("input_poultry_manure") else 0, 1 if d.get("input_livestock_manure") else 0,
            1 if d.get("input_compost_worm") else 0, 1 if d.get("input_biological") else 0,
            1 if d.get("input_regular_compost") else 0, 1 if d.get("input_plasticulture") else 0,
            1 if d.get("input_mulch") else 0,
        ])
    ws4 = wb.create_sheet("4_Energy_Audit")
    ws4.append(["Batch Number","Grow Room / Zone","Lighting Type","Lighting Wattage (W)",
                "Light Hours/Day","Weeks Veg","Weeks Flower","HVAC System Type",
                "HVAC kWh/Month","CO2 Supplementation","CO2 Source","Irrigation System",
                "Est. Lighting kWh/Cycle","Est. Total kWh/Cycle"])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        if d.get("grow_room") or d.get("lighting_type"):
            w = d.get("lighting_wattage_w") or 0
            h = d.get("light_hours_per_day") or 0
            days = ((d.get("weeks_veg") or 0) + (d.get("weeks_flower") or 0)) * 7
            l_kwh = round(w * h * days / 1000, 1)
            t_kwh = round(l_kwh + (d.get("hvac_kwh_month") or 0) * days / 30, 1)
            ws4.append([bn, d.get("grow_room",""), d.get("lighting_type",""),
                d.get("lighting_wattage_w") or "", d.get("light_hours_per_day") or "",
                d.get("weeks_veg") or "", d.get("weeks_flower") or "",
                d.get("hvac_type",""), d.get("hvac_kwh_month") or "",
                d.get("co2_supplementation",""), d.get("co2_source",""),
                d.get("irrigation_system",""), l_kwh, t_kwh])
    ws5 = wb.create_sheet("5_Nutrient_Log")
    ws5.append(["Batch Number","Application Date","Crop Stage","Product / Amendment Name",
                "Type","NPK / Analysis","Rate Applied","Units","Application Method","Notes"])
    for bn, b in batches_dict.items():
        df = b.get("nutrients_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws5.append([bn] + list(row))
    ws6 = wb.create_sheet("6_Pest_Control_Log")
    ws6.append(["Batch Number","Date","Event Type","Target Pest / Disease",
                "Product / Method Name","Type","EPA Reg. # / Active Ingredient",
                "Rate Applied","Application Method","PHI (days)","Applied By","Notes"])
    for bn, b in batches_dict.items():
        df = b.get("pests_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws6.append([bn] + list(row))
    ws7 = wb.create_sheet("7_Yield_Selling")
    ws7.append(["Batch Number","Strain Name","Season / Cycle",
                "Flower Sold (g)","Flower Sale Date","Flower Price ($/g)","Flower Revenue ($)",
                "Pre-Rolls Sold (units)","Pre-Roll Sale Date","Pre-Roll Price ($/unit)","Pre-Roll Revenue ($)",
                "Biomass Sold (g)","Biomass Sale Date","Biomass Price ($/kg)","Biomass Revenue ($)"])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws7.append([bn, d.get("strain_name",""), d.get("season_cycle",""),
            d.get("flower_sold_g") or "", d.get("flower_sale_date",""),
            d.get("flower_price_per_g") or "", d.get("flower_revenue") or "",
            d.get("preroll_sold_units") or "", d.get("preroll_sale_date",""),
            d.get("preroll_price_per_unit") or "", d.get("preroll_revenue") or "",
            d.get("biomass_sold_g") or "", d.get("biomass_sale_date",""),
            d.get("biomass_price_per_kg") or "", d.get("biomass_revenue") or ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def build_postharvest_excel(batches_dict):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "1_Harvest_Batch_Log"
    ws1.append(["Harvest Batch ID","Flower Room / Zone","Harvest Date",
                "Initial Plant Count","Harvested Plant Count","Plant Loss (#)","Plant Loss Notes"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        initial = d.get("initial_plant_count") or 0
        harvested = d.get("harvested_plant_count") or 0
        ws1.append([bid, d.get("flower_room",""), d.get("harvest_date",""),
            initial or "", harvested or "",
            (initial - harvested) if (initial and harvested) else "",
            d.get("plant_loss_notes","")])
    ws2 = wb.create_sheet("2_Processing_Weights")
    ws2.append(["Harvest Batch ID",
                "Total Wet Weight (g)","Harvest Waste (g)",
                "Bucking Date","Bucked Weight (g)","Bucking Waste (g)",
                "Drying Date","Dried Weight (g)","Drying Waste (g)","Dry Time (days)",
                "Trim Date","Trimmed Flower (g)","Trim Weight (g)","Trim Waste (g)","Moisture Loss (%)"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        wet = d.get("total_wet_weight_g") or 0; dry = d.get("dried_weight_g") or 0
        ml = round((1 - dry/wet)*100, 1) if wet and dry else ""
        ws2.append([bid, wet or "", d.get("harvest_waste_g") or "",
            d.get("bucking_date",""), d.get("bucked_weight_g") or "", d.get("bucking_waste_g") or "",
            d.get("drying_date",""), dry or "", d.get("drying_waste_g") or "", d.get("dry_time_days") or "",
            d.get("trim_date",""), d.get("trimmed_flower_g") or "", d.get("trim_weight_g") or "",
            d.get("trim_waste_g") or "", ml])
    ws3 = wb.create_sheet("3_Curing")
    ws3.append(["Harvest Batch ID","Label","Date Into Cure Room","Flower Weight Entry (g)","Trim Weight Entry (g)"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws3.append([bid, d.get("cure_label",""), d.get("cure_entry_date",""),
                    d.get("cure_flower_entry_g") or "", d.get("cure_trim_entry_g") or ""])
    ws3.append([])
    ws3.append(["Harvest Batch ID","Check Date","Days Since Start",
                "Flower Weight (g)","Trim Weight (g)","RH %","Action Taken"])
    for bid, b in batches_dict.items():
        cdf = b.get("curing_log_df")
        if cdf is not None and not cdf.empty:
            for _, row in cdf.iterrows():
                ws3.append([bid] + list(row))
    ws4 = wb.create_sheet("4_Quality_Testing")
    ws4.append(["Harvest Batch ID","Date Submitted","Lab Name","Sample ID / Tracking #",
                "Date Tested","Date Results Received","Result","If Failed — Why?","COA Link"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws4.append([bid, d.get("lab_date_submitted",""), d.get("lab_name",""), d.get("sample_id",""),
                    d.get("date_tested",""), d.get("date_received",""), d.get("test_result",""),
                    d.get("fail_reason",""), d.get("coa_link","")])
    ws5 = wb.create_sheet("5_Packaging_Inventory")
    ws5.append(["Harvest Batch ID","Date Packaged","Flower Available (g)",
                "1g Units","Eighths (3.5g)","Quarters (7g)","Half Oz (14g)","1 Oz (28g)",
                "Pre-Rolls (units)","Other SKU","End of Stock Date"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws5.append([bid, d.get("packaging_date",""), d.get("flower_available_for_packaging_g") or "",
            d.get("pkg_1g") or "", d.get("pkg_eighth") or "", d.get("pkg_quarter") or "",
            d.get("pkg_half") or "", d.get("pkg_oz") or "", d.get("pkg_prerolls") or "",
            d.get("pkg_other",""), d.get("end_stock_date","")])
    ws6 = wb.create_sheet("6_Sales_Tracking")
    ws6.append(["Harvest Batch ID","Trim Sold (g)","Trim Price ($/g)","Trim Revenue ($)",
                "Popcorn Sold (g)","Popcorn Price ($/g)","Other Byproduct",
                "Buyer / Dispensary","Sale Date","Invoice / PO #","Notes"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws6.append([bid, d.get("trim_sold_g") or "", d.get("trim_price_per_g") or "",
            d.get("trim_revenue") or "", d.get("popcorn_sold_g") or "",
            d.get("popcorn_price_per_g") or "", d.get("other_byproduct",""),
            d.get("buyer_name",""), d.get("sale_date",""),
            d.get("invoice_num",""), d.get("sale_notes","")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def build_logs_excel(batches_dict):
    """Nutrient + Pest logs as a 2-sheet Excel."""
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Nutrient_Log"
    ws1.append(["Batch Number","Application Date","Crop Stage","Product / Amendment Name",
                "Type","NPK / Analysis","Rate Applied","Units","Application Method","Notes"])
    for bn, b in batches_dict.items():
        df = b.get("nutrients_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws1.append([bn] + list(row))
    ws2 = wb.create_sheet("Pest_Control_Log")
    ws2.append(["Batch Number","Date","Event Type","Target Pest / Disease",
                "Product / Method Name","Type","EPA Reg. # / Active Ingredient",
                "Rate Applied","Application Method","PHI (days)","Applied By","Notes"])
    for bn, b in batches_dict.items():
        df = b.get("pests_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws2.append([bn] + list(row))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def build_curing_excel(batches_dict):
    """Curing log as Excel."""
    wb = Workbook()
    ws = wb.active; ws.title = "Curing_Log"
    ws.append(["Harvest Batch ID","Check Date (mm/dd/yy)","Days Since Cure Start",
               "Flower Weight (g)","Trim Weight (g)","RH % Inside Container","Action Taken"])
    for bn, b in batches_dict.items():
        cdf = b.get("curing_log_df")
        if cdf is not None and not cdf.empty:
            for _, row in cdf.iterrows():
                ws.append([bn] + list(row))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def _build_combined_report(batch_id, ph, pos):
    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    ph_d = ph.get("data", {}) if ph else {}
    pos_d = pos.get("data", {}) if pos else {}
    ws.append(["Field","Value","Source"])
    ws.append(["Batch Number / ID", batch_id, ""])
    ws.append([])
    ws.append(["--- PRE-HARVEST ---","",""])
    for k, v in ph_d.items():
        ws.append([k, v, "Pre-Harvest"])
    ws.append([])
    ws.append(["--- POST-HARVEST ---","",""])
    for k, v in pos_d.items():
        ws.append([k, v, "Post-Harvest"])
    if ph and ph.get("plants_df") is not None and not ph["plants_df"].empty:
        ws2 = wb.create_sheet("Individual_Plants"); df = ph["plants_df"]
        ws2.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows(): ws2.append([batch_id] + list(row))
    if ph and ph.get("nutrients_df") is not None and not ph["nutrients_df"].empty:
        ws3 = wb.create_sheet("Nutrient_Log"); df = ph["nutrients_df"]
        ws3.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows(): ws3.append([batch_id] + list(row))
    if ph and ph.get("pests_df") is not None and not ph["pests_df"].empty:
        ws4 = wb.create_sheet("Pest_Log"); df = ph["pests_df"]
        ws4.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows(): ws4.append([batch_id] + list(row))
    if pos and pos.get("curing_log_df") is not None and not pos["curing_log_df"].empty:
        ws5 = wb.create_sheet("Curing_Log"); df = pos["curing_log_df"]
        ws5.append(["Harvest Batch ID"] + list(df.columns))
        for _, row in df.iterrows(): ws5.append([batch_id] + list(row))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE HEADER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="disclaimer-box">
<b>Record-keeping tool only.</b> Not a compliance filing system. Always maintain records consistent
with NYS OCM requirements. Data is stored in your browser session — download before closing.
</div>
""", unsafe_allow_html=True)

st.title("🌿 Crop Overview")
st.caption("Integrated pre-harvest, post-harvest, and batch data — NYS cannabis & hemp")
st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD SECTION
# ══════════════════════════════════════════════════════════════════════════════

with st.expander("📤 Upload Existing Excel Templates", expanded=False):
    u_col1, u_col2 = st.columns(2)

    with u_col1:
        st.markdown("**Pre-Harvest Template**")
        if not HAS_OPENPYXL:
            st.error("openpyxl not installed — Excel upload/download disabled.")
        else:
            up_ph = st.file_uploader("PreHarvest_GrowingAndSelling_Template.xlsx",
                                     type=["xlsx"], key="crop_upload_ph")
            if up_ph:
                def _read_section(ws, section_row):
                    header_r = section_row + 1; data_start = section_row + 3; data_end = section_row + 24
                    headers = [ws.cell(header_r, c).value for c in range(1, ws.max_column + 1)]
                    headers = [str(h) if h is not None else f"_col{c}" for c, h in enumerate(headers, 1)]
                    records = []
                    for r in range(data_start, data_end + 1):
                        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
                        if any(v is not None for v in vals):
                            records.append(dict(zip(headers, vals)))
                    return records
                def _read_log_sheet(ws, header_row, data_start_row):
                    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
                    headers = [str(h) if h is not None else f"_col{c}" for c, h in enumerate(headers, 1)]
                    records = []
                    for r in range(data_start_row, ws.max_row + 1):
                        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
                        if any(v is not None for v in vals):
                            records.append(dict(zip(headers, vals)))
                    return records
                with st.spinner("Parsing…"):
                    try:
                        wb = openpyxl.load_workbook(io.BytesIO(up_ph.read()), data_only=True)
                        batches = {}
                        if "1_PreHarvest_Aggregate" in wb.sheetnames:
                            ws = wb["1_PreHarvest_Aggregate"]
                            sections = [_read_section(ws, r) for r in [4, 29, 54, 79, 104, 129]]
                            max_len = max((len(s) for s in sections), default=0)
                            for i in range(max_len):
                                merged = {}
                                for s in sections:
                                    if i < len(s): merged.update(s[i])
                                bn = str(merged.get("Batch Number") or merged.get("Farm Name") or f"Batch_{i+1}")
                                batches.setdefault(bn, {})["aggregate"] = merged
                                batches[bn]["key"] = bn
                        imported = 0
                        for bn, b in batches.items():
                            agg = b.get("aggregate", {})
                            st.session_state.preharvest_batches[bn] = {
                                "data": {
                                    "farm_name": agg.get("Farm Name",""),
                                    "strain_name": agg.get("Strain Name",""),
                                    "seed_source": agg.get("Seed Source",""),
                                    "flowering_strategy": agg.get("Flowering Strategy","Photoperiod"),
                                    "cultivated_area_sqm": agg.get("Cultivated Area (sqM)"),
                                    "growing_media": agg.get("Growing Media",""),
                                    "soil_type": agg.get("Soil Type",""),
                                    "germination_rate": agg.get("Germination Rate (%)"),
                                    "thc_pct": agg.get("THC (%)"),
                                    "cbd_pct": agg.get("CBD (%)"),
                                },
                                "plants_df": None, "nutrients_df": None, "pests_df": None,
                            }
                            if bn not in st.session_state.postharvest_batches:
                                st.session_state.postharvest_batches[bn] = {"data": {}, "curing_log_df": None}
                            imported += 1
                        if imported:
                            st.success(f"Imported {imported} pre-harvest batch(es).")
                            st.rerun()
                        else:
                            st.warning("No data rows found in template.")
                    except Exception as e:
                        st.error(f"Could not parse file: {e}")

    with u_col2:
        st.markdown("**Post-Harvest Template**")
        if HAS_OPENPYXL:
            up_pos = st.file_uploader("PostHarvest_DataCollection_Template.xlsx",
                                      type=["xlsx"], key="crop_upload_pos")
            if up_pos:
                with st.spinner("Parsing…"):
                    try:
                        def _rs(ws, sr):
                            hr = sr+1; ds = sr+3; de = min(sr+24, ws.max_row)
                            hdrs = [ws.cell(hr,c).value for c in range(1, ws.max_column+1)]
                            hdrs = [str(h) if h is not None else f"_col{c}" for c,h in enumerate(hdrs,1)]
                            recs = []
                            for r in range(ds, de+1):
                                vals = [ws.cell(r,c).value for c in range(1,len(hdrs)+1)]
                                if any(v is not None for v in vals): recs.append(dict(zip(hdrs,vals)))
                            return recs
                        wb2 = openpyxl.load_workbook(io.BytesIO(up_pos.read()), data_only=True)
                        imported2 = 0
                        if "1_HarvestBatch_Log" in wb2.sheetnames:
                            ws = wb2["1_HarvestBatch_Log"]
                            for rec in _rs(ws, 4):
                                bid = str(rec.get("Harvest Batch Name / ID",""))
                                if bid and bid not in ("None",""):
                                    st.session_state.postharvest_batches[bid] = {
                                        "data": {
                                            "harvest_date": str(rec.get("Harvest Date (mm/dd/yyyy)","")),
                                            "flower_room": str(rec.get("Flower Room / Zone","")),
                                        },
                                        "curing_log_df": None,
                                    }
                                    if bid not in st.session_state.preharvest_batches:
                                        st.session_state.preharvest_batches[bid] = {
                                            "data": {}, "plants_df": None, "nutrients_df": None, "pests_df": None,
                                        }
                                    imported2 += 1
                        if imported2:
                            st.success(f"Imported {imported2} post-harvest batch(es).")
                            st.rerun()
                        else:
                            st.warning("No data rows found in template.")
                    except Exception as e:
                        st.error(f"Could not parse file: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# BATCH MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("### Batch Management")
st.caption("One batch ID links pre-harvest and post-harvest data. Creating a batch here adds it to both.")

all_batch_ids = sorted(set(
    list(st.session_state.preharvest_batches.keys()) +
    list(st.session_state.postharvest_batches.keys())
))

col_sel, col_new_id, col_btn, col_del = st.columns([3, 2, 1, 1])
with col_sel:
    selected_batch = st.selectbox(
        "Select batch to view / edit",
        all_batch_ids if all_batch_ids else ["(none)"],
        key="crop_selected_batch",
        disabled=not all_batch_ids,
    )
    if not all_batch_ids:
        selected_batch = None

with col_new_id:
    new_bid = st.text_input("New batch ID", placeholder="e.g. BATCH-2025-001",
                            key="crop_new_bid", label_visibility="collapsed")
with col_btn:
    st.write("")
    if st.button("➕ Create", use_container_width=True):
        bid = new_bid.strip()
        if not bid:
            st.error("Enter a batch ID.")
        elif bid in st.session_state.preharvest_batches or bid in st.session_state.postharvest_batches:
            st.error(f"'{bid}' already exists.")
        else:
            st.session_state.preharvest_batches[bid] = {
                "data": {}, "plants_df": None, "nutrients_df": None, "pests_df": None,
            }
            st.session_state.postharvest_batches[bid] = {"data": {}, "curing_log_df": None}
            st.session_state.crop_selected_batch = bid
            st.rerun()

with col_del:
    st.write("")
    if selected_batch and selected_batch != "(none)":
        if st.button("🗑 Delete", use_container_width=True,
                     help="Remove this batch from both pre- and post-harvest"):
            st.session_state.preharvest_batches.pop(selected_batch, None)
            st.session_state.postharvest_batches.pop(selected_batch, None)
            st.rerun()

if not selected_batch or selected_batch == "(none)":
    st.info("Create a new batch or upload an existing template to get started.")
    st.stop()

# Ensure the selected batch exists in both dicts
if selected_batch not in st.session_state.preharvest_batches:
    st.session_state.preharvest_batches[selected_batch] = {
        "data": {}, "plants_df": None, "nutrients_df": None, "pests_df": None,
    }
if selected_batch not in st.session_state.postharvest_batches:
    st.session_state.postharvest_batches[selected_batch] = {"data": {}, "curing_log_df": None}

ph_batch  = st.session_state.preharvest_batches[selected_batch]
pos_batch = st.session_state.postharvest_batches[selected_batch]
ph_bd  = ph_batch.setdefault("data", {})
pos_bd = pos_batch.setdefault("data", {})

st.markdown(
    f"Editing: <span class='batch-badge-ph'>🌱 {selected_batch}</span>",
    unsafe_allow_html=True,
)
st.write("")

# ══════════════════════════════════════════════════════════════════════════════
# OUTER TABS
# ══════════════════════════════════════════════════════════════════════════════

outer_tabs = st.tabs(["🌱 Pre-Harvest", "🌾 Post-Harvest", "📊 Batch Dashboard"])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — PRE-HARVEST
# ─────────────────────────────────────────────────────────────────────────────
with outer_tabs[0]:
    tabs = st.tabs([
        "🌱 Batch & Growing",
        "🌿 Individual Plants",
        "📋 Inputs Survey",
        "⚡ Energy Audit",
        "🧪 Nutrient Log",
        "🐛 Pest Control",
        "💰 Yield & Selling",
    ])

    # ── TAB 1.1: Batch Identity & Growing Environment ──────────────────────
    with tabs[0]:
        st.markdown("#### Batch Identity")
        c1, c2, c3 = st.columns(3)
        with c1:
            ph_bd["farm_name"] = st.text_input("Farm Name", value=ph_bd.get("farm_name",""),
                                               key=f"ph_{selected_batch}_farm")
        with c2:
            ph_bd["strain_name"] = st.text_input("Strain Name", value=ph_bd.get("strain_name",""),
                                                 key=f"ph_{selected_batch}_strain")
        with c3:
            ph_bd["seed_source"] = st.text_input("Seed Source", value=ph_bd.get("seed_source",""),
                                                 placeholder="Breeder, clone source…",
                                                 key=f"ph_{selected_batch}_seed")
        c1, c2 = st.columns(2)
        with c1:
            ph_bd["flowering_strategy"] = st.selectbox(
                "Flowering Strategy", ["Photoperiod","Autoflower"],
                index=0 if ph_bd.get("flowering_strategy","Photoperiod") == "Photoperiod" else 1,
                key=f"ph_{selected_batch}_flowering")
        with c2:
            ph_bd["season_cycle"] = st.text_input("Season / Cycle", value=ph_bd.get("season_cycle",""),
                                                   placeholder="e.g. Spring 2025, Cycle 2",
                                                   key=f"ph_{selected_batch}_season")
        st.divider()
        st.markdown("#### Growing Environment")
        c1, c2, c3 = st.columns(3)
        with c1:
            ph_bd["cultivated_area_sqm"] = st.number_input("Cultivated Area (sqM)", min_value=0.0, step=10.0,
                value=float(ph_bd.get("cultivated_area_sqm") or 0.0), key=f"ph_{selected_batch}_area")
        with c2:
            ph_bd["growing_media"] = st.text_input("Growing Media", value=ph_bd.get("growing_media",""),
                placeholder="Soil, Coco, Hydro…", key=f"ph_{selected_batch}_media")
        with c3:
            ph_bd["soil_type"] = st.text_input("Soil Type", value=ph_bd.get("soil_type",""),
                placeholder="Loam, Sandy loam…", key=f"ph_{selected_batch}_soil_type")
        c1, c2 = st.columns(2)
        with c1:
            aspects = ["—","N","NE","E","SE","S","SW","W","NW","Flat"]
            cur_asp = ph_bd.get("field_aspect","—")
            ph_bd["field_aspect"] = st.selectbox("Field Aspect", aspects,
                index=aspects.index(cur_asp) if cur_asp in aspects else 0,
                key=f"ph_{selected_batch}_aspect")
        with c2:
            steeps = ["—","Flat","Gentle slope","Moderate slope","Steep"]
            cur_st = ph_bd.get("field_steepness","—")
            ph_bd["field_steepness"] = st.selectbox("Field Steepness", steeps,
                index=steeps.index(cur_st) if cur_st in steeps else 0,
                key=f"ph_{selected_batch}_steepness")
        st.markdown("**Plant Count by Environment**")
        pc1, pc2, pc3, pc4 = st.columns(4)
        with pc1:
            ph_bd["n_plants_outdoor"] = st.number_input("# Outdoors", min_value=0, step=1,
                value=int(ph_bd.get("n_plants_outdoor") or 0), key=f"ph_{selected_batch}_nout")
        with pc2:
            ph_bd["n_plants_hoop"] = st.number_input("# Hoop House", min_value=0, step=1,
                value=int(ph_bd.get("n_plants_hoop") or 0), key=f"ph_{selected_batch}_nhoop")
        with pc3:
            ph_bd["n_plants_greenhouse"] = st.number_input("# Greenhouse", min_value=0, step=1,
                value=int(ph_bd.get("n_plants_greenhouse") or 0), key=f"ph_{selected_batch}_ngh")
        with pc4:
            ph_bd["n_plants_indoor"] = st.number_input("# Indoors", min_value=0, step=1,
                value=int(ph_bd.get("n_plants_indoor") or 0), key=f"ph_{selected_batch}_nin")
        total_plants = sum(int(ph_bd.get(k) or 0)
                          for k in ["n_plants_outdoor","n_plants_hoop","n_plants_greenhouse","n_plants_indoor"])
        st.metric("Total Plants This Batch", total_plants)
        st.divider()
        st.markdown("#### Germination & Flowering")
        g1, g2, g3 = st.columns(3)
        with g1:
            ph_bd["germination_rate"] = st.number_input("Germination Rate (%)", 0.0, 100.0, step=0.5,
                value=float(ph_bd.get("germination_rate") or 0.0), key=f"ph_{selected_batch}_germ")
        with g2:
            ph_bd["first_flower_date"] = st.text_input("Approx. First Flower Date (mm/dd/yyyy)",
                value=ph_bd.get("first_flower_date",""), key=f"ph_{selected_batch}_ff_date")
        with g3:
            ph_bd["days_in_ground"] = st.number_input("Days in Ground", min_value=0, step=1,
                value=int(ph_bd.get("days_in_ground") or 0), key=f"ph_{selected_batch}_days")
        st.divider()
        st.markdown("#### Disease")
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            dp_opts = ["No (0)","Yes (1)"]
            ph_bd["disease_presence"] = st.selectbox("Disease Present?", dp_opts,
                index=1 if ph_bd.get("disease_presence") == "Yes (1)" else 0,
                key=f"ph_{selected_batch}_dis_yn")
        with d2:
            ph_bd["disease_name"] = st.text_input("Disease Name", value=ph_bd.get("disease_name",""),
                placeholder="Botrytis, PM, Fusarium…", key=f"ph_{selected_batch}_dis_name")
        with d3:
            sev_opts = ["—","Mild","Moderate","Severe"]
            cur_sev = ph_bd.get("disease_severity","—")
            ph_bd["disease_severity"] = st.selectbox("Severity", sev_opts,
                index=sev_opts.index(cur_sev) if cur_sev in sev_opts else 0,
                key=f"ph_{selected_batch}_dis_sev")
        with d4:
            ph_bd["disease_date"] = st.text_input("Date First Seen (mm/dd/yy)",
                value=ph_bd.get("disease_date",""), key=f"ph_{selected_batch}_dis_date")
        st.divider()
        st.markdown("#### Harvest Quality — Trichome Assessment")
        hq1, hq2, hq3, hq4 = st.columns(4)
        with hq1:
            ph_bd["pct_white_trichomes"] = st.number_input("% White (Cloudy)", 0.0, 100.0, step=1.0,
                value=float(ph_bd.get("pct_white_trichomes") or 0.0), key=f"ph_{selected_batch}_wh_tri")
        with hq2:
            ph_bd["pct_clear_trichomes"] = st.number_input("% Clear", 0.0, 100.0, step=1.0,
                value=float(ph_bd.get("pct_clear_trichomes") or 0.0), key=f"ph_{selected_batch}_cl_tri")
        with hq3:
            ph_bd["pct_amber_trichomes"] = st.number_input("% Amber", 0.0, 100.0, step=1.0,
                value=float(ph_bd.get("pct_amber_trichomes") or 0.0), key=f"ph_{selected_batch}_am_tri")
        with hq4:
            ph_bd["biomass_weight_kg"] = st.number_input("Total Biomass Weight (kg)", 0.0, step=0.1,
                value=float(ph_bd.get("biomass_weight_kg") or 0.0), key=f"ph_{selected_batch}_biomass")
        st.divider()
        st.markdown("#### Lab Results — Cannabinoids & Terpenes")
        st.caption("Enter values from COA. Full COA attachment goes in Post-Harvest → Quality Testing tab.")
        lr1, lr2, lr3 = st.columns(3)
        with lr1:
            ph_bd["thc_pct"] = st.number_input("THC (%)", 0.0, 100.0, step=0.1,
                value=float(ph_bd.get("thc_pct") or 0.0), key=f"ph_{selected_batch}_thc")
        with lr2:
            ph_bd["cbd_pct"] = st.number_input("CBD (%)", 0.0, 100.0, step=0.1,
                value=float(ph_bd.get("cbd_pct") or 0.0), key=f"ph_{selected_batch}_cbd")
        with lr3:
            ph_bd["other_cannabinoids"] = st.text_input("Other Cannabinoids (notes)",
                value=ph_bd.get("other_cannabinoids",""), placeholder="CBG: 0.5%, CBN: 0.2%…",
                key=f"ph_{selected_batch}_other_cann")
        lr4, lr5 = st.columns(2)
        with lr4:
            ph_bd["b_myrcene_pct"] = st.number_input("β-Myrcene (%)", 0.0, 100.0, step=0.001,
                value=float(ph_bd.get("b_myrcene_pct") or 0.0),
                key=f"ph_{selected_batch}_myrcene", format="%.3f")
        with lr5:
            ph_bd["b_caryophyllene_pct"] = st.number_input("β-Caryophyllene (%)", 0.0, 100.0, step=0.001,
                value=float(ph_bd.get("b_caryophyllene_pct") or 0.0),
                key=f"ph_{selected_batch}_caryoph", format="%.3f")

    # ── TAB 1.2: Individual Plants ─────────────────────────────────────────
    with tabs[1]:
        st.markdown("#### Individual Plant Data")
        st.caption("One row per plant. Batch Number is inherited from the selected batch.")
        _plant_schema = {
            "Plant Number": pd.Series(dtype="Int64"),
            "Plant ID / Tag": pd.Series(dtype="str"),
            "Environment": pd.Series(dtype="str"),
            "Growing Media": pd.Series(dtype="str"),
            "Seed Planting Date": pd.Series(dtype="str"),
            "Transplanting Date": pd.Series(dtype="str"),
            "Topped (Y/N)": pd.Series(dtype="str"),
            "Pruned (Y/N)": pd.Series(dtype="str"),
            "Trellised (Y/N)": pd.Series(dtype="str"),
            "Height at Harvest (cm)": pd.Series(dtype="Float64"),
            "Stem Width (mm)": pd.Series(dtype="Float64"),
            "Node Count": pd.Series(dtype="Int64"),
            "Bud Size": pd.Series(dtype="str"),
        }
        existing_pl = ph_batch.get("plants_df")
        default_pl = pd.DataFrame(_plant_schema) if (existing_pl is None or existing_pl.empty) else existing_pl
        edited_pl = st.data_editor(default_pl, num_rows="dynamic", use_container_width=True,
            key=f"plants_editor_{selected_batch}",
            column_config={
                "Plant Number": st.column_config.NumberColumn("Plant #", min_value=1, step=1),
                "Environment": st.column_config.SelectboxColumn("Environment",
                    options=["Outdoors","Hoop House","Greenhouse","Indoors","Aquaponics","Other"]),
                "Topped (Y/N)": st.column_config.SelectboxColumn("Topped", options=["Y","N"]),
                "Pruned (Y/N)": st.column_config.SelectboxColumn("Pruned", options=["Y","N"]),
                "Trellised (Y/N)": st.column_config.SelectboxColumn("Trellised", options=["Y","N"]),
                "Height at Harvest (cm)": st.column_config.NumberColumn("Height (cm)", min_value=0.0, format="%.1f"),
                "Stem Width (mm)": st.column_config.NumberColumn("Stem (mm)", min_value=0.0, format="%.1f"),
                "Node Count": st.column_config.NumberColumn("Nodes", min_value=0, step=1),
                "Bud Size": st.column_config.SelectboxColumn("Bud Size", options=["S","M","L","XL"]),
            })
        st.session_state.preharvest_batches[selected_batch]["plants_df"] = edited_pl
        if not edited_pl.empty:
            st.caption(f"{len(edited_pl)} plant records")

    # ── TAB 1.3: Inputs Survey ─────────────────────────────────────────────
    with tabs[2]:
        st.markdown("#### Inputs Survey — Do You Use?")
        st.caption("Check all inputs used this batch / season.")
        inputs_list = [
            ("Mineral / Conventional Soil Amendments","input_mineral_soil"),
            ("Blood Meal","input_blood_meal"),
            ("Green Sand (potassium / mineral source)","input_greensand"),
            ("Feather Meal (slow-release N)","input_feather_meal"),
            ("Poultry Manure / Chicken Litter","input_poultry_manure"),
            ("Livestock Manure (cattle, horse, other)","input_livestock_manure"),
            ("Compost / Worm Castings (Vermicompost)","input_compost_worm"),
            ("Biological Inoculants (Mycorrhizae, PGPR, etc.)","input_biological"),
            ("Regular Compost (not vermicompost)","input_regular_compost"),
            ("Plasticulture (plastic mulch, row covers)","input_plasticulture"),
            ("Organic Mulch / Cover Crops","input_mulch"),
        ]
        c1, c2, c3 = st.columns(3)
        for idx, (label, key) in enumerate(inputs_list):
            with [c1, c2, c3][idx % 3]:
                ph_bd[key] = st.checkbox(label, value=bool(ph_bd.get(key, False)),
                                         key=f"ph_{selected_batch}_{key}")

    # ── TAB 1.4: Energy Audit ──────────────────────────────────────────────
    with tabs[3]:
        n_indoor = int(ph_bd.get("n_plants_indoor") or 0)
        if n_indoor == 0:
            st.info("No indoor plants entered. Energy Audit applies to indoor grows only. "
                    "Update plant count in **Batch & Growing** tab if needed.")
        st.markdown("#### Energy Audit — Indoor Operations")
        st.caption("Complete for each indoor grow zone. Leave blank if fully outdoor or hoop house.")
        ea1, ea2, ea3 = st.columns(3)
        with ea1:
            ph_bd["grow_room"] = st.text_input("Grow Room / Zone", value=ph_bd.get("grow_room",""),
                placeholder="e.g. Veg Room A", key=f"ph_{selected_batch}_grow_room")
        with ea2:
            lt_opts = ["—","LED","HPS","CMH","Fluorescent","Other"]
            cur_lt = ph_bd.get("lighting_type","—")
            ph_bd["lighting_type"] = st.selectbox("Lighting Type", lt_opts,
                index=lt_opts.index(cur_lt) if cur_lt in lt_opts else 0,
                key=f"ph_{selected_batch}_light_type")
        with ea3:
            ph_bd["lighting_wattage_w"] = st.number_input("Lighting Wattage (W)", min_value=0.0, step=100.0,
                value=float(ph_bd.get("lighting_wattage_w") or 0.0), key=f"ph_{selected_batch}_wattage")
        ea4, ea5, ea6 = st.columns(3)
        with ea4:
            ph_bd["light_hours_per_day"] = st.number_input("Light Hours / Day", 0.0, 24.0, step=0.5,
                value=float(ph_bd.get("light_hours_per_day") or 18.0), key=f"ph_{selected_batch}_l_hrs")
        with ea5:
            ph_bd["weeks_veg"] = st.number_input("Weeks in Veg", min_value=0, step=1,
                value=int(ph_bd.get("weeks_veg") or 0), key=f"ph_{selected_batch}_wks_veg")
        with ea6:
            ph_bd["weeks_flower"] = st.number_input("Weeks in Flower", min_value=0, step=1,
                value=int(ph_bd.get("weeks_flower") or 0), key=f"ph_{selected_batch}_wks_fl")
        ea7, ea8 = st.columns(2)
        with ea7:
            ph_bd["hvac_type"] = st.text_input("HVAC System Type", value=ph_bd.get("hvac_type",""),
                placeholder="Mini-split, chiller, swamp cooler…", key=f"ph_{selected_batch}_hvac_type")
        with ea8:
            ph_bd["hvac_kwh_month"] = st.number_input("HVAC Est. kWh / Month", min_value=0.0, step=10.0,
                value=float(ph_bd.get("hvac_kwh_month") or 0.0), key=f"ph_{selected_batch}_hvac_kwh")
        ea9, ea10, ea11 = st.columns(3)
        with ea9:
            co2_opts = ["No","Yes"]
            ph_bd["co2_supplementation"] = st.selectbox("CO2 Supplementation", co2_opts,
                index=1 if ph_bd.get("co2_supplementation") == "Yes" else 0,
                key=f"ph_{selected_batch}_co2")
        with ea10:
            ph_bd["co2_source"] = st.text_input("CO2 Source", value=ph_bd.get("co2_source",""),
                placeholder="Tank, generator, fermentation…",
                key=f"ph_{selected_batch}_co2_src",
                disabled=(ph_bd.get("co2_supplementation") != "Yes"))
        with ea11:
            irr_opts = ["—","Drip","Hand water","Flood / drain","Ebb and flow","Aeroponics","Other"]
            cur_irr = ph_bd.get("irrigation_system","—")
            ph_bd["irrigation_system"] = st.selectbox("Irrigation System", irr_opts,
                index=irr_opts.index(cur_irr) if cur_irr in irr_opts else 0,
                key=f"ph_{selected_batch}_irr")
        w = ph_bd.get("lighting_wattage_w") or 0
        h = ph_bd.get("light_hours_per_day") or 0
        days = ((ph_bd.get("weeks_veg") or 0) + (ph_bd.get("weeks_flower") or 0)) * 7
        l_kwh = round(w * h * days / 1000, 1) if days > 0 else 0
        hvac_kwh = round((ph_bd.get("hvac_kwh_month") or 0) * days / 30, 1) if days > 0 else 0
        if l_kwh > 0 or hvac_kwh > 0:
            st.divider()
            ek1, ek2, ek3 = st.columns(3)
            ek1.metric("Lighting kWh / Cycle", f"{l_kwh:,.0f}")
            ek2.metric("HVAC kWh / Cycle", f"{hvac_kwh:,.0f}")
            ek3.metric("Total Est. kWh / Cycle", f"{l_kwh + hvac_kwh:,.0f}")

    # ── TAB 1.5: Nutrient Log ──────────────────────────────────────────────
    with tabs[4]:
        st.markdown("#### Nutrient Amendment Log")
        st.caption("One row per application event.")
        _nut_schema = {
            "Application Date (mm/dd/yy)": pd.Series(dtype="str"),
            "Crop Stage": pd.Series(dtype="str"),
            "Product / Amendment Name": pd.Series(dtype="str"),
            "Type": pd.Series(dtype="str"),
            "NPK / Analysis": pd.Series(dtype="str"),
            "Rate Applied": pd.Series(dtype="Float64"),
            "Units": pd.Series(dtype="str"),
            "Application Method": pd.Series(dtype="str"),
            "Notes": pd.Series(dtype="str"),
        }
        existing_n = ph_batch.get("nutrients_df")
        default_n = pd.DataFrame(_nut_schema) if (existing_n is None or existing_n.empty) else existing_n
        edited_n = st.data_editor(default_n, num_rows="dynamic", use_container_width=True,
            key=f"nut_editor_{selected_batch}",
            column_config={
                "Crop Stage": st.column_config.SelectboxColumn("Crop Stage",
                    options=["Seedling","Veg","Early Flower","Mid Flower","Late Flower","Flush","Other"]),
                "Type": st.column_config.SelectboxColumn("Type",
                    options=["Organic","Synthetic","Biological","pH Amendment","Other"]),
                "Application Method": st.column_config.SelectboxColumn("Method",
                    options=["Top dress","Drench","Foliar","Fertigation","Broadcast","Other"]),
                "Units": st.column_config.SelectboxColumn("Units",
                    options=["g/plant","kg/acre","mL/L","oz/gallon","lbs/acre","Other"]),
                "Rate Applied": st.column_config.NumberColumn("Rate", min_value=0.0, format="%.2f"),
            })
        st.session_state.preharvest_batches[selected_batch]["nutrients_df"] = edited_n
        if not edited_n.empty:
            st.caption(f"{len(edited_n)} amendment event(s) recorded")

    # ── TAB 1.6: Pest Control ──────────────────────────────────────────────
    with tabs[5]:
        st.markdown("#### Pest Control Application Log")
        st.caption("One row per scouting or application event.")
        _pest_schema = {
            "Date (mm/dd/yy)": pd.Series(dtype="str"),
            "Event Type": pd.Series(dtype="str"),
            "Target Pest / Disease": pd.Series(dtype="str"),
            "Product / Method Name": pd.Series(dtype="str"),
            "Type": pd.Series(dtype="str"),
            "EPA Reg. # / Active Ingredient": pd.Series(dtype="str"),
            "Rate Applied": pd.Series(dtype="str"),
            "Application Method": pd.Series(dtype="str"),
            "PHI (days)": pd.Series(dtype="Float64"),
            "Applied By": pd.Series(dtype="str"),
            "Notes": pd.Series(dtype="str"),
        }
        existing_p = ph_batch.get("pests_df")
        default_p = pd.DataFrame(_pest_schema) if (existing_p is None or existing_p.empty) else existing_p
        edited_p = st.data_editor(default_p, num_rows="dynamic", use_container_width=True,
            key=f"pest_editor_{selected_batch}",
            column_config={
                "Event Type": st.column_config.SelectboxColumn("Event Type",
                    options=["Scouting","Preventative Application","Curative Application"]),
                "Type": st.column_config.SelectboxColumn("Type",
                    options=["Pesticide (OMRI listed)","Pesticide (synthetic)",
                             "Biocontrol","Physical/Mechanical","None (scouting)"]),
                "Application Method": st.column_config.SelectboxColumn("Method",
                    options=["Foliar spray","Soil drench","Biocontrol release","Trap placement","Other"]),
                "PHI (days)": st.column_config.NumberColumn("PHI (days)", min_value=0, step=1),
            })
        st.session_state.preharvest_batches[selected_batch]["pests_df"] = edited_p
        if not edited_p.empty:
            st.caption(f"{len(edited_p)} event(s) recorded")

    # ── TAB 1.7: Yield & Selling ───────────────────────────────────────────
    with tabs[6]:
        st.markdown("#### Yield & Selling Information")
        st.caption("Processing weights (bucked, dried, trimmed) go in Post-Harvest.")
        y1, y2 = st.columns(2)
        with y1:
            st.markdown("**Flower**")
            ph_bd["flower_sold_g"] = st.number_input("Total Sold (g)", min_value=0.0, step=1.0,
                value=float(ph_bd.get("flower_sold_g") or 0.0), key=f"ph_{selected_batch}_fl_sold")
            ph_bd["flower_price_per_g"] = st.number_input("Price ($/g)", min_value=0.0, step=0.01,
                value=float(ph_bd.get("flower_price_per_g") or 0.0),
                key=f"ph_{selected_batch}_fl_price", format="%.2f")
            ph_bd["flower_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
                value=ph_bd.get("flower_sale_date",""), key=f"ph_{selected_batch}_fl_date")
            fl_rev = round((ph_bd.get("flower_sold_g") or 0) * (ph_bd.get("flower_price_per_g") or 0), 2)
            ph_bd["flower_revenue"] = fl_rev
            if fl_rev > 0: st.metric("Flower Revenue", f"${fl_rev:,.2f}")
        with y2:
            st.markdown("**Pre-Rolls**")
            ph_bd["preroll_sold_units"] = st.number_input("Total Sold (units)", min_value=0, step=1,
                value=int(ph_bd.get("preroll_sold_units") or 0), key=f"ph_{selected_batch}_pr_sold")
            ph_bd["preroll_price_per_unit"] = st.number_input("Price ($/unit)", min_value=0.0, step=0.01,
                value=float(ph_bd.get("preroll_price_per_unit") or 0.0),
                key=f"ph_{selected_batch}_pr_price", format="%.2f")
            ph_bd["preroll_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
                value=ph_bd.get("preroll_sale_date",""), key=f"ph_{selected_batch}_pr_date")
            pr_rev = round((ph_bd.get("preroll_sold_units") or 0) * (ph_bd.get("preroll_price_per_unit") or 0), 2)
            ph_bd["preroll_revenue"] = pr_rev
            if pr_rev > 0: st.metric("Pre-Roll Revenue", f"${pr_rev:,.2f}")
        st.divider()
        st.markdown("**Biomass**")
        bm1, bm2, bm3 = st.columns(3)
        with bm1:
            ph_bd["biomass_sold_g"] = st.number_input("Total Sold (g)", min_value=0.0, step=10.0,
                value=float(ph_bd.get("biomass_sold_g") or 0.0), key=f"ph_{selected_batch}_bm_sold")
        with bm2:
            ph_bd["biomass_price_per_kg"] = st.number_input("Price ($/kg)", min_value=0.0, step=1.0,
                value=float(ph_bd.get("biomass_price_per_kg") or 0.0), key=f"ph_{selected_batch}_bm_price")
        with bm3:
            ph_bd["biomass_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
                value=ph_bd.get("biomass_sale_date",""), key=f"ph_{selected_batch}_bm_date")
        bm_rev = round((ph_bd.get("biomass_sold_g") or 0) / 1000 * (ph_bd.get("biomass_price_per_kg") or 0), 2)
        ph_bd["biomass_revenue"] = bm_rev
        total_rev = (ph_bd.get("flower_revenue") or 0) + (ph_bd.get("preroll_revenue") or 0) + bm_rev
        if total_rev > 0:
            st.divider()
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Flower", f"${ph_bd.get('flower_revenue',0):,.2f}")
            r2.metric("Pre-Rolls", f"${ph_bd.get('preroll_revenue',0):,.2f}")
            r3.metric("Biomass", f"${bm_rev:,.2f}")
            r4.metric("Total Revenue", f"${total_rev:,.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — POST-HARVEST
# ─────────────────────────────────────────────────────────────────────────────
with outer_tabs[1]:
    pos_tabs = st.tabs([
        "📦 Harvest Log",
        "⚖️ Processing Weights",
        "🫙 Curing",
        "🔬 Quality Testing",
        "📦 Packaging",
        "💵 Sales Tracking",
    ])

    # ── TAB 2.1: Harvest Batch Log ─────────────────────────────────────────
    with pos_tabs[0]:
        st.markdown("#### Harvest Batch Log")
        c1, c2 = st.columns(2)
        with c1:
            pos_bd["flower_room"] = st.text_input("Flower Room / Zone", value=pos_bd.get("flower_room",""),
                placeholder="e.g. Greenhouse A, Field 2", key=f"pos_{selected_batch}_room")
            pos_bd["harvest_date"] = st.text_input("Harvest Date (mm/dd/yyyy)",
                value=pos_bd.get("harvest_date",""), key=f"pos_{selected_batch}_hdate")
        with c2:
            pos_bd["initial_plant_count"] = st.number_input("Initial Plant Count", min_value=0, step=1,
                value=int(pos_bd.get("initial_plant_count") or 0), key=f"pos_{selected_batch}_init_cnt")
            pos_bd["harvested_plant_count"] = st.number_input("Harvested Plant Count", min_value=0, step=1,
                value=int(pos_bd.get("harvested_plant_count") or 0), key=f"pos_{selected_batch}_harv_cnt")
        initial = int(pos_bd.get("initial_plant_count") or 0)
        harvested = int(pos_bd.get("harvested_plant_count") or 0)
        plant_loss = initial - harvested if initial >= harvested else 0
        if initial > 0:
            pc1, pc2, pc3 = st.columns(3)
            pc1.metric("Initial Count", initial)
            pc2.metric("Harvested", harvested)
            pc3.metric("Plant Loss", plant_loss,
                       delta=f"{-plant_loss}" if plant_loss > 0 else "None",
                       delta_color="inverse" if plant_loss > 0 else "off")
        pos_bd["plant_loss_notes"] = st.text_area("Plant Loss Notes", value=pos_bd.get("plant_loss_notes",""),
            placeholder="Reason for any plant loss (disease, hermie, compliance, etc.)",
            key=f"pos_{selected_batch}_loss_notes", height=80)

    # ── TAB 2.2: Processing Weights ────────────────────────────────────────
    with pos_tabs[1]:
        st.markdown("#### Processing Weights")
        st.caption("Track weight at every step. All weights in grams (g).")
        st.markdown("**Harvest Weights**")
        hw1, hw2 = st.columns(2)
        with hw1:
            pos_bd["total_wet_weight_g"] = st.number_input("Total Wet Weight (g)", min_value=0.0, step=10.0,
                value=float(pos_bd.get("total_wet_weight_g") or 0.0), key=f"pos_{selected_batch}_wet_wt")
        with hw2:
            pos_bd["harvest_waste_g"] = st.number_input("Waste From Harvest (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("harvest_waste_g") or 0.0), key=f"pos_{selected_batch}_harv_waste")
        st.divider(); st.markdown("**Bucking**")
        bk1, bk2, bk3 = st.columns(3)
        with bk1:
            pos_bd["bucking_date"] = st.text_input("Bucking Date (mm/dd/yyyy)",
                value=pos_bd.get("bucking_date",""), key=f"pos_{selected_batch}_buck_date")
        with bk2:
            pos_bd["bucked_weight_g"] = st.number_input("Bucked Weight (g)", min_value=0.0, step=10.0,
                value=float(pos_bd.get("bucked_weight_g") or 0.0), key=f"pos_{selected_batch}_buck_wt")
        with bk3:
            pos_bd["bucking_waste_g"] = st.number_input("Waste from Bucking (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("bucking_waste_g") or 0.0), key=f"pos_{selected_batch}_buck_waste")
        st.divider(); st.markdown("**Drying**")
        dr1, dr2, dr3, dr4 = st.columns(4)
        with dr1:
            pos_bd["drying_date"] = st.text_input("Drying Date (mm/dd/yyyy)",
                value=pos_bd.get("drying_date",""), key=f"pos_{selected_batch}_dry_date")
        with dr2:
            pos_bd["dried_weight_g"] = st.number_input("Dried Weight (g)", min_value=0.0, step=10.0,
                value=float(pos_bd.get("dried_weight_g") or 0.0), key=f"pos_{selected_batch}_dry_wt")
        with dr3:
            pos_bd["drying_waste_g"] = st.number_input("Waste from Drying (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("drying_waste_g") or 0.0), key=f"pos_{selected_batch}_dry_waste")
        with dr4:
            pos_bd["dry_time_days"] = st.number_input("Dry Time (days)", min_value=0, step=1,
                value=int(pos_bd.get("dry_time_days") or 0), key=f"pos_{selected_batch}_dry_days")
        st.divider(); st.markdown("**Trimming**")
        tr1, tr2, tr3, tr4 = st.columns(4)
        with tr1:
            pos_bd["trim_date"] = st.text_input("Trim Date (mm/dd/yyyy)",
                value=pos_bd.get("trim_date",""), key=f"pos_{selected_batch}_trim_date")
        with tr2:
            pos_bd["trimmed_flower_g"] = st.number_input("Trimmed Flower (g)", min_value=0.0, step=10.0,
                value=float(pos_bd.get("trimmed_flower_g") or 0.0), key=f"pos_{selected_batch}_trim_fl")
        with tr3:
            pos_bd["trim_weight_g"] = st.number_input("Trim Weight (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("trim_weight_g") or 0.0), key=f"pos_{selected_batch}_trim_wt")
        with tr4:
            pos_bd["trim_waste_g"] = st.number_input("Trim Waste (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("trim_waste_g") or 0.0), key=f"pos_{selected_batch}_trim_waste")
        wet = pos_bd.get("total_wet_weight_g") or 0
        dried = pos_bd.get("dried_weight_g") or 0
        trimmed = pos_bd.get("trimmed_flower_g") or 0
        trim_byproduct = pos_bd.get("trim_weight_g") or 0
        if wet > 0:
            st.divider(); st.markdown("**Weight Flow Summary**")
            wf1, wf2, wf3, wf4 = st.columns(4)
            wf1.metric("Wet Weight", f"{wet:,.0f} g")
            wf2.metric("Dried Weight", f"{dried:,.0f} g",
                       delta=f"{((dried/wet)-1)*100:.1f}% vs wet" if wet else None, delta_color="off")
            wf3.metric("Trimmed Flower", f"{trimmed:,.0f} g")
            wf4.metric("Trim / Byproduct", f"{trim_byproduct:,.0f} g")
            if wet and dried:
                moisture_loss = round((1 - dried / wet) * 100, 1)
                st.caption(f"Moisture loss: **{moisture_loss}%** | Dry/wet ratio: **{dried/wet:.3f}**"
                           + (f" | Flower recovery from dry: **{trimmed/dried*100:.1f}%**" if dried else ""))

    # ── TAB 2.3: Curing ────────────────────────────────────────────────────
    with pos_tabs[2]:
        st.markdown("#### Curing")
        st.markdown("**Cure Room Entry**")
        ce1, ce2 = st.columns(2)
        with ce1:
            pos_bd["cure_label"] = st.text_input("Strain / Batch Label on Container",
                value=pos_bd.get("cure_label",""), key=f"pos_{selected_batch}_cure_label")
            pos_bd["cure_entry_date"] = st.text_input("Date Brought Into Cure Room (mm/dd/yyyy)",
                value=pos_bd.get("cure_entry_date",""), key=f"pos_{selected_batch}_cure_date")
        with ce2:
            pos_bd["cure_flower_entry_g"] = st.number_input("Flower Weight at Cure Entry (g)",
                min_value=0.0, step=1.0,
                value=float(pos_bd.get("cure_flower_entry_g") or 0.0), key=f"pos_{selected_batch}_cure_fl")
            pos_bd["cure_trim_entry_g"] = st.number_input("Trim Weight at Cure Entry (g)",
                min_value=0.0, step=1.0,
                value=float(pos_bd.get("cure_trim_entry_g") or 0.0), key=f"pos_{selected_batch}_cure_tr")
        st.divider(); st.markdown("**Cure Monitoring Log**")
        st.caption("Add a row each time you check or burp the cure containers.")
        _cure_schema = {
            "Check Date (mm/dd/yy)": pd.Series(dtype="str"),
            "Days Since Cure Start": pd.Series(dtype="Int64"),
            "Flower Weight (g)": pd.Series(dtype="Float64"),
            "Trim Weight (g)": pd.Series(dtype="Float64"),
            "RH % Inside Container": pd.Series(dtype="Float64"),
            "Action Taken": pd.Series(dtype="str"),
        }
        existing_c = pos_batch.get("curing_log_df")
        default_c = pd.DataFrame(_cure_schema) if (existing_c is None or existing_c.empty) else existing_c
        edited_c = st.data_editor(default_c, num_rows="dynamic", use_container_width=True,
            key=f"cure_editor_{selected_batch}",
            column_config={
                "Days Since Cure Start": st.column_config.NumberColumn("Days", min_value=0, step=1),
                "Flower Weight (g)": st.column_config.NumberColumn("Flower (g)", min_value=0.0, format="%.1f"),
                "Trim Weight (g)": st.column_config.NumberColumn("Trim (g)", min_value=0.0, format="%.1f"),
                "RH % Inside Container": st.column_config.NumberColumn("RH %", min_value=0.0, max_value=100.0, format="%.1f"),
                "Action Taken": st.column_config.SelectboxColumn("Action",
                    options=["Burp","Seal","Note only","Removed from cure","Other"]),
            })
        st.session_state.postharvest_batches[selected_batch]["curing_log_df"] = edited_c
        if not edited_c.empty:
            st.caption(f"{len(edited_c)} cure check(s) logged")

    # ── TAB 2.4: Quality Testing ───────────────────────────────────────────
    with pos_tabs[3]:
        st.markdown("#### Quality Testing — Certificate of Analysis (COA)")
        st.markdown("**Lab Submission**")
        qt1, qt2 = st.columns(2)
        with qt1:
            pos_bd["lab_date_submitted"] = st.text_input("Date Submitted to Lab (mm/dd/yyyy)",
                value=pos_bd.get("lab_date_submitted",""), key=f"pos_{selected_batch}_lab_sub")
            pos_bd["lab_name"] = st.text_input("Lab Name", value=pos_bd.get("lab_name",""),
                key=f"pos_{selected_batch}_lab_name")
        with qt2:
            pos_bd["sample_id"] = st.text_input("Sample ID / Tracking #", value=pos_bd.get("sample_id",""),
                key=f"pos_{selected_batch}_sample_id")
        st.divider(); st.markdown("**Test Results**")
        st.caption("THC% and CBD% go in Pre-Harvest → Lab Results tab.")
        tr1, tr2 = st.columns(2)
        with tr1:
            pos_bd["date_tested"] = st.text_input("Date Tested (mm/dd/yyyy)",
                value=pos_bd.get("date_tested",""), key=f"pos_{selected_batch}_dt_tested")
            pos_bd["date_received"] = st.text_input("Date Results Received (mm/dd/yyyy)",
                value=pos_bd.get("date_received",""), key=f"pos_{selected_batch}_dt_received")
        with tr2:
            result_opts = ["—","Pass","Fail","Retest Pass","Retest Fail","Pending"]
            cur_res = pos_bd.get("test_result","—")
            pos_bd["test_result"] = st.selectbox("Result", result_opts,
                index=result_opts.index(cur_res) if cur_res in result_opts else 0,
                key=f"pos_{selected_batch}_result")
        if pos_bd.get("test_result") in ("Fail","Retest Fail"):
            pos_bd["fail_reason"] = st.text_area("If Failed — Why?",
                value=pos_bd.get("fail_reason",""),
                placeholder="e.g. Pesticides detected, mold count exceeded limit…",
                key=f"pos_{selected_batch}_fail_reason", height=80)
        else:
            pos_bd["fail_reason"] = pos_bd.get("fail_reason","")
        pos_bd["coa_link"] = st.text_input("COA File Path or Link",
            value=pos_bd.get("coa_link",""),
            placeholder="e.g. /files/coa_batch001.pdf",
            key=f"pos_{selected_batch}_coa")
        result = pos_bd.get("test_result","—")
        if result in ("Pass","Retest Pass"):
            st.success(f"✅ COA Result: **{result}**")
        elif result in ("Fail","Retest Fail"):
            st.error(f"❌ COA Result: **{result}** — {pos_bd.get('fail_reason','')}")
        elif result == "Pending":
            st.warning("⏳ Results pending")

    # ── TAB 2.5: Packaging ─────────────────────────────────────────────────
    with pos_tabs[4]:
        st.markdown("#### Packaging & Inventory")
        pk1, pk2 = st.columns(2)
        with pk1:
            pos_bd["packaging_date"] = st.text_input("Date Entered Into Packaging (mm/dd/yyyy)",
                value=pos_bd.get("packaging_date",""), key=f"pos_{selected_batch}_pkg_date")
        with pk2:
            pos_bd["flower_available_for_packaging_g"] = st.number_input(
                "Total Flower Available for Packaging (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("flower_available_for_packaging_g") or 0.0),
                key=f"pos_{selected_batch}_fl_avail")
        st.divider(); st.markdown("**Packaged Units by SKU**")
        sk1, sk2, sk3 = st.columns(3)
        with sk1:
            pos_bd["pkg_1g"] = st.number_input("1 Gram (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_1g") or 0), key=f"pos_{selected_batch}_pkg_1g")
            pos_bd["pkg_eighth"] = st.number_input("Eighths / 3.5g (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_eighth") or 0), key=f"pos_{selected_batch}_pkg_8th")
        with sk2:
            pos_bd["pkg_quarter"] = st.number_input("Quarters / 7g (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_quarter") or 0), key=f"pos_{selected_batch}_pkg_qtr")
            pos_bd["pkg_half"] = st.number_input("Half Oz / 14g (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_half") or 0), key=f"pos_{selected_batch}_pkg_half")
        with sk3:
            pos_bd["pkg_oz"] = st.number_input("1 Oz / 28g (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_oz") or 0), key=f"pos_{selected_batch}_pkg_oz")
            pos_bd["pkg_prerolls"] = st.number_input("Pre-Rolls (units)", min_value=0, step=1,
                value=int(pos_bd.get("pkg_prerolls") or 0), key=f"pos_{selected_batch}_pkg_pr")
        pos_bd["pkg_other"] = st.text_input("Other SKU — describe",
            value=pos_bd.get("pkg_other",""), key=f"pos_{selected_batch}_pkg_other")
        pos_bd["end_stock_date"] = st.text_input("End of Stock Date (mm/dd/yyyy)",
            value=pos_bd.get("end_stock_date",""), key=f"pos_{selected_batch}_end_stock")
        total_packaged_g = (
            (pos_bd.get("pkg_1g") or 0) * 1 +
            (pos_bd.get("pkg_eighth") or 0) * 3.5 +
            (pos_bd.get("pkg_quarter") or 0) * 7 +
            (pos_bd.get("pkg_half") or 0) * 14 +
            (pos_bd.get("pkg_oz") or 0) * 28
        )
        available = pos_bd.get("flower_available_for_packaging_g") or 0
        if total_packaged_g > 0:
            st.divider()
            pg1, pg2, pg3 = st.columns(3)
            pg1.metric("Total Packaged (g)", f"{total_packaged_g:,.1f}")
            if available:
                pg2.metric("Available (g)", f"{available:,.1f}")
                pg3.metric("Remaining / Variance (g)", f"{available - total_packaged_g:+,.1f}",
                           delta_color="normal" if available - total_packaged_g >= 0 else "inverse")

    # ── TAB 2.6: Sales Tracking ────────────────────────────────────────────
    with pos_tabs[5]:
        st.markdown("#### Sales Tracking — Trim & Byproducts")
        st.caption("Flower, pre-roll, and biomass sales go in Pre-Harvest → Yield & Selling.")
        s1, s2 = st.columns(2)
        with s1:
            pos_bd["trim_sold_g"] = st.number_input("Trim Total Sold (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("trim_sold_g") or 0.0), key=f"pos_{selected_batch}_trim_sold")
            pos_bd["trim_price_per_g"] = st.number_input("Trim Price ($/g)", min_value=0.0, step=0.01,
                value=float(pos_bd.get("trim_price_per_g") or 0.0),
                key=f"pos_{selected_batch}_trim_price", format="%.3f")
            trim_rev = round((pos_bd.get("trim_sold_g") or 0) * (pos_bd.get("trim_price_per_g") or 0), 2)
            pos_bd["trim_revenue"] = trim_rev
            if trim_rev > 0: st.metric("Trim Revenue", f"${trim_rev:,.2f}")
        with s2:
            pos_bd["popcorn_sold_g"] = st.number_input("Popcorn / Smalls Sold (g)", min_value=0.0, step=1.0,
                value=float(pos_bd.get("popcorn_sold_g") or 0.0), key=f"pos_{selected_batch}_pop_sold")
            pos_bd["popcorn_price_per_g"] = st.number_input("Popcorn / Smalls Price ($/g)",
                min_value=0.0, step=0.01,
                value=float(pos_bd.get("popcorn_price_per_g") or 0.0),
                key=f"pos_{selected_batch}_pop_price", format="%.3f")
            pop_rev = round((pos_bd.get("popcorn_sold_g") or 0) * (pos_bd.get("popcorn_price_per_g") or 0), 2)
            if pop_rev > 0: st.metric("Popcorn Revenue", f"${pop_rev:,.2f}")
        pos_bd["other_byproduct"] = st.text_input("Other Byproduct — describe",
            value=pos_bd.get("other_byproduct",""),
            placeholder="e.g. extract input material, 500g @ $2.00/g",
            key=f"pos_{selected_batch}_other_byp")
        st.divider(); st.markdown("**Buyer / Channel (optional)**")
        b1, b2, b3 = st.columns(3)
        with b1:
            pos_bd["buyer_name"] = st.text_input("Buyer / Dispensary Name",
                value=pos_bd.get("buyer_name",""), key=f"pos_{selected_batch}_buyer")
        with b2:
            pos_bd["sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
                value=pos_bd.get("sale_date",""), key=f"pos_{selected_batch}_sale_date")
        with b3:
            pos_bd["invoice_num"] = st.text_input("Invoice / PO #",
                value=pos_bd.get("invoice_num",""), key=f"pos_{selected_batch}_invoice")
        pos_bd["sale_notes"] = st.text_area("Notes", value=pos_bd.get("sale_notes",""),
            key=f"pos_{selected_batch}_sale_notes", height=60)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — BATCH DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
with outer_tabs[2]:
    ph_batches  = st.session_state.preharvest_batches
    pos_batches = st.session_state.postharvest_batches
    all_ids = sorted(set(list(ph_batches.keys()) + list(pos_batches.keys())))

    def _has_data(d, *keys):
        return any(d.get(k) not in (None, "", 0, 0.0) for k in keys)

    def _completeness(label, d, required_keys, optional_keys=None):
        total = len(required_keys) + len(optional_keys or [])
        filled = sum(1 for k in (required_keys + (optional_keys or [])) if d.get(k) not in (None, "", 0))
        pct = round(filled / total * 100) if total else 0
        status = "complete" if pct == 100 else ("missing" if pct == 0 else "partial")
        return label, status, pct

    col_sel, col_all = st.columns([3, 1])
    with col_sel:
        dash_selected = st.selectbox("Select Batch", all_ids, key="overview_batch",
                                     index=all_ids.index(selected_batch) if selected_batch in all_ids else 0)
    with col_all:
        st.write("")
        show_all = st.checkbox("Show all batches summary", value=False)

    ph  = ph_batches.get(dash_selected)
    pos = pos_batches.get(dash_selected)
    ph_d  = ph.get("data", {}) if ph else {}
    pos_d = pos.get("data", {}) if pos else {}

    # Completeness
    st.markdown("#### Data Completeness")
    checks = []
    if ph:
        checks += [
            _completeness("Batch Identity", ph_d, ["farm_name","strain_name","flowering_strategy"], ["seed_source","season_cycle"]),
            _completeness("Growing Env.", ph_d, ["cultivated_area_sqm"], ["growing_media","n_plants_outdoor","n_plants_greenhouse"]),
            _completeness("Lab Results", ph_d, ["thc_pct","cbd_pct"], ["b_myrcene_pct"]),
            _completeness("Yield / Selling", ph_d, [], ["flower_sold_g","flower_price_per_g"]),
        ]
    else:
        checks.append(("Pre-Harvest", "missing", 0))
    if pos:
        checks += [
            _completeness("Harvest Log", pos_d, ["harvest_date","harvested_plant_count"], ["flower_room"]),
            _completeness("Processing Wts", pos_d, ["total_wet_weight_g","dried_weight_g"], ["trimmed_flower_g"]),
            _completeness("Quality / COA", pos_d, ["test_result"], ["lab_name","coa_link"]),
            _completeness("Packaging", pos_d, [], ["flower_available_for_packaging_g","pkg_1g","pkg_eighth"]),
        ]
    else:
        checks.append(("Post-Harvest", "missing", 0))

    icon_map = {"complete": "✅", "partial": "⚠️", "missing": "❌"}
    ch_cols = st.columns(len(checks))
    for col, (label, status, pct) in zip(ch_cols, checks):
        with col:
            st.markdown(f"**{label}**")
            st.markdown(f"{icon_map[status]} {pct}%", help=f"Status: {status} ({pct}% of fields filled)")
    st.divider()

    # Key metrics
    st.markdown("#### Key Metrics")
    km_cols = st.columns(4)
    with km_cols[0]:
        st.markdown('<p class="section-header">Strain & Grow</p>', unsafe_allow_html=True)
        st.markdown(f"**Strain:** {ph_d.get('strain_name') or '—'}")
        st.markdown(f"**Farm:** {ph_d.get('farm_name') or '—'}")
        st.markdown(f"**Type:** {ph_d.get('flowering_strategy') or '—'}")
        st.markdown(f"**Season:** {ph_d.get('season_cycle') or '—'}")
    with km_cols[1]:
        st.markdown('<p class="section-header">Plants & Environment</p>', unsafe_allow_html=True)
        total_pl = sum(int(ph_d.get(k) or 0) for k in ["n_plants_outdoor","n_plants_hoop","n_plants_greenhouse","n_plants_indoor"])
        harv_pl = int(pos_d.get("harvested_plant_count") or 0)
        area = ph_d.get("cultivated_area_sqm")
        st.markdown(f"**Total Plants:** {total_pl or '—'}")
        st.markdown(f"**Harvested:** {harv_pl or '—'}")
        if total_pl and harv_pl:
            loss_pct = round((total_pl - harv_pl) / total_pl * 100, 1)
            st.markdown(f"**Plant Loss:** {total_pl - harv_pl} ({loss_pct}%)")
        st.markdown(f"**Cultivated Area:** {f'{area} sqM' if area else '—'}")
    with km_cols[2]:
        st.markdown('<p class="section-header">Harvest Date & COA</p>', unsafe_allow_html=True)
        result = pos_d.get("test_result") or "—"
        result_fmt = f"✅ {result}" if "Pass" in str(result) else (f"❌ {result}" if "Fail" in str(result) else result)
        st.markdown(f"**Harvest Date:** {pos_d.get('harvest_date') or '—'}")
        st.markdown(f"**COA Result:** {result_fmt}")
        st.markdown(f"**THC:** {f\"{ph_d.get('thc_pct')}%\" if ph_d.get('thc_pct') else '—'}")
        st.markdown(f"**CBD:** {f\"{ph_d.get('cbd_pct')}%\" if ph_d.get('cbd_pct') else '—'}")
    with km_cols[3]:
        st.markdown('<p class="section-header">Revenue</p>', unsafe_allow_html=True)
        fl_rev = ph_d.get("flower_revenue") or 0
        pr_rev = ph_d.get("preroll_revenue") or 0
        bm_rev = ph_d.get("biomass_revenue") or 0
        trim_rev = pos_d.get("trim_revenue") or 0
        total_rev = fl_rev + pr_rev + bm_rev + trim_rev
        if total_rev > 0:
            st.markdown(f"**Flower:** ${fl_rev:,.2f}")
            st.markdown(f"**Pre-Rolls:** ${pr_rev:,.2f}")
            st.markdown(f"**Biomass:** ${bm_rev:,.2f}")
            st.markdown(f"**Trim:** ${trim_rev:,.2f}")
            st.markdown(f"**Total: ${total_rev:,.2f}**")
        else:
            st.markdown("No revenue data entered.")
    st.divider()

    # Weight flow
    wet = pos_d.get("total_wet_weight_g") or 0
    dried = pos_d.get("dried_weight_g") or 0
    trimmed = pos_d.get("trimmed_flower_g") or 0
    trim_byp = pos_d.get("trim_weight_g") or 0
    packaged = (
        (pos_d.get("pkg_1g") or 0) * 1 +
        (pos_d.get("pkg_eighth") or 0) * 3.5 +
        (pos_d.get("pkg_quarter") or 0) * 7 +
        (pos_d.get("pkg_half") or 0) * 14 +
        (pos_d.get("pkg_oz") or 0) * 28
    )
    if wet > 0:
        st.markdown("#### Weight Flow")
        if HAS_PLOTLY and dried > 0:
            stages = ["Wet Weight","Dried Weight","Trimmed Flower","Packaged"]
            values = [wet, dried, trimmed if trimmed else dried, packaged if packaged else (trimmed or dried)]
            fig = go.Figure(go.Funnel(y=stages, x=values, textinfo="value+percent initial",
                marker_color=["#4CAF50","#2196F3","#FF9800","#9C27B0"],
                connector={"line": {"color":"royalblue","dash":"solid","width":2}}))
            fig.update_layout(height=300, margin=dict(l=20,r=20,t=20,b=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            wf_cols = st.columns(4)
            for col, (label, val) in zip(wf_cols, [
                ("Wet Weight (g)",wet),("Dried (g)",dried),("Trimmed Flower (g)",trimmed),("Packaged (g)",packaged)
            ]):
                col.metric(label, f"{val:,.0f}" if val else "—")
        if wet and dried:
            moisture = round((1 - dried / wet) * 100, 1)
            notes = [f"Moisture loss: **{moisture}%**"]
            if trimmed and dried: notes.append(f"Flower recovery from dry: **{trimmed/dried*100:.1f}%**")
            if trim_byp: notes.append(f"Trim byproduct: **{trim_byp:,.0f} g**")
            st.caption(" | ".join(notes))
        st.divider()

    # Revenue chart
    if total_rev > 0 and HAS_PLOTLY:
        st.markdown("#### Revenue Breakdown")
        rev_labels, rev_values = [], []
        for label, val in [("Flower",fl_rev),("Pre-Rolls",pr_rev),("Biomass",bm_rev),("Trim",trim_rev)]:
            if val > 0:
                rev_labels.append(label); rev_values.append(val)
        if rev_labels:
            fig_rev = px.pie(values=rev_values, names=rev_labels,
                             title=f"Total Revenue: ${total_rev:,.2f}", hole=0.4,
                             color_discrete_sequence=["#4CAF50","#FF9800","#9C27B0","#2196F3"])
            fig_rev.update_layout(height=320, legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig_rev, use_container_width=True)
            st.divider()

    # Cannabinoids
    if _has_data(ph_d, "thc_pct","cbd_pct","b_myrcene_pct","b_caryophyllene_pct"):
        st.markdown("#### Cannabinoids & Terpenes")
        cn_cols = st.columns(5)
        for col, (label, key, fmt) in zip(cn_cols, [
            ("THC (%)","thc_pct","{:.1f}%"), ("CBD (%)","cbd_pct","{:.1f}%"),
            ("Other Cannabinoids","other_cannabinoids","{}"),
            ("β-Myrcene (%)","b_myrcene_pct","{:.3f}%"),
            ("β-Caryophyllene (%)","b_caryophyllene_pct","{:.3f}%"),
        ]):
            val = ph_d.get(key)
            if val:
                try: col.metric(label, fmt.format(val))
                except (TypeError, ValueError): col.metric(label, str(val))
        st.divider()

    # Pest log summary
    if ph and ph.get("pests_df") is not None and not ph["pests_df"].empty:
        pest_df = ph["pests_df"]
        st.markdown("#### Pest Control Log Summary")
        n_events = len(pest_df)
        n_apps = len(pest_df[pest_df.get("Event Type", pd.Series()).str.contains("Application", na=False)])
        pc1, pc2 = st.columns(2)
        pc1.metric("Total Events", n_events); pc2.metric("Application Events", n_apps)
        with st.expander("View pest control log"):
            st.dataframe(pest_df, use_container_width=True, hide_index=True)
        st.divider()

    # Nutrient log summary
    if ph and ph.get("nutrients_df") is not None and not ph["nutrients_df"].empty:
        nut_df = ph["nutrients_df"]
        st.markdown("#### Nutrient Amendment Log Summary")
        st.metric("Total Amendment Events", len(nut_df))
        with st.expander("View nutrient log"):
            st.dataframe(nut_df, use_container_width=True, hide_index=True)
        st.divider()

    # All-batches table
    if show_all and all_ids:
        st.markdown("#### All Batches Summary")
        rows = []
        for bid in all_ids:
            ph_b = ph_batches.get(bid, {}); pos_b = pos_batches.get(bid, {})
            ph_dd = ph_b.get("data", {}); pos_dd = pos_b.get("data", {})
            total_pl2 = sum(int(ph_dd.get(k) or 0) for k in ["n_plants_outdoor","n_plants_hoop","n_plants_greenhouse","n_plants_indoor"])
            wet_w = pos_dd.get("total_wet_weight_g") or 0; dry_w = pos_dd.get("dried_weight_g") or 0
            rev = sum([ph_dd.get("flower_revenue") or 0, ph_dd.get("preroll_revenue") or 0,
                       ph_dd.get("biomass_revenue") or 0, pos_dd.get("trim_revenue") or 0])
            rows.append({
                "Batch ID": bid, "Strain": ph_dd.get("strain_name") or "—",
                "Season": ph_dd.get("season_cycle") or "—",
                "Harvest Date": pos_dd.get("harvest_date") or "—",
                "Total Plants": total_pl2 or "—",
                "Wet Wt (g)": f"{wet_w:,.0f}" if wet_w else "—",
                "Dry Wt (g)": f"{dry_w:,.0f}" if dry_w else "—",
                "Moisture Loss": f"{round((1-dry_w/wet_w)*100,1)}%" if wet_w and dry_w else "—",
                "COA Result": pos_dd.get("test_result") or "—",
                "THC (%)": ph_dd.get("thc_pct") or "—",
                "Total Revenue": f"${rev:,.2f}" if rev else "—",
                "Pre-Harvest": "✅" if bid in ph_batches else "❌",
                "Post-Harvest": "✅" if bid in pos_batches else "❌",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        ab_csv = pd.DataFrame(rows).to_csv(index=False).encode()
        ab_col1, ab_col2 = st.columns(2)
        with ab_col1:
            st.download_button("⬇ All Batches Summary (CSV)", data=ab_csv,
                file_name="all_batches_summary.csv", mime="text/csv", use_container_width=True)
        with ab_col2:
            if HAS_OPENPYXL:
                ab_wb = Workbook(); ab_ws = ab_wb.active; ab_ws.title = "All_Batches"
                if rows:
                    ab_ws.append(list(rows[0].keys()))
                    for r in rows: ab_ws.append(list(r.values()))
                ab_buf = io.BytesIO(); ab_wb.save(ab_buf); ab_buf.seek(0)
                st.download_button("⬇ All Batches Summary (Excel)", data=ab_buf.getvalue(),
                    file_name="all_batches_summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        st.divider()

    # Per-batch download
    st.markdown(f"### Download — {dash_selected}")
    row_ph = {"Batch Number": dash_selected}; row_ph.update(ph_d)
    row_pos = {"Harvest Batch ID": dash_selected}; row_pos.update(pos_d)
    combined_df = pd.DataFrame([row_ph, row_pos])
    dd1, dd2 = st.columns(2)
    with dd1:
        st.download_button(f"⬇ {dash_selected} — Combined Summary (CSV)",
            data=combined_df.to_csv(index=False).encode(),
            file_name=f"{dash_selected}_combined_summary.csv", mime="text/csv",
            use_container_width=True)
    with dd2:
        if HAS_OPENPYXL and (ph or pos):
            xl = _build_combined_report(dash_selected, ph, pos)
            st.download_button(f"⬇ {dash_selected} — Full Report (Excel)", data=xl,
                file_name=f"{dash_selected}_full_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# DOWNLOAD SECTION
# ══════════════════════════════════════════════════════════════════════════════

st.divider()
st.markdown("### Download All Data")
st.caption("Downloads cover all batches in this session. Session data is not persisted after closing — download to keep your records.")

ph_bats  = st.session_state.preharvest_batches
pos_bats = st.session_state.postharvest_batches

dl_c1, dl_c2 = st.columns(2)

with dl_c1:
    st.markdown("**Pre-Harvest**")
    if ph_bats:
        rows_ph = [{"Batch Number": bn, **b.get("data", {})} for bn, b in ph_bats.items()]
        ph_csv = pd.DataFrame(rows_ph).to_csv(index=False).encode()
        btn1, btn2 = st.columns(2)
        with btn1:
            st.download_button("⬇ Summary (CSV)", data=ph_csv,
                file_name="preharvest_summary.csv", mime="text/csv", use_container_width=True)
        with btn2:
            if HAS_OPENPYXL:
                st.download_button("⬇ Full Data (Excel)", data=build_preharvest_excel(ph_bats),
                    file_name="preharvest_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    st.markdown("**Nutrient & Pest Logs**")
    all_logs = []
    for bn, b in ph_bats.items():
        for df_key in ["nutrients_df","pests_df"]:
            df = b.get(df_key)
            if df is not None and not df.empty:
                tmp = df.copy()
                tmp.insert(0, "Log Type", "Nutrient" if df_key == "nutrients_df" else "Pest")
                tmp.insert(0, "Batch Number", bn)
                all_logs.append(tmp)
    if all_logs:
        combined_logs = pd.concat(all_logs, ignore_index=True)
        lg1, lg2 = st.columns(2)
        with lg1:
            st.download_button("⬇ Logs (CSV)", data=combined_logs.to_csv(index=False).encode(),
                file_name="preharvest_logs.csv", mime="text/csv", use_container_width=True)
        with lg2:
            if HAS_OPENPYXL:
                st.download_button("⬇ Logs (Excel)", data=build_logs_excel(ph_bats),
                    file_name="preharvest_logs.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

with dl_c2:
    st.markdown("**Post-Harvest**")
    if pos_bats:
        rows_pos = [{"Harvest Batch ID": bn, **b.get("data", {})} for bn, b in pos_bats.items()]
        pos_csv = pd.DataFrame(rows_pos).to_csv(index=False).encode()
        btn3, btn4 = st.columns(2)
        with btn3:
            st.download_button("⬇ Summary (CSV)", data=pos_csv,
                file_name="postharvest_summary.csv", mime="text/csv", use_container_width=True)
        with btn4:
            if HAS_OPENPYXL:
                st.download_button("⬇ Full Data (Excel)", data=build_postharvest_excel(pos_bats),
                    file_name="postharvest_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    st.markdown("**Curing Log**")
    all_curing = []
    for bn, b in pos_bats.items():
        cdf = b.get("curing_log_df")
        if cdf is not None and not cdf.empty:
            tmp = cdf.copy(); tmp.insert(0, "Harvest Batch ID", bn)
            all_curing.append(tmp)
    if all_curing:
        combined_curing = pd.concat(all_curing, ignore_index=True)
        cr1, cr2 = st.columns(2)
        with cr1:
            st.download_button("⬇ Curing Log (CSV)", data=combined_curing.to_csv(index=False).encode(),
                file_name="curing_log.csv", mime="text/csv", use_container_width=True)
        with cr2:
            if HAS_OPENPYXL:
                st.download_button("⬇ Curing Log (Excel)", data=build_curing_excel(pos_bats),
                    file_name="curing_log.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

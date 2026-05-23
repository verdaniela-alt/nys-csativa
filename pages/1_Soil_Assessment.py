"""
1_Soil_Assessment.py — NY Cannabis/Hemp Soil Assessment Tool
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd
import plotly.graph_objects as go

from utils.nutrient_data import (
    NUTRIENTS, AMENDMENTS, QUICK_AMEND, LAB_FACTORS, UNIT_CONVERSIONS,
    LIME_RATE_TABLE, LIME_TILLAGE_FACTORS, lime_rate_lookup,
)
from utils.soil_api import get_soil_data
from utils.sidebar import render_sidebar

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Soil Assessment | NYS Cannabis Tool",
    page_icon="🌱",
    layout="wide",
)

render_sidebar()

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.disclaimer-box {
    background: #fff3cd;
    border: 2px solid #e0a800;
    border-radius: 8px;
    padding: 16px 20px;
    margin-bottom: 20px;
    font-size: 0.92rem;
}
.disclaimer-box b { color: #856404; }
.deficient  { background-color: #ffd6d6; color: #8b0000; font-weight: bold; padding: 2px 8px; border-radius: 4px; }
.adequate   { background-color: #d6f0d6; color: #006400; font-weight: bold; padding: 2px 8px; border-radius: 4px; }
.excess     { background-color: #fff3cd; color: #856404; font-weight: bold; padding: 2px 8px; border-radius: 4px; }
.soil-card  { background: #f0f7ff; border-left: 4px solid #1565C0; padding: 12px 16px; border-radius: 6px; margin-bottom: 12px; }
.amend-card { background: #fafafa; border: 1px solid #ddd; border-radius: 8px; padding: 14px 16px; margin-bottom: 10px; }
.amend-card h4 { margin: 0 0 6px 0; }
.tag { display: inline-block; padding: 2px 9px; border-radius: 10px; font-size: 0.78rem; font-weight: bold; margin-right: 4px; }
.tag-organic    { background: #d4edda; color: #155724; }
.tag-synthetic  { background: #e2e3e5; color: #383d41; }
.tag-powder     { background: #cfe2ff; color: #084298; }
.tag-liquid     { background: #d1ecf1; color: #0c5460; }
.tag-granular   { background: #fff3cd; color: #856404; }
.tag-pellet     { background: #f8d7da; color: #721c24; }
</style>
""", unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────
if "soil_data"       not in st.session_state: st.session_state.soil_data       = None
if "assessment_done" not in st.session_state: st.session_state.assessment_done = False

# ─────────────────────────────────────────────────────────────────────────────
# DISCLAIMER — TOP OF PAGE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="disclaimer-box">
<b>⚠️ Important Disclaimer — Please Read Before Using This Tool</b><br><br>
This tool provides <b>general agronomic guidance only</b>, based on cannabis and hemp
soil fertility literature and NY State extension frameworks.
<b>The suggestions presented here are possible options, not prescriptions.</b>
Results should be interpreted by a qualified professional before any action is taken.<br><br>
<b>Limitations you should be aware of:</b><br>
• Recommendations are based on <b>inherent soil properties</b> (from a recent soil test) and do
  <b>not</b> account for your current management practices — including cover cropping, compost or
  manure applications, irrigation practices, or other amendments already applied.<br>
• This tool does <b>not</b> differentiate between production systems. Recommendations may differ
  for field, greenhouse, hoop house, or container production.<br>
• The USDA NRCS SSURGO data pulled from your address reflects soil survey classifications
  collected largely in the 1970s. These describe <b>inherent</b> (relatively permanent) soil
  properties such as texture, drainage class, and slope — not <b>dynamic</b> properties such as
  current organic matter, recent pH changes, or the effect of your management on nutrient levels.<br>
• Nutrient targets are drawn from the best available literature for <em>Cannabis sativa</em>,
  which is limited and often extrapolated from related crops (e.g., tomatoes). As new research
  emerges, targets may be updated.<br><br>
<b>This tool and its developers assume no responsibility or liability</b> for any
decisions, actions, crop losses, financial outcomes, or regulatory consequences
arising from the use of this tool.<br><br>
Always consult a <b>certified crop advisor (CCA)</b>, licensed agronomist, or your local
Cornell Cooperative Extension office before making large-scale amendment applications.
Compliance with all applicable <b>NY State OCM and Ag &amp; Markets</b> regulations is
the sole responsibility of the grower.
</div>
""", unsafe_allow_html=True)

# ── Page header ───────────────────────────────────────────────────────────────
st.title("🌱 NY Cannabis & Hemp Soil Assessment")
st.caption("Outdoor & greenhouse cultivation — New York State | Data: USDA NRCS SSURGO + your lab report")
st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — SITE LOCATION
# ─────────────────────────────────────────────────────────────────────────────
with st.expander("📍 Step 1: Site Location & Soil Survey Lookup", expanded=True):
    st.markdown("Enter your farm address to automatically pull USDA NRCS soil data for your field.")

    col1, col2 = st.columns([3, 1])
    with col1:
        address_input = st.text_input(
            "Farm / Field Address",
            placeholder="e.g.  42 Example Farm Rd, Penn Yan, NY 14527",
            key="address_input",
        )
    with col2:
        st.write("")
        st.write("")
        lookup_btn = st.button("🔍 Look Up Soil Data", use_container_width=True)

    if lookup_btn and address_input.strip():
        with st.spinner("Geocoding address and querying NRCS Soil Data Access…"):
            try:
                soil = get_soil_data(address_input.strip())
                st.session_state.soil_data = soil
                st.success(f"✅ Matched: **{soil['matched_address']}**  "
                           f"| Coordinates: {soil['lat']:.5f}°N, {soil['lon']:.5f}°W")
            except ValueError as e:
                st.error(f"❌ Address not found: {e}")
            except Exception as e:
                st.error(f"❌ Lookup error: {e}")

    # ── NY State map (always shown; pin appears after lookup) ─────────────────
    _soil_now = st.session_state.soil_data
    _pin_lat  = [_soil_now["lat"]]  if _soil_now else []
    _pin_lon  = [_soil_now["lon"]]  if _soil_now else []
    _pin_text = [_soil_now.get("matched_address", "Farm location")] if _soil_now else []

    _fig_map = go.Figure(go.Scattermapbox(
        lat=_pin_lat,
        lon=_pin_lon,
        mode="markers+text",
        marker=go.scattermapbox.Marker(size=18, color="#e53935"),
        text=_pin_text,
        textposition="top right",
        textfont=dict(size=12, color="#e53935"),
        hoverinfo="text",
    ))
    _fig_map.update_layout(
        mapbox=dict(
            style="open-street-map",
            center=dict(lat=42.9, lon=-75.7),
            zoom=6,
        ),
        margin=dict(l=0, r=0, t=0, b=0),
        height=420,
    )
    st.markdown("**📍 New York State — Field Location**")
    st.caption("Enter your address above and click Look Up to place a pin on the map.")
    st.plotly_chart(_fig_map, use_container_width=True, key="ny_state_map")

    # ── NRCS Soil Survey Results (shown after lookup) ──────────────────────────
    if _soil_now:
        soil    = _soil_now
        comp    = soil.get("comp")
        horizon = soil.get("horizon")

        if comp or horizon:
            st.markdown("### 🗺️ NRCS Soil Survey Results")
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Soil Component**")
                if comp:
                    st.markdown(f"""
<div class="soil-card">
<b>Map Unit:</b> {comp.get('map_unit','—')}<br>
<b>Dominant Series:</b> {comp.get('series','—')} ({comp.get('pct','—')}% of map unit)<br>
<b>Taxonomy:</b> {comp.get('taxonomy','—')}<br>
<b>Drainage Class:</b> {comp.get('drainage','—')}<br>
<b>Hydrologic Group:</b> {comp.get('hydro_grp','—')}<br>
<b>Representative Slope:</b> {comp.get('slope','—')}%
</div>""", unsafe_allow_html=True)
                else:
                    st.warning("No component data — address may be outside SSURGO coverage.")
            with c2:
                st.markdown("**Surface Horizon (0 cm)**")
                if horizon:
                    st.markdown(f"""
<div class="soil-card">
<b>Horizon:</b> {horizon.get('horizon','—')} ({horizon.get('depth_top','?')}–{horizon.get('depth_bot','?')} cm)<br>
<b>Texture Class:</b> {horizon.get('texture','—')}<br>
<b>pH (1:1 H₂O):</b> {horizon.get('ph','—') or '—'}<br>
<b>CEC:</b> {f"{horizon['cec']} meq/100g" if horizon.get('cec') else '—'}<br>
<b>Organic Matter:</b> {f"{horizon['om']}%" if horizon.get('om') else '—'}
</div>""", unsafe_allow_html=True)
                else:
                    st.warning("No horizon data returned.")
        else:
            st.warning("⚠️ No NRCS data found. You can still enter data manually below.")

    st.caption("Data from USDA NRCS SSURGO. For an interactive map visit "
               "[SoilWeb](https://casoilresource.lawr.ucdavis.edu/gmap/).")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — CROP & LAB SELECTION
# ─────────────────────────────────────────────────────────────────────────────
with st.expander("🌿 Step 2: Crop Type & Laboratory", expanded=True):
    col1, col2 = st.columns(2)

    with col1:
        crop = st.selectbox(
            "Crop Type",
            ["Hemp (fiber / grain / CBD)", "Cannabis (MJ, Adult-Use / Medical)"],
            key="crop_select",
        )

    with col2:
        lab = st.selectbox(
            "Soil Laboratory / Extraction Method",
            list(LAB_FACTORS.keys()),
            key="lab_select",
            help="Selects the conversion method for P and K. Modified Morgan labs (Agro-One MM, UVM/UConn/UMass): P ×1.8 and K ×1.0 to estimate Mehlich III equivalents (Cornell NMSP v7). Dairy One and Agro-One Mehlich III labs need no conversion.",
        )

    crop_key = "hemp" if "Hemp" in crop else "mj"

    if "Modified Morgan" in lab and "Cornell Soil Health" not in lab:
        st.info(
            "ℹ️ **Modified Morgan lab detected.**  \n"
            "P will be multiplied ×**1.8** and K ×**1.0** to approximate Mehlich III equivalents "
            "(Cornell NMSP Conversion Tool v7, Ketterings 2005).  \n"
            "**Note:** P, K, Ca, Mg, Mn, Zn, and Al are typically reported in **lbs/acre** — "
            "set the unit selectors below to **lbs / acre** for those nutrients.  \n"
            "Fe is typically reported without a unit label (enter as ppm).  \n"
            "S, Cu, and B are not typically reported by Modified Morgan labs."
        )
    elif "Cornell Soil Health" in lab:
        st.info(
            "ℹ️ **Cornell Soil Health Lab detected.**  \n"
            "Values are already Cornell Modified Morgan equivalents reported in **lbs/acre**. "
            "Set unit selectors to **lbs / acre**. No M3 conversion is applied — "
            "values are compared directly to Modified Morgan lbs/acre targets."
        )
    elif "Logan Labs" in lab:
        st.warning(
            "⚠️ **Logan Labs — P₂O₅ conversion required before entering phosphorus.**  \n\n"
            "Logan Labs reports phosphorus as **lbs/acre P₂O₅**, not elemental P. "
            "Entering the value directly will overstate P by ~2.3× and may produce incorrect results.  \n\n"
            "**Before entering P:** multiply the value on your report by **0.437**  \n"
            "Example: report shows 200 lbs/ac → enter **87 lbs/ac** (200 × 0.437)  \n\n"
            "**Units for Logan Labs:**  \n"
            "• Macronutrients (P, K, Ca, Mg): **lbs / acre** — set the macronutrient unit selector to lbs/acre  \n"
            "• Micronutrients (B, Cu, Fe, Mn, Zn) and Al: **ppm** — leave micronutrient unit at ppm  \n"
            "• No MM→M3 conversion is applied — Logan Labs values are already Mehlich III equivalents."
        )
    elif lab and "Mehlich" in lab:
        st.info(
            "ℹ️ **Mehlich III lab detected.** Values are already M3 equivalents — "
            "enter in the units shown on your report (typically ppm for most nutrients)."
        )

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3a — UPLOAD LAB REPORT (optional)
# ─────────────────────────────────────────────────────────────────────────────

# Column name → nutrient key mapping (case-insensitive)
_UPLOAD_COL_MAP = {
    "ph": "pH",
    "organic matter": "Organic Matter", "om": "Organic Matter",
    "om%": "Organic Matter", "om (%)": "Organic Matter", "% om": "Organic Matter",
    "p": "P (Phosphorus)", "phosphorus": "P (Phosphorus)",
    "p (phosphorus)": "P (Phosphorus)", "p_ppm": "P (Phosphorus)", "p_lbs": "P (Phosphorus)",
    "k": "K (Potassium)", "potassium": "K (Potassium)",
    "k (potassium)": "K (Potassium)", "k_ppm": "K (Potassium)",
    "ca": "Ca (Calcium)", "calcium": "Ca (Calcium)", "ca (calcium)": "Ca (Calcium)",
    "mg": "Mg (Magnesium)", "magnesium": "Mg (Magnesium)", "mg (magnesium)": "Mg (Magnesium)",
    "s": "S (Sulfur)", "sulfur": "S (Sulfur)", "s (sulfur)": "S (Sulfur)",
    "zn": "Zn (Zinc)", "zinc": "Zn (Zinc)", "zn (zinc)": "Zn (Zinc)", "zn_n": "Zn (Zinc)",
    "mn": "Mn (Manganese)", "manganese": "Mn (Manganese)", "mn (manganese)": "Mn (Manganese)",
    "fe": "Fe (Iron)", "iron": "Fe (Iron)", "fe (iron)": "Fe (Iron)",
    "cu": "Cu (Copper)", "copper": "Cu (Copper)", "cu (copper)": "Cu (Copper)", "cu_n": "Cu (Copper)",
    "b": "B (Boron)", "boron": "B (Boron)", "b (boron)": "B (Boron)",
    "na": "Na (Sodium)", "sodium": "Na (Sodium)", "na (sodium)": "Na (Sodium)",
    "al": "Al (Aluminum)", "aluminum": "Al (Aluminum)", "al (aluminum)": "Al (Aluminum)",
    "aluminium": "Al (Aluminum)",
    "cec": "CEC", "total exchange capacity": "CEC", "t.e.c.": "CEC",
    "ca%": "Base Saturation Ca%", "ca sat": "Base Saturation Ca%",
    "ca saturation": "Base Saturation Ca%", "% ca": "Base Saturation Ca%",
    "k%": "Base Saturation K%", "k sat": "Base Saturation K%",
    "k saturation": "Base Saturation K%", "% k": "Base Saturation K%",
}

_TEMPLATE_CSV = (
    "Sample ID,pH,Organic Matter (%),P,K,Ca,Mg,S,Zn,Mn,Fe,Cu,B,Na,Al,CEC,Ca%,K%\n"
    "# Units: enter ppm for Mehlich III labs; lbs/acre for Modified Morgan labs\n"
    "# IMPORTANT - Logan Labs users: P is reported as P2O5 lbs/acre on your report.\n"
    "# Multiply your P value by 0.437 before entering here (e.g. 200 lbs/ac P2O5 -> 87 lbs/ac elemental P).\n"
    "Sample-1,,,,,,,,,,,,,,,,\n"
)

def _find_header_row(df_raw):
    """Find the row index that looks like a nutrient header."""
    nutrient_keywords = {"ph", "phosphorus", "potassium", "calcium", "magnesium", "p", "k", "ca", "mg"}
    for i, row in df_raw.iterrows():
        cells = {str(v).strip().lower() for v in row if v is not None and str(v).strip()}
        if len(cells & nutrient_keywords) >= 2:
            return i
    return 0

def _parse_lab_file(file_bytes, filename):
    """Parse CSV or Excel lab report. Returns (list of {sample_id, values}, error)."""
    import io
    try:
        if filename.lower().endswith(".csv"):
            df_raw = pd.read_csv(io.BytesIO(file_bytes), header=None, comment="#")
        else:
            df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    except Exception as e:
        return None, str(e)

    hdr_idx = _find_header_row(df_raw)
    headers = [str(v).strip() for v in df_raw.iloc[hdr_idx]]

    col_to_nutrient = {}
    sample_col = None
    for ci, h in enumerate(headers):
        h_lo = h.lower()
        if h_lo in ("sample", "sample id", "sample_id", "id", "name", "sample name"):
            sample_col = ci
        elif h_lo in _UPLOAD_COL_MAP:
            col_to_nutrient[ci] = _UPLOAD_COL_MAP[h_lo]

    if not col_to_nutrient:
        return None, "No recognized nutrient columns found. Check column headers match: pH, P, K, Ca, Mg, etc."

    results = []
    for ri in range(hdr_idx + 1, len(df_raw)):
        row = df_raw.iloc[ri]
        sample_id = str(row.iloc[sample_col]).strip() if sample_col is not None else f"Row {ri}"
        if sample_id in ("nan", "", "None"):
            sample_id = f"Row {ri}"
        vals = {}
        for ci, nname in col_to_nutrient.items():
            raw_val = row.iloc[ci]
            try:
                v = float(raw_val)
                if pd.notna(v):
                    vals[nname] = v
            except (ValueError, TypeError):
                pass
        if vals:
            results.append({"sample_id": sample_id, "values": vals})

    if not results:
        return None, "File parsed but no numeric data rows found."
    return results, None


# ── PDF parsing helpers ───────────────────────────────────────────────────────

# Regex patterns: nutrient name → list of patterns tried in order
_PDF_PATTERNS = [
    ("pH",                   [r"(?:Soil\s+)?pH(?:\s+of\s+Soil\s+Sample)?\s+(\d+\.?\d*)"]),
    ("Organic Matter",       [r"(?:Organic\s+Matter|%\s*OM)[^<\n]*?(\d+\.?\d*)\s*%",
                              r"(?:%\s*)?Organic\s+Matter[,\s%]*(\d+\.?\d*)",
                              r"%\s*OM\s+(\d+\.?\d*)",
                              r"Organic\s+Matter\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("P (Phosphorus)",       [r"Phosphorus?\s*(?:\(P\))?\s*(?:\[Mehlich\s*III\])?\s+(\d+\.?\d*)",
                              r"Mehlich\s+III\s+Phospho\w+[^<\n]*?(\d+\.?\d*)",
                              r"Phosphorus\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("K (Potassium)",        [r"Potassium\s*(?:\(K\))?\s+(\d+\.?\d*)"]),
    ("Ca (Calcium)",         [r"Calcium\s*(?:\(Ca\))?\s+(?:lbs/acre\s+)?(\d[\d,]*\.?\d*)"]),
    ("Mg (Magnesium)",       [r"Magnesium\s*(?:\(Mg\))?\s+(?:lbs/acre\s+)?(\d+\.?\d*)"]),
    ("S (Sulfur)",           [r"Sulfur\s*(?:\(S\))?\s+(\d+\.?\d*)",
                              r"Sulfur\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("Fe (Iron)",            [r"Iron\s*(?:\(Fe\))?[,\s]*(?:lbs/acre)?\s+(\d+\.?\d*)",
                              r"Iron\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("Mn (Manganese)",       [r"Manganese\s*(?:\(Mn\))?[,\s]*(?:lbs/acre)?\s+(\d+\.?\d*)",
                              r"Manganese\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("Zn (Zinc)",            [r"Zinc\s*(?:\(Zn\))?[,\s]*(?:lbs/acre)?\s+(\d+\.?\d*)",
                              r"Zinc\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("Cu (Copper)",          [r"Copper\s*(?:\(Cu\))?[,\s]*(?:lbs/acre)?\s+(\d+\.?\d*)",
                              r"Copper\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("B (Boron)",            [r"Boron\s*(?:\(B\))?\s+(\d+\.?\d*)",
                              r"Boron\s*\([^)]*\)\s+(\d+\.?\d*)"]),
    ("Na (Sodium)",          [r"Sodium\s*(?:\(Na\))?\s+(\d+\.?\d*)"]),
    ("Al (Aluminum)",        [r"Alum(?:inum|inium)\s*(?:\(Al\))?[,\s]*(?:lbs/acre|ppm)?\s+(\d+\.?\d*)"]),
    ("CEC",                  [r"(?:Total\s+)?(?:Exchange\s+Capacity|EC\s*[\(\[]?Cation|CEC)[^\n]*?(\d+\.?\d*)\s*meq",
                              r"Total\s+Exchange\s+Capacity[^\n]*?(\d+\.?\d*)"]),
    ("Base Saturation Ca%",  [r"(?:Calcium|Ca)\s*\(60\s*to\s*70%\)\s+(\d+\.?\d*)",
                              r"%\s*Calcium\s*\(?60[–-]70%\)?\s+(\d+\.?\d*)",
                              r"Calcium\s*\(%\s*Sat\w*\)\s*(\d+\.?\d*)"]),
    ("Base Saturation K%",   [r"(?:Potassium|K)\s*\(2\s*to\s*5%\)\s+(\d+\.?\d*)",
                              r"%\s*Potassium\s*\(?2[–-]5%\)?\s+(\d+\.?\d*)",
                              r"Potassium\s*\(%\s*Sat\w*\)\s*(\d+\.?\d*)"]),
    ("Base Saturation Mg%",  [r"(?:Magnesium|Mg)\s*\(10\s*to\s*20%\)\s+(\d+\.?\d*)",
                              r"%\s*Magnesium\s*\(?10[–\-]20%\)?\s+(\d+\.?\d*)",
                              r"(?:Magnesium|Mg)\s*\(10[–\-]20%\)\s+(\d+\.?\d*)",
                              r"Magnesium\s*\(%\s*Sat\w*\)\s*(\d+\.?\d*)"]),
    ("EC (Soluble Salts)",   [r"Soluble\s+Salts[,\s]*(?:mmhos/cm|mS/cm|dS/m)?\s+(\d+\.?\d*)"]),
    ("Buffer pH",            [r"Buffer\s+pH\s+(\d+\.?\d*)"]),
]

def _parse_pdf_page_text(text, selected_lab):
    """
    Extract nutrient values from one page of PDF text using keyword regex.
    Returns {nutrient_name: float}, warnings list.
    """
    import re
    values = {}
    warnings = []

    # Check for below-detection-limit tokens so we can skip them cleanly
    bdl_pattern = re.compile(r"<\s*0?\.\d+|ND|BDL|not\s+detected", re.IGNORECASE)

    for nutrient, patterns in _PDF_PATTERNS:
        if nutrient == "Buffer pH":
            continue  # informational only, not an app input field
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                raw = m.group(0)
                # Check if the match is preceded by a BDL token on the same line
                line_start = text.rfind("\n", 0, m.start()) + 1
                line = text[line_start: text.find("\n", m.end())]
                if bdl_pattern.search(line):
                    warnings.append(f"**{nutrient}**: below detection limit — left blank.")
                    break
                # Strip commas (e.g. "3,076" → 3076)
                val_str = m.group(1).replace(",", "")
                try:
                    val = float(val_str)
                    # Calcium sanity: Dairy One reports Ca as lbs/acre (can be 3000+),
                    # New Age Labs reports in ppm (also can be 1500+) — both fine
                    values[nutrient] = val
                    break
                except ValueError:
                    pass

    # Apply P₂O₅ → elemental P conversion for Logan Labs
    is_logan = "Logan" in selected_lab
    p_converted = False
    if is_logan and "P (Phosphorus)" in values:
        raw_p = values["P (Phosphorus)"]
        values["P (Phosphorus)"] = round(raw_p * 0.437, 1)
        warnings.append(
            f"**P (Phosphorus)**: Logan Labs reports P as P₂O₅. "
            f"Raw value {raw_p} lbs/ac × 0.437 = **{values['P (Phosphorus)']} lbs/ac elemental P** — "
            f"this conversion was applied automatically."
        )
        p_converted = True

    # University of Maine Soil Testing Service: K, Ca, Mg (lbs/acre) and CEC
    # are only in the numerical results row — supplement standard pattern results
    if _is_maine_soil_testing_pdf(text):
        maine_vals = _parse_maine_numerical_row(text)
        for k, v in maine_vals.items():
            if k not in values:
                values[k] = v
        if maine_vals:
            warnings.append(
                "**University of Maine Soil Testing Service format detected.**  \n"
                "K, Ca, and Mg values are in **lbs/acre** — set the macronutrient unit "
                "selector to **lbs/acre** before running the assessment."
            )

    return values, warnings


def _parse_pdf_logan_table(page_obj, col_idx):
    """
    Parse a Logan Labs PDF page (multi-column table) and extract values for one sample column.
    col_idx: 0-based index into the data columns (0 = first sample).
    Returns {nutrient_name: float}, warnings list.
    """
    import re
    warnings = []
    tables = page_obj.extract_tables()
    if not tables:
        return {}, ["Could not extract table from Logan Labs PDF page."]

    # Logan Labs table: rows are nutrients, columns are samples
    # Row label is col 0; data starts at col 1
    table = tables[0]

    # Build row-label → value mapping for the selected column
    data_col = col_idx + 1  # +1 to skip row label column

    ROW_MAP = {
        "ph of soil sample":        "pH",
        "organic matter, percent":  "Organic Matter",
        "sulfur":                   "S (Sulfur)",
        "mehlich iii phosphorous":  "P (Phosphorus)",
        "mehlich iii phosphorus":   "P (Phosphorus)",
        "potassium":                "K (Potassium)",
        "calcium":                  "Ca (Calcium)",
        "magnesium":                "Mg (Magnesium)",
        "sodium":                   "Na (Sodium)",
        "calcium (60 to 70%)":      "Base Saturation Ca%",
        "magnesium (10 to 20%)":    "Base Saturation Mg%",
        "potassium (2 to 5%)":      "Base Saturation K%",
        "boron (p.p.m.)":           "B (Boron)",
        "iron (p.p.m.)":            "Fe (Iron)",
        "manganese (p.p.m.)":       "Mn (Manganese)",
        "copper (p.p.m.)":          "Cu (Copper)",
        "zinc (p.p.m.)":            "Zn (Zinc)",
        "aluminum (p.p.m.)":        "Al (Aluminum)",
        "total exchange capacity":  "CEC",
    }

    values = {}
    prev_label = None

    for row in table:
        if not row or data_col >= len(row):
            continue
        label_raw = str(row[0] or "").strip().lower()
        cell_raw  = str(row[data_col] or "").strip()

        # For Ca/Mg/K, Logan has 3 sub-rows: Desired / Value Found / Deficit
        # We want "Value Found" (second sub-row)
        if label_raw in ("value found", "") and prev_label in (
            "Ca (Calcium)", "Mg (Magnesium)", "K (Potassium)"
        ):
            try:
                values[prev_label] = float(cell_raw.replace(",", ""))
            except ValueError:
                pass
            continue

        nutrient = ROW_MAP.get(label_raw)
        if nutrient:
            prev_label = nutrient
            # Skip Desired/Deficit rows — only take direct value rows
            if nutrient in ("Ca (Calcium)", "Mg (Magnesium)", "K (Potassium)"):
                continue  # value comes from "Value Found" sub-row
            try:
                val = float(cell_raw.replace(",", ""))
                values[nutrient] = val
            except ValueError:
                if cell_raw.startswith(">"):
                    try:
                        values[nutrient] = float(cell_raw[1:].replace(",", ""))
                        warnings.append(f"**{nutrient}**: reported as '{cell_raw}' — entered as {values[nutrient]}.")
                    except ValueError:
                        pass

    # P₂O₅ conversion
    if "P (Phosphorus)" in values:
        raw_p = values["P (Phosphorus)"]
        values["P (Phosphorus)"] = round(raw_p * 0.437, 1)
        warnings.append(
            f"**P (Phosphorus)**: Logan Labs reports P as P₂O₅. "
            f"Raw value {raw_p} lbs/ac × 0.437 = **{values['P (Phosphorus)']} lbs/ac elemental P** — "
            f"applied automatically."
        )

    return values, warnings


def _get_pdf_pages(file_bytes):
    """
    Open PDF with pdfplumber and return list of {page_num, label, text, page_obj}.
    Skips pages that appear to be blank or cover letters (no nutrient data).
    """
    import pdfplumber, io, re
    results = []
    try:
        pdf = pdfplumber.open(io.BytesIO(file_bytes))
    except Exception as e:
        return None, str(e)

    nutrient_signals = re.compile(
        r"phospho|potassium|calcium|magnesium|organic matter|pH|sulfur|boron|iron|manganese",
        re.IGNORECASE
    )

    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        has_data = bool(nutrient_signals.search(text))
        # Build a label from sample location if detectable
        label = None
        for pat in [r"Field/Location:\s*(.+)", r"Field Name:\s*(.+)"]:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                candidate = m.group(1).strip()
                if candidate and ":" not in candidate:
                    label = candidate
                    break
        if not label:
            # New Age Labs: "Sample Location: Collected By:\nNorth Field ..."
            m = re.search(r"Sample Location:[^\n]*\n([^\n]+)", text, re.IGNORECASE)
            if m:
                # First word(s) before a multi-space gap or another name
                raw = m.group(1).strip()
                # Take only the first field (before 2+ spaces)
                parts = re.split(r"\s{2,}", raw)
                label = parts[0].strip() if parts else raw
        if not label:
            m = re.search(r"Sample Name:\s*(.+)", text, re.IGNORECASE)
            if m:
                candidate = m.group(1).strip()
                if candidate and ":" not in candidate:
                    label = candidate
        if not label:
            # U of Maine: first line "MM/DD/YYYY LABNO SAMPLE_NAME COUNTY SIZE"
            m = re.search(
                r"^\d{2}/\d{2}/\d{4}\s+\d+\s+(.+?)\s+\w+\s+[\d,]+\s+(?:Acres?|sq)",
                text, re.MULTILINE | re.IGNORECASE
            )
            if m:
                label = m.group(1).strip()
        results.append({
            "page_num":  i + 1,
            "label":     label or f"Page {i + 1}",
            "text":      text,
            "page_obj":  page,
            "has_data":  has_data,
        })

    pdf.close()
    return results, None


def _is_logan_pdf(text):
    """Detect Logan Labs format (multi-column table on one page)."""
    import re
    return bool(re.search(r"logan\s+labs", text, re.IGNORECASE)) or \
           bool(re.search(r"mehlich\s+iii\s+phospho", text, re.IGNORECASE) and
                re.search(r"p\.p\.m\.", text, re.IGNORECASE))


def _is_maine_soil_testing_pdf(text):
    """Detect University of Maine Soil Testing Service format."""
    import re
    return bool(re.search(r"maine\s+soil\s+testing\s+service|university\s+of\s+maine", text, re.IGNORECASE))


def _parse_maine_numerical_row(text):
    """
    Extract K, Ca, Mg (lbs/acre) and CEC from the U of Maine 'Level Found' numerical row.
    Row format: Level [pH] [LimeIdx] [P-lb/A] [K-lb/A] [Mg-lb/A] [Ca-lb/A] [CEC(me/100g)] ...
    Returns dict of extracted values (may be empty).
    """
    import re
    result = {}
    m = re.search(
        r"Level\s+[\d.]+\s+[\d.]+\s+[\d.]+\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d.]+)",
        text
    )
    if m:
        result["K (Potassium)"]  = float(m.group(1).replace(",", ""))
        result["Mg (Magnesium)"] = float(m.group(2).replace(",", ""))
        result["Ca (Calcium)"]   = float(m.group(3).replace(",", ""))
        result["CEC"]            = float(m.group(4))
    return result


def _get_logan_column_names(page_obj):
    """Extract sample column header names from a Logan Labs table."""
    tables = page_obj.extract_tables()
    if not tables:
        return []
    table = tables[0]
    # Header rows: first rows typically contain sample location / sample ID
    # Scan first 3 rows for non-empty cells after col 0
    names = []
    for row in table[:4]:
        if not row:
            continue
        candidates = [str(c).strip() for c in row[1:] if c and str(c).strip() not in ("", "None")]
        if candidates and len(candidates) > len(names):
            names = candidates
    return names if names else [f"Sample {i+1}" for i in range(5)]


# ── Step 2b UI ────────────────────────────────────────────────────────────────
with st.expander("📎 Step 2b — Upload Lab Report (PDF, CSV, or Excel — optional)", expanded=False):
    st.markdown(
        "Upload your lab report to auto-fill the fields below. "
        "Accepted formats: **PDF**, **CSV**, or **Excel (.xlsx)**."
    )

    up_col1, up_col2 = st.columns([3, 1])
    with up_col1:
        lab_file = st.file_uploader(
            "Choose file", type=["pdf", "csv", "xlsx"], key="lab_upload",
            label_visibility="collapsed",
        )
    with up_col2:
        st.download_button(
            "⬇ Download CSV template",
            data=_TEMPLATE_CSV.encode(),
            file_name="soil_test_template.csv",
            mime="text/csv",
            use_container_width=True,
        )

    if lab_file is not None:
        file_bytes = lab_file.read()
        fname = lab_file.name.lower()

        # ── PDF path ──────────────────────────────────────────────────────────
        if fname.endswith(".pdf"):
            pages, pdf_err = _get_pdf_pages(file_bytes)

            if pdf_err:
                st.error(f"Could not read PDF: {pdf_err}")

            elif not pages:
                st.error("No pages found in PDF.")

            else:
                data_pages = [p for p in pages if p["has_data"]]
                all_pages  = pages

                if not data_pages:
                    st.warning(
                        "⚠️ No nutrient data detected in this PDF. "
                        "It may be a scanned image (not a digital PDF). "
                        "Please use the CSV template instead, or contact your lab for a digital copy."
                    )
                else:
                    # Page selector — show only pages with detected data, but let user override
                    page_options = {
                        f"Page {p['page_num']} — {p['label']}": p
                        for p in data_pages
                    }
                    if len(page_options) > 1:
                        st.info(
                            f"Found **{len(data_pages)} page(s) with soil data** in this PDF. "
                            "Select the page that contains the sample you want to analyze."
                        )
                        selected_label = st.selectbox(
                            "Select page / sample",
                            list(page_options.keys()),
                            key="pdf_page_sel",
                        )
                        selected_page = page_options[selected_label]
                    else:
                        selected_page = data_pages[0]
                        st.success(f"Detected soil data on **Page {selected_page['page_num']}** — {selected_page['label']}.")

                    page_text = selected_page["text"]
                    page_obj  = selected_page["page_obj"]

                    # Logan Labs: multi-column table → column selector
                    is_logan = _is_logan_pdf(page_text)
                    col_warnings = []

                    if is_logan:
                        col_names = _get_logan_column_names(page_obj)
                        if len(col_names) > 1:
                            st.info(
                                f"Logan Labs format detected — **{len(col_names)} samples** found on this page. "
                                "Select the sample column you want to analyze."
                            )
                            chosen_col_name = st.selectbox(
                                "Select sample column",
                                col_names,
                                key="pdf_logan_col_sel",
                            )
                            col_idx = col_names.index(chosen_col_name)
                        else:
                            col_idx = 0
                            chosen_col_name = col_names[0] if col_names else "Sample 1"

                        parsed_vals, col_warnings = _parse_pdf_logan_table(page_obj, col_idx)
                        sample_label = f"{selected_page['label']} — {chosen_col_name}"
                    else:
                        parsed_vals, col_warnings = _parse_pdf_page_text(page_text, lab)
                        sample_label = selected_page["label"]

                    if not parsed_vals:
                        st.error(
                            "Could not extract nutrient values from the selected page. "
                            "The PDF may be a scanned image or use an unsupported layout. "
                            "Please enter values manually or use the CSV template."
                        )
                    else:
                        # ── Conversion warnings ───────────────────────────────
                        if col_warnings:
                            for w in col_warnings:
                                st.info(f"ℹ️ {w}")

                        # ── Preview table ─────────────────────────────────────
                        st.markdown(f"**Extracted values — {sample_label}**")
                        preview_rows = [
                            {"Nutrient": k, "Value extracted from PDF": v}
                            for k, v in parsed_vals.items()
                        ]
                        st.dataframe(
                            pd.DataFrame(preview_rows),
                            hide_index=True,
                            use_container_width=True,
                        )

                        # ── User responsibility confirmation ──────────────────
                        st.warning(
                            "⚠️ **Please verify these values before applying.**  \n\n"
                            "Check that each value above matches what is printed on your lab report "
                            f"for the sample you selected (**{sample_label}**). "
                            "PDF extraction is automated and may occasionally misread values — "
                            "particularly for scanned reports, unusual formatting, or detection-limit entries.  \n\n"
                            "**By clicking 'Apply' below, you confirm that you have reviewed these values "
                            "and that they correspond to the correct sample and page. "
                            "The accuracy of your soil assessment depends entirely on the accuracy of the data entered.**"
                        )

                        if st.button(
                            "✅ I have reviewed the values — Apply to input fields",
                            type="primary",
                            use_container_width=True,
                            key="pdf_apply_btn",
                        ):
                            for nname, val in parsed_vals.items():
                                st.session_state[f"nutrient_{nname}"] = val
                            st.success(
                                f"✅ Fields populated from **{sample_label}**. "
                                "Scroll down to review and adjust any values before running the assessment."
                            )
                            st.rerun()

        # ── CSV / Excel path (unchanged) ──────────────────────────────────────
        else:
            parsed, err = _parse_lab_file(file_bytes, fname)
            if err:
                st.error(f"Could not parse file: {err}")
            else:
                sample_ids = [r["sample_id"] for r in parsed]
                if len(parsed) > 1:
                    chosen_id = st.selectbox("Select sample", sample_ids, key="upload_sample_sel")
                    chosen = next(r for r in parsed if r["sample_id"] == chosen_id)
                else:
                    chosen = parsed[0]
                    st.success(f"Parsed: **{chosen['sample_id']}** — {len(chosen['values'])} nutrient columns detected.")

                preview_rows = [{"Nutrient": k, "Value from file": v} for k, v in chosen["values"].items()]
                st.dataframe(pd.DataFrame(preview_rows), hide_index=True, use_container_width=True)

                st.warning(
                    "⚠️ **Please verify these values before applying.**  \n\n"
                    "Check that each value matches your lab report for the correct sample. "
                    "**By clicking 'Apply' below, you confirm that you have reviewed these values "
                    "and that they correspond to the correct sample. "
                    "The accuracy of your soil assessment depends entirely on the accuracy of the data entered.**"
                )

                if st.button(
                    "✅ I have reviewed the values — Apply to input fields",
                    type="primary",
                    use_container_width=True,
                    key="csv_apply_btn",
                ):
                    for nname, val in chosen["values"].items():
                        st.session_state[f"nutrient_{nname}"] = val
                    st.success("Fields populated — scroll down to review and adjust if needed.")
                    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — SOIL TEST INPUT
# ─────────────────────────────────────────────────────────────────────────────
with st.expander("🧪 Step 3: Enter Soil Test Results", expanded=True):
    st.markdown(
        "Enter values from your lab report. Leave blank if not measured. "
        "**pH, Organic Matter %, CEC, and Base Saturation %** always use their own fixed units. "
        "For all other sections, select the unit your lab report uses."
    )

    lab_factors = LAB_FACTORS.get(lab, {})

    # ── Section unit selectors ────────────────────────────────────────────
    unit_options = list(UNIT_CONVERSIONS.keys())
    unit_help = (
        "• ppm (mg/kg) — most common for Mehlich III labs\n"
        "• lbs/acre — some labs report exchangeable cations this way\n"
        "• kg/ha — metric equivalent\n\n"
        "The app will automatically convert to ppm before comparing to targets."
    )

    u_col1, u_col2, u_col3 = st.columns(3)
    with u_col1:
        unit_macro = st.selectbox(
            "Macronutrient units (P, K, Ca, Mg, S)",
            unit_options, key="unit_macro", help=unit_help,
        )
    with u_col2:
        unit_micro = st.selectbox(
            "Micronutrient units (Zn, Mn, Fe, Cu, B)",
            unit_options, key="unit_micro", help=unit_help,
        )
    with u_col3:
        unit_salts = st.selectbox(
            "Salts & Other units (Na, Al)",
            unit_options, key="unit_salts", help=unit_help,
        )

    # Map each nutrient to its section unit factor
    section_units = {
        "P (Phosphorus)":    unit_macro,
        "K (Potassium)":     unit_macro,
        "Ca (Calcium)":      unit_macro,
        "Mg (Magnesium)":    unit_macro,
        "S (Sulfur)":        unit_macro,
        "Zn (Zinc)":         unit_micro,
        "Mn (Manganese)":    unit_micro,
        "Fe (Iron)":         unit_micro,
        "Cu (Copper)":       unit_micro,
        "B (Boron)":         unit_micro,
        "Na (Sodium)":       unit_salts,
        "Al (Aluminum)":     unit_salts,
        # EC and base saturation always use their own fixed units — not section-mapped
    }

    # ── Lime rate inputs (buffer pH + tillage) ───────────────────────────
    st.divider()
    st.markdown("**🪨 Lime Rate Inputs** *(optional — required for lime rate calculation)*")
    lime_col1, lime_col2 = st.columns(2)
    with lime_col1:
        buffer_ph_val = st.number_input(
            "Modified Mehlich Buffer pH",
            min_value=4.0, max_value=7.5, value=None,
            step=0.1, format="%.1f",
            key="buffer_ph",
            placeholder="e.g. 6.2",
            help=(
                "Reported on most Cornell Soil Health Lab and Agro-One reports as "
                "'Buffer pH' or 'Mehlich Buffer pH'. This is NOT the same as soil pH. "
                "If your report does not include a buffer pH, leave blank."
            ),
        )
    with lime_col2:
        tillage_key = st.selectbox(
            "Tillage depth",
            list(LIME_TILLAGE_FACTORS.keys()),
            key="tillage_depth",
            help=(
                "Affects how deep the lime needs to be incorporated. "
                "Deeper tillage requires more lime per acre. "
                "Source: Cornell NMSP Lime Calculator v2.0 (2014)."
            ),
        )

    st.divider()

    # ── Per-nutrient help text ────────────────────────────────────────────
    help_texts = {
        "pH":                  "Dimensionless. Look for 'pH' or 'Soil pH' on your report.",
        "Organic Matter":      "Always enter as %. Look for '%OM' or 'Organic Matter %'.",
        "P (Phosphorus)":      "For Modified Morgan labs (Agro-One MM, UVM/UConn/UMass): enter lbs/acre value; auto-converted ×1.8 to Mehlich III equivalents (Cornell NMSP v7). Dairy One / Agro-One Mehlich III: no conversion applied.",
        "K (Potassium)":       "For Modified Morgan labs: enter lbs/acre value; auto-converted ×1.0 to Mehlich III equivalents. Dairy One / Agro-One Mehlich III: no conversion applied.",
        "Ca (Calcium)":        "Dairy One/Agro-One: enter the lbs/acre value and set unit to lbs/acre. This is NOT the same as Base Saturation Ca%.",
        "Mg (Magnesium)":      "Dairy One/Agro-One: enter the lbs/acre value and set unit to lbs/acre.",
        "S (Sulfur)":          "Dairy One/Agro-One: S is not typically reported. Leave blank if not on your report.",
        "Zn (Zinc)":           "Dairy One/Agro-One: reported in lbs/acre — set micronutrient unit to lbs/acre.",
        "Mn (Manganese)":      "Dairy One/Agro-One: reported in lbs/acre — set micronutrient unit to lbs/acre.",
        "Fe (Iron)":           "Dairy One/Agro-One: Fe appears without unit label on the report — enter as ppm (leave unit selector at ppm).",
        "Cu (Copper)":         "Dairy One/Agro-One: Cu not typically reported. Leave blank if not on your report.",
        "B (Boron)":           "Dairy One/Agro-One: B not typically reported. Leave blank if not on your report.",
        "Na (Sodium)":         "Dairy One/Agro-One: Na not typically reported. Leave blank if not on your report.",
        "Al (Aluminum)":       "Dairy One/Agro-One: reported in lbs/acre — set Salts & Other unit to lbs/acre.",
        "CEC":                   "Always enter in meq/100g. Also reported as 'Total Exchange Capacity (M.E.)' or 'T.E.C.' on some lab reports — these are the same thing. CEC is an inherent soil property shown here for context.",
        "EC (Soluble Salts)":    "Enter in dS/m or mS/cm (same value). Some labs report this as 'Soluble Salts' or 'EC'. Most relevant in greenhouse or container production. Leave blank if not on your report.",
        "Base Saturation Ca%":   "Always enter as %. This is the % of CEC occupied by Ca ions — different from Ca in lbs/acre.",
        "Base Saturation K%":    "Always enter as %. This is the % of CEC occupied by K ions — different from K in lbs/acre.",
        "Base Saturation Mg%":   "Always enter as %. This is the % of CEC occupied by Mg ions. Target 10–20%. Leave blank if not on your report.",
    }

    # Override P help text for Logan Labs to warn about P₂O₅
    if "Logan Labs" in lab:
        help_texts["P (Phosphorus)"] = (
            "⚠️ Logan Labs reports P as lbs/acre P₂O₅ (not elemental P). "
            "Multiply the value on your report by 0.437 before entering here. "
            "Example: report shows 200 lbs/ac → enter 87 lbs/ac. "
            "Set the macronutrient unit selector above to lbs/acre."
        )
        help_texts["K (Potassium)"] = (
            "Logan Labs reports K in lbs/acre. Enter the value directly. "
            "Set the macronutrient unit selector above to lbs/acre."
        )
        help_texts["Ca (Calcium)"] = (
            "Logan Labs reports Ca in lbs/acre. Enter the value directly and set unit to lbs/acre."
        )
        help_texts["Mg (Magnesium)"] = (
            "Logan Labs reports Mg in lbs/acre. Enter the value directly and set unit to lbs/acre."
        )

    groups = {
        "Basic Properties": ["pH", "Organic Matter"],
        "Macronutrients":   ["P (Phosphorus)", "K (Potassium)", "Ca (Calcium)", "Mg (Magnesium)", "S (Sulfur)"],
        "Micronutrients":   ["Zn (Zinc)", "Mn (Manganese)", "Fe (Iron)", "Cu (Copper)", "B (Boron)"],
        "Salts & Other":    ["Na (Sodium)", "Al (Aluminum)", "CEC", "EC (Soluble Salts)",
                             "Base Saturation Ca%", "Base Saturation K%", "Base Saturation Mg%"],
    }

    user_values  = {}
    section_unit_map = {}   # nname → unit string (for gap analysis)

    for group_name, nutrient_names in groups.items():
        st.markdown(f"**{group_name}**")
        cols = st.columns(min(len(nutrient_names), 4))
        for i, nname in enumerate(nutrient_names):
            nutrient = next((n for n in NUTRIENTS if n["name"] == nname), None)
            if nutrient is None:
                continue
            col = cols[i % 4]
            with col:
                # Display unit: fixed for non-convertible, section unit for others
                if not nutrient["allow_unit_conversion"]:
                    display_unit = nutrient["unit"]
                else:
                    display_unit = section_units.get(nname, "ppm (mg/kg)")

                section_unit_map[nname] = display_unit
                mm_note = " *(MM→M3)*" if nname in lab_factors else ""

                # Special label for CEC
                label = "CEC / Total Exchange Capacity (M.E.)" if nname == "CEC" else nname

                val = st.number_input(
                    f"{label} [{display_unit}]{mm_note}",
                    min_value=0.0,
                    max_value=50000.0,
                    value=None,
                    step=0.1,
                    format="%.2f",
                    key=f"nutrient_{nname}",
                    placeholder="—",
                    help=help_texts.get(nname, ""),
                )
                user_values[nname] = val
        st.write("")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — RUN ASSESSMENT
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
run_btn = st.button("▶ Run Gap Analysis", type="primary", use_container_width=True)
if run_btn:
    st.session_state.assessment_done = True

if st.session_state.assessment_done:
    st.markdown("## 📊 Gap Analysis Results")
    st.caption(f"Crop: **{crop}** | Lab: **{lab}**")

    lab_factors = LAB_FACTORS.get(lab, {})
    is_mm_lab   = bool(lab_factors)   # True for any Modified Morgan lab

    rows = []
    deficient_nutrients = []
    excess_nutrients    = []

    for n in NUTRIENTS:
        nname = n["name"]
        raw   = user_values.get(nname)

        unit_str = n["unit"] if n["unit"] != "—" else ""

        entered_unit = section_unit_map.get(nname, "ppm (mg/kg)")
        unit_conv    = UNIT_CONVERSIONS.get(entered_unit, 1.0) if n["allow_unit_conversion"] else 1.0
        mm_conv      = lab_factors.get(nname, 1.0)

        # ── Decide which target set to use ───────────────────────────────────
        # When lab is Modified Morgan AND user entered lbs/acre AND MM-specific
        # targets exist for this nutrient, compare directly in lbs/acre —
        # no unit or MM→M3 conversion needed.
        use_mm_direct = (
            is_mm_lab
            and "lbs" in entered_unit
            and n.get("mm_mj_min") is not None
            and n["allow_unit_conversion"]
        )

        if use_mm_direct:
            t_min    = n["mm_hemp_min"] if crop_key == "hemp" else n["mm_mj_min"]
            t_max    = n["mm_hemp_max"] if crop_key == "hemp" else n["mm_mj_max"]
            t_min_m3 = f"{t_min} lbs/ac (MM)"
            t_max_m3 = f"{t_max} lbs/ac (MM)"
        else:
            t_min    = n["hemp_min"] if crop_key == "hemp" else n["mj_min"]
            t_max    = n["hemp_max"] if crop_key == "hemp" else n["mj_max"]
            t_min_m3 = f"{t_min} {unit_str}".strip()
            t_max_m3 = f"{t_max} {unit_str}".strip()

        # Targets converted back to entered units (for M3 path only)
        if not use_mm_direct:
            needs_target_conv = n["allow_unit_conversion"] and (unit_conv != 1.0 or mm_conv != 1.0)
            if needs_target_conv:
                factor = unit_conv * mm_conv
                t_min_eu = round(t_min / factor, 1)
                t_max_eu = round(t_max / factor, 1)
                eu_label = "lbs/ac" if "lbs" in entered_unit else ("kg/ha" if "kg" in entered_unit else "ppm")
                t_min_eu_str = f"{t_min_eu} {eu_label}"
                t_max_eu_str = f"{t_max_eu} {eu_label}"
            else:
                t_min_eu_str = "—"
                t_max_eu_str = "—"
        else:
            t_min_eu_str = "—"   # target already shown in MM lbs/ac column
            t_max_eu_str = "—"

        if raw is None:
            rows.append({
                "Nutrient":                  nname,
                "Value entered":             "—",
                "Conversion → M3 ppm":       "—",
                "Target min":                t_min_m3,
                "Target max":                t_max_m3,
                "Target min (your units)":   t_min_eu_str,
                "Target max (your units)":   t_max_eu_str,
                "Status":                    "— No data",
                "Note":                      n["note"],
            })
            continue

        # ── Has data ─────────────────────────────────────────────────────────
        if use_mm_direct:
            converted  = raw   # compare directly
            show_conv  = "Direct MM lbs/acre comparison (no conversion)"
        else:
            converted = round(raw * unit_conv * mm_conv, 2)
            needs_conv = (unit_conv != 1.0 or mm_conv != 1.0)
            if needs_conv:
                eq = str(raw)
                if unit_conv != 1.0:
                    eu_abbrev = "lbs/ac" if "lbs" in entered_unit else ("kg/ha" if "kg" in entered_unit else entered_unit)
                    eq += f" × {unit_conv} ({eu_abbrev}→ppm)"
                if mm_conv != 1.0:
                    eq += f" × {mm_conv} (MM→M3)"
                show_conv = f"{eq} = {converted} ppm"
            else:
                show_conv = "—"

        is_info = n.get("is_informational", False)
        if converted < t_min:
            if is_info:
                status = "ℹ Below typical range"
            else:
                status = "⚠ DEFICIENT"
                deficient_nutrients.append(nname)
        elif converted > t_max:
            if is_info:
                status = "ℹ Above typical range"
            else:
                status = "▲ EXCESS"
                excess_nutrients.append(nname)
        else:
            status = "✓ ADEQUATE" if not is_info else "ℹ Within typical range"

        rows.append({
            "Nutrient":                  nname,
            "Value entered":             raw,
            "Conversion → M3 ppm":       show_conv,
            "Target min":                t_min_m3,
            "Target max":                t_max_m3,
            "Target min (your units)":   t_min_eu_str,
            "Target max (your units)":   t_max_eu_str,
            "Status":                    status,
            "Note":                      n["note"],
        })

    df = pd.DataFrame(rows)

    # Store anonymized gap data for the optional data-sharing block below
    st.session_state["_soil_share"] = {
        "crop":      crop,
        "lab":       lab,
        "user_values": {k: v for k, v in user_values.items() if v is not None},
        "deficient": list(deficient_nutrients),
        "excess":    list(excess_nutrients),
    }

    def style_status(val):
        if "DEFICIENT" in str(val): return "background-color: #ffd6d6; color: #8b0000; font-weight: bold"
        if "EXCESS"    in str(val): return "background-color: #fff3cd; color: #856404; font-weight: bold"
        if "ADEQUATE"  in str(val): return "background-color: #d6f0d6; color: #006400; font-weight: bold"
        if "ℹ"         in str(val): return "background-color: #e8f4fd; color: #1565C0; font-style: italic"
        return "color: #888; font-style: italic"

    styled = df.style.map(style_status, subset=["Status"])
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # ── Contextual alerts based on pH and M3 extraction caveats ──────────────
    ph_val = user_values.get("pH")

    if ph_val and ph_val > 7.0:
        st.warning(
            f"⚠️ **High pH alert (pH {ph_val:.1f}):** Above pH 7.0, Mehlich III extraction "
            "may report adequate or excess Fe, Mn, Zn, and B even when these nutrients are not "
            "plant-available. **Address pH first** — lower it to 6.2–6.8 with elemental sulfur "
            "or acidifying fertilizers before acting on micronutrient status. "
            "Gypsum (calcium sulfate) is a useful option to add S and Ca without raising pH further."
        )

    _fe_excess = "Fe (Iron)" in excess_nutrients
    _mn_excess = "Mn (Manganese)" in excess_nutrients
    if (_fe_excess or _mn_excess) and (ph_val is None or ph_val < 7.5):
        _flagged = " and ".join(
            n for n, f in [("Fe", _fe_excess), ("Mn", _mn_excess)] if f
        )
        st.info(
            f"ℹ️ **{_flagged} flagged as EXCESS:** Mehlich III routinely over-extracts Fe and Mn "
            "from mineral soils — values above the target range are common in healthy soils and "
            "often do not indicate actual toxicity. Before acting, check for visual symptoms "
            "(Fe: leaf bronzing or dark spots; Mn: necrotic spots between veins). "
            "If pH is in the 6.2–6.8 range and plants look healthy, no correction is typically needed."
        )

    # ── Download buttons ─────────────────────────────────────────────────────
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    def _build_report_xlsx(gap_df, deficient_list, excess_list, crop_label, lab_label):
        wb = openpyxl.Workbook()

        # ── Sheet 1: Gap Analysis ────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Gap Analysis"
        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF")
        def_fill  = PatternFill("solid", fgColor="FFD6D6")
        exc_fill  = PatternFill("solid", fgColor="FFF3CD")
        ok_fill   = PatternFill("solid", fgColor="D6F0D6")

        cols = list(gap_df.columns)
        for ci, h in enumerate(cols, 1):
            c = ws1.cell(1, ci, h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(wrap_text=True)

        for ri, row in gap_df.iterrows():
            status = str(row.get("Status", ""))
            row_fill = def_fill if "DEFICIENT" in status else (
                       exc_fill if "EXCESS" in status else (
                       ok_fill  if "ADEQUATE" in status else None))
            for ci, col in enumerate(cols, 1):
                c = ws1.cell(ri + 2, ci, str(row[col]) if row[col] != "—" else "")
                if row_fill:
                    c.fill = row_fill
                c.alignment = Alignment(wrap_text=True)

        ws1.column_dimensions["A"].width = 22
        for letter in ["B","C","D","E","F","G"]:
            ws1.column_dimensions[letter].width = 26
        ws1.column_dimensions["H"].width = 40

        # ── Sheet 2: Amendment Recommendations ──────────────────────────────
        ws2 = wb.create_sheet("Amendments")
        amend_headers = ["Deficiency Addressed", "Amendment", "Type", "Form",
                         "How to Apply", "Typical Rate", "Notes",
                         "Est. Price (low)", "Est. Price (high)", "Price Unit"]
        for ci, h in enumerate(amend_headers, 1):
            c = ws2.cell(1, ci, h)
            c.font = header_font
            c.fill = header_fill

        ri2 = 2
        seen_a = set()
        for nname in deficient_list:
            short = nname.split("(")[0].strip().lower()
            for a in AMENDMENTS:
                cond_lower = a["condition"].lower()
                if (short in cond_lower or nname.lower() in cond_lower) and ">" not in a["condition"]:
                    key_a = a["amendment"]
                    if key_a in seen_a:
                        continue
                    seen_a.add(key_a)
                    row_vals = [
                        nname,
                        a["amendment"],
                        "Organic / OMRI" if a.get("organic") else "Conventional",
                        a.get("form", ""),
                        a.get("application", ""),
                        a.get("rate", ""),
                        a.get("notes", ""),
                        a.get("price_low", ""),
                        a.get("price_high", ""),
                        a.get("price_unit", ""),
                    ]
                    for ci, v in enumerate(row_vals, 1):
                        ws2.cell(ri2, ci, v)
                    ri2 += 1

        for excess_n in excess_list:
            action = QUICK_AMEND.get(excess_n, {}).get("high", "Reduce inputs; re-test in 60–90 days")
            row_vals = [f"{excess_n} (EXCESS)", "—", "—", "—", action, "—", "—", "", "", ""]
            for ci, v in enumerate(row_vals, 1):
                ws2.cell(ri2, ci, v)
            ri2 += 1

        if ri2 == 2:
            ws2.cell(2, 1, "No amendments indicated — all nutrients within target range.")

        for letter, w in zip(["A","B","C","D","E","F","G","H","I","J"],
                              [24, 28, 18, 14, 36, 28, 40, 14, 14, 14]):
            ws2.column_dimensions[letter].width = w

        # ── Sheet 3: Summary ─────────────────────────────────────────────────
        ws3 = wb.create_sheet("Summary")
        import datetime
        summary_rows = [
            ("Report generated", datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Crop type", crop_label),
            ("Laboratory", lab_label),
            ("Deficient nutrients", ", ".join(deficient_list) if deficient_list else "None"),
            ("Excess nutrients",    ", ".join(excess_list)    if excess_list    else "None"),
            ("Nutrients in range",
             ", ".join(r["Nutrient"] for _, r in gap_df.iterrows()
                       if "ADEQUATE" in str(r.get("Status","")))),
        ]
        for ri3, (k, v) in enumerate(summary_rows, 1):
            ws3.cell(ri3, 1, k).font = Font(bold=True)
            ws3.cell(ri3, 2, v)
        ws3.column_dimensions["A"].width = 26
        ws3.column_dimensions["B"].width = 60

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    _report_bytes = _build_report_xlsx(df, deficient_nutrients, excess_nutrients, crop, lab)

    dl_c1, dl_c2 = st.columns(2)
    with dl_c1:
        st.download_button(
            "⬇ Download Full Report (Excel)",
            data=_report_bytes,
            file_name="soil_assessment_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_c2:
        st.download_button(
            "⬇ Download Gap Analysis (CSV)",
            data=df.to_csv(index=False),
            file_name="soil_gap_analysis.csv",
            mime="text/csv",
            use_container_width=True,
        )

    # ── Possible Amendments ────────────────────────────────────────────────
    st.divider()
    st.markdown("## 🌾 Possible Amendments")
    st.caption(
        "The following are **possible options** based on the deficiencies and excesses identified above. "
        "These are not prescriptions. Always verify rates and products with a certified crop advisor "
        "before purchasing or applying anything. "
        "**Prices are estimates based on 2025–2026 US market data and will vary by supplier, region, "
        "quantity, and season — treat as rough budgeting guidance only.**"
    )

    if not deficient_nutrients and not excess_nutrients:
        st.success("🎉 All entered nutrients are within target range — no amendments indicated based on this data.")
    else:
        def form_tag(form_str):
            f = form_str.lower()
            if "liquid"   in f: cls, label = "tag-liquid",   "💧 Liquid"
            elif "pellet" in f: cls, label = "tag-pellet",   "🔵 Pellet"
            elif "granul" in f or "prill" in f: cls, label = "tag-granular", "🟡 Granular"
            else:               cls, label = "tag-powder",   "⚪ Powder"
            return f'<span class="tag {cls}">{label}</span>'

        def organic_tag(is_organic):
            if is_organic:
                return '<span class="tag tag-organic">🌿 Organic / OMRI</span>'
            return '<span class="tag tag-synthetic">🔬 Conventional</span>'

        if deficient_nutrients:
            st.markdown(f"### ⚠ Addressing Deficiencies: {', '.join(deficient_nutrients)}")
            shown = set()
            for nname in deficient_nutrients:
                short = nname.split("(")[0].strip().lower()
                for a in AMENDMENTS:
                    cond_lower = a["condition"].lower()
                    if short in cond_lower or nname.lower() in cond_lower:
                        # Skip amendments designed for high/excess values (e.g. "pH > 7.0")
                        if ">" in a["condition"]:
                            continue
                        key = a["amendment"]
                        if key in shown:
                            continue
                        shown.add(key)
                        price_str = (
                            f"<b>💲 Estimated price:</b> "
                            f"${a['price_low']:.2f}–${a['price_high']:.2f} {a['price_unit']}"
                            f"<br><i style='font-size:0.85rem;color:#666'>{a['price_note']}</i>"
                        ) if "price_low" in a else ""
                        st.markdown(f"""
<div class="amend-card">
<h4>{a['amendment']}</h4>
{organic_tag(a['organic'])} {form_tag(a['form'])}
<br><br>
<b>Addresses:</b> {a['condition']}<br>
<b>How to apply:</b> {a['application']}<br>
<b>Typical rate:</b> {a['rate']}<br>
<b>Notes:</b> {a['notes']}<br><br>
{price_str}
</div>""", unsafe_allow_html=True)

        # ── Lime rate calculator (shown when pH is deficient) ─────────────
        if "pH" in deficient_nutrients:
            st.divider()
            st.markdown("### 🪨 Lime Rate Estimate")
            _bph = st.session_state.get("buffer_ph", None)
            _till = st.session_state.get("tillage_depth", list(LIME_TILLAGE_FACTORS.keys())[0])
            _soil_ph = user_values.get("pH", None)

            if _bph is None or _bph == 0.0:
                st.info(
                    "Enter a **Modified Mehlich Buffer pH** in the Lime Rate Inputs section above "
                    "to get a site-specific lime rate recommendation."
                )
            else:
                _rate_100 = lime_rate_lookup(_bph, _till, target_min_ph=6.4)
                _rate_80  = round(_rate_100 / 0.80, 2)   # typical ag lime ~80% ENV
                _rate_90  = round(_rate_100 / 0.90, 2)   # high-quality lime ~90% ENV

                st.markdown(f"""
<div style="background:#e8f5e9; border-left:5px solid #2e7d32; border-radius:8px;
            padding:18px 22px; margin-bottom:12px;">
<h4 style="margin:0 0 10px 0; color:#1b5e20;">🌿 Lime Rate — Cannabis / Hemp (target pH ≥ 6.4)</h4>
<p style="margin:4px 0"><b>Buffer pH entered:</b> {_bph:.1f} &nbsp;|&nbsp;
   <b>Tillage depth:</b> {_till} &nbsp;|&nbsp;
   <b>Tillage factor:</b> ×{LIME_TILLAGE_FACTORS[_till]}</p>
<hr style="border:none; border-top:1px solid #a5d6a7; margin:10px 0">
<table style="width:100%; border-collapse:collapse; font-size:0.95rem;">
<tr><th style="text-align:left; padding:4px 8px; background:#c8e6c9">Lime product quality</th>
    <th style="text-align:center; padding:4px 8px; background:#c8e6c9">%ENV</th>
    <th style="text-align:center; padding:4px 8px; background:#c8e6c9">Rate (tons/acre)</th></tr>
<tr><td style="padding:4px 8px">Reference rate (100% ENV)</td>
    <td style="text-align:center; padding:4px 8px">100%</td>
    <td style="text-align:center; padding:4px 8px"><b>{_rate_100:.2f}</b></td></tr>
<tr style="background:#f1f8e9"><td style="padding:4px 8px">High-quality ag lime (typical)</td>
    <td style="text-align:center; padding:4px 8px">90%</td>
    <td style="text-align:center; padding:4px 8px"><b>{_rate_90:.2f}</b></td></tr>
<tr><td style="padding:4px 8px">Standard ag lime (typical)</td>
    <td style="text-align:center; padding:4px 8px">80%</td>
    <td style="text-align:center; padding:4px 8px"><b>{_rate_80:.2f}</b></td></tr>
</table>
<p style="margin:10px 0 4px 0; font-size:0.85rem; color:#2e7d32">
<b>Note:</b> Actual application rate = table rate ÷ (%ENV / 100). Check the bag or supplier sheet
for your lime product's %ENV (Effective Neutralizing Value). Apply at least <b>3–6 months before
planting</b> and incorporate thoroughly. For large applications (&gt;3 tons/acre), split into two
applications one season apart. <br>
<b>Source:</b> Cornell NMSP Lime Guidelines Calculator v2.0 (March 2014) — target minimum pH 6.4,
using Modified Mehlich buffer pH method.
</p>
</div>""", unsafe_allow_html=True)

                # Acres input for total lime needed
                _acres_lime = st.session_state.get("soil_amendment_acres", 1.0)
                st.caption(
                    f"At {_acres_lime:.1f} acres: "
                    f"**{round(_rate_90 * _acres_lime, 1)} tons** (90% ENV) or "
                    f"**{round(_rate_80 * _acres_lime, 1)} tons** (80% ENV) total. "
                    f"Adjust acreage in the Amendment Budget section below."
                )

        if excess_nutrients:
            st.markdown(f"### ▲ Addressing Excess Levels: {', '.join(excess_nutrients)}")
            for nname in excess_nutrients:
                action = QUICK_AMEND.get(nname, {}).get("high", "Reduce inputs; re-test in 60–90 days")
                st.markdown(f"""
<div class="amend-card" style="border-left: 3px solid #e67e22;">
<h4>{nname} — Excess</h4>
<b>Suggested action:</b> {action}
</div>""", unsafe_allow_html=True)

    # ── Soil health context ───────────────────────────────────────────────
    with st.expander("💡 Soil Health Context & General Guidelines"):
        st.markdown("""
**pH** is the most important lever — it controls availability of nearly every nutrient.
Target **6.2–6.8** for cannabis and hemp. At pH < 6.0, Al and Mn become toxic;
at pH > 7.2, Fe, Zn, Mn, and B become unavailable.

**Organic matter** feeds the soil microbiome, buffers nutrients, and improves water retention.
A minimum of 3% OM is recommended; 5–8% is ideal for NY outdoor cultivation.

**Ca:Mg ratio** should be approximately 5:1 to 8:1 by weight (ppm).
High Mg relative to Ca can cause compaction in the silt loam soils common in NY.

**K:Mg ratio** should stay below 3:1.
High K suppresses Mg uptake — a common cause of Mg deficiency in NY cannabis fields.

**Micronutrients** (Zn, Mn, Fe, Cu, B) are primarily affected by pH.
Foliar applications are the fastest in-season correction.
Note: Mehlich III tends to over-extract Fe and Mn from mineral soils — "excess" results for
these two nutrients are common and do not automatically indicate a toxicity problem.
Always confirm with plant tissue symptoms before correcting.

---
**Lab-specific notes:**

**Logan Labs:** Phosphorus is reported as **lbs/acre P₂O₅**, not elemental P.
Multiply the P value on your Logan Labs report by **0.437** before entering it in this tool
(e.g. 200 lbs/ac P₂O₅ → enter 87 lbs/ac). Entering P₂O₅ directly will overstate phosphorus
by ~2.3×, which can flip a correct DEFICIENT call to ADEQUATE or ADEQUATE to EXCESS.

**K% base saturation targets:** This tool uses a 2–5% target based on cannabis-calibrated
Mehlich III data (Cornell NMSP). Some independent agronomists and labs use 4–6% or 6–8%
for specialty crops. If your lab report says K% is "low" at 4–5%, that is not necessarily a
disagreement — it reflects a different calibration standard, not an error.

For questions about NY-specific hemp or cannabis licensing and agronomy,
contact your local Cornell Cooperative Extension office.
""")

    # ── Amendment Budget Estimator ─────────────────────────────────────────
    if deficient_nutrients:
        st.divider()
        st.markdown("## 💰 Amendment Budget Estimator")
        st.caption(
            "Rough cost estimate for the amendments indicated above. "
            "Enter your field size to calculate total estimated cost. "
            "This estimate flows automatically into the Economics Tool."
        )

        col_a, _ = st.columns([1, 3])
        with col_a:
            acres_est = st.number_input(
                "Field size (acres)",
                min_value=0.01, max_value=10000.0, value=1.0, step=0.25,
                key="budget_acres",
                help="Total cultivated acreage for this field or season"
            )

        # Build cost table: one row per deficient nutrient, first matching amendment
        budget_rows = []
        total_low = 0.0
        total_high = 0.0
        seen_amend = set()

        for nname in deficient_nutrients:
            short = nname.split("(")[0].strip().lower()
            for a in AMENDMENTS:
                cond_lower = a["condition"].lower()
                if (short in cond_lower or nname.lower() in cond_lower) and a.get("cost_acre_low", 0) > 0:
                    key_a = a["amendment"]
                    if key_a in seen_amend:
                        continue
                    seen_amend.add(key_a)
                    low  = round(a["cost_acre_low"]  * acres_est, 2)
                    high = round(a["cost_acre_high"] * acres_est, 2)
                    total_low  += low
                    total_high += high
                    budget_rows.append({
                        "Deficiency":  nname,
                        "Amendment":   a["amendment"],
                        "$/acre (low)": f"${a['cost_acre_low']:.0f}",
                        "$/acre (high)":f"${a['cost_acre_high']:.0f}",
                        f"Est. cost ({acres_est:.2f} ac) low":  f"${low:,.0f}",
                        f"Est. cost ({acres_est:.2f} ac) high": f"${high:,.0f}",
                    })
                    break

        if budget_rows:
            budget_mid = (total_low + total_high) / 2
            st.dataframe(pd.DataFrame(budget_rows), use_container_width=True, hide_index=True)

            c1, c2, c3 = st.columns(3)
            c1.metric("Estimated Low", f"${total_low:,.0f}")
            c2.metric("Estimated Mid", f"${budget_mid:,.0f}")
            c3.metric("Estimated High", f"${total_high:,.0f}")

            # Store in session state for economics tool
            st.session_state["soil_amendment_cost_low"] = total_low
            st.session_state["soil_amendment_cost_high"] = total_high
            st.session_state["soil_amendment_cost_mid"] = budget_mid
            st.session_state["soil_amendment_acres"] = acres_est

            # Ask user if they want to carry amendment estimates to Economics Tool
            _carry_done = st.session_state.get("_soil_carry_done", False)
            _carry_yes  = st.session_state.get("_soil_carry_econ", False)

            if not _carry_done:
                st.markdown("**📊 Would you like to carry these amendment estimates into the Economics Tool?**")
                st.caption("Your amendment cost will be pre-filled under Variable Costs in the Economics Tool.")
                _cc1, _cc2, _ = st.columns([1, 1, 4])
                with _cc1:
                    if st.button("Yes, carry to Economics", key="_carry_econ_yes",
                                 type="primary", use_container_width=True):
                        st.session_state["_soil_carry_done"] = True
                        st.session_state["_soil_carry_econ"] = True
                        st.rerun()
                with _cc2:
                    if st.button("No thanks", key="_carry_econ_no", use_container_width=True):
                        for _k in ["soil_amendment_cost_low", "soil_amendment_cost_high",
                                   "soil_amendment_cost_mid", "soil_amendment_acres"]:
                            st.session_state.pop(_k, None)
                        st.session_state["_soil_carry_done"] = True
                        st.session_state["_soil_carry_econ"] = False
                        st.rerun()
            elif _carry_yes:
                st.success(
                    "✅ Amendment estimates will be pre-filled in the Economics Tool under Variable Costs.",
                    icon="📊"
                )
            else:
                st.info("Amendment estimates were not carried to the Economics Tool.", icon="👍")
        else:
            st.caption("No per-acre cost data available for the detected deficiencies.")

if st.session_state.assessment_done and "_soil_share" in st.session_state:
    from utils.data_share import render_share_block
    _sd = st.session_state["_soil_share"]
    _uv = _sd["user_values"]
    _SOIL_COLS = [
        "timestamp", "county", "crop_type", "lab_type",
        "pH", "OM_pct", "P_entered", "K_entered", "Ca_entered", "Mg_entered",
        "S_entered", "Zn_entered", "Mn_entered", "Fe_entered", "Cu_entered",
        "B_entered", "Na_entered", "Al_entered", "CEC", "Base_Sat_Ca_pct", "Base_Sat_K_pct",
        "deficient_nutrients", "excess_nutrients",
    ]
    def _soil_row_builder(county):
        return [{
            "county":              county,
            "crop_type":           _sd["crop"],
            "lab_type":            _sd["lab"],
            "pH":                  _uv.get("pH", ""),
            "OM_pct":              _uv.get("Organic Matter", ""),
            "P_entered":           _uv.get("P (Phosphorus)", ""),
            "K_entered":           _uv.get("K (Potassium)", ""),
            "Ca_entered":          _uv.get("Ca (Calcium)", ""),
            "Mg_entered":          _uv.get("Mg (Magnesium)", ""),
            "S_entered":           _uv.get("S (Sulfur)", ""),
            "Zn_entered":          _uv.get("Zn (Zinc)", ""),
            "Mn_entered":          _uv.get("Mn (Manganese)", ""),
            "Fe_entered":          _uv.get("Fe (Iron)", ""),
            "Cu_entered":          _uv.get("Cu (Copper)", ""),
            "B_entered":           _uv.get("B (Boron)", ""),
            "Na_entered":          _uv.get("Na (Sodium)", ""),
            "Al_entered":          _uv.get("Al (Aluminum)", ""),
            "CEC":                 _uv.get("CEC", ""),
            "Base_Sat_Ca_pct":     _uv.get("Base Saturation Ca%", ""),
            "Base_Sat_K_pct":      _uv.get("Base Saturation K%", ""),
            "deficient_nutrients": ", ".join(_sd["deficient"]),
            "excess_nutrients":    ", ".join(_sd["excess"]),
        }]
    render_share_block("soil", "Soil Data", _SOIL_COLS, _soil_row_builder, county_widget=True)

st.divider()
st.markdown("## How This Tool Works")
steps = st.columns(4)
step_data = [
    ("1️⃣", "Enter Address", "Type your farm or field address. The tool geocodes it and queries the USDA NRCS SSURGO database for your soil series, texture, drainage class, and baseline pH."),
    ("2️⃣", "Select Crop & Lab", "Choose Hemp or Adult-Use Cannabis and select your soil laboratory. Modified Morgan values are automatically converted to Mehlich III equivalents."),
    ("3️⃣", "Enter Lab Results", "Type in values from your soil test report — pH, OM, macro- and micronutrients. Leave fields blank if not measured."),
    ("4️⃣", "Get Recommendations", "Instant color-coded gap analysis (Deficient / Adequate / Excess) plus specific amendment products, rates, and timing for NY conditions."),
]
for col, (icon, title, desc) in zip(steps, step_data):
    with col:
        st.markdown(f"**{icon} {title}**")
        st.caption(desc)

st.caption(
    "Built for NYS licensed cultivators · "
    "Soil data: USDA NRCS SSURGO · Geocoding: US Census Geocoder · "
    "Nutrient targets: NY State Extension / CCE frameworks · "
    "Lab conversions (MM→M3): Cornell NMSP Conversion Tools (nmsp.cals.cornell.edu/software/conv-tools.html) · "
    "Lime rates: Cornell NMSP Lime Guidelines Calculator v2.0 (2014) · "
    "Amendment rates: representative ranges only — consult a CCA."
)

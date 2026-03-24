"""
3_Pre_Harvest.py — NY Cannabis/Hemp Pre-Harvest Data Collection Tool
Upload existing PreHarvest Excel template OR enter data manually.
Download results as CSV or Excel.
"""

import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

st.set_page_config(
    page_title="Pre-Harvest | NYS Cannabis Tool",
    page_icon="🌱",
    layout="wide",
)

st.markdown("""
<style>
.disclaimer-box {
    background: #fff3cd; border: 2px solid #e0a800;
    border-radius: 8px; padding: 16px 20px; margin-bottom: 20px; font-size: 0.92rem;
}
.batch-badge {
    display: inline-block; padding: 3px 12px; border-radius: 12px;
    background: #d4edda; color: #155724; font-weight: bold; font-size: 0.9rem;
}
</style>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
if "preharvest_batches" not in st.session_state:
    st.session_state.preharvest_batches = {}

# ── Upload parser ──────────────────────────────────────────────────────────────

def _read_section(ws, section_row):
    """Read column headers (section_row+1) and data rows (section_row+3 to section_row+24)."""
    header_r = section_row + 1
    data_start = section_row + 3
    data_end = section_row + 24
    headers = [ws.cell(header_r, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h is not None else f"_col{c}" for c, h in enumerate(headers, 1)]
    records = []
    for r in range(data_start, data_end + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if any(v is not None for v in vals):
            records.append(dict(zip(headers, vals)))
    return records


def _read_log_sheet(ws, header_row, data_start_row):
    """Read log-style sheets where headers are on a fixed row."""
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h is not None else f"_col{c}" for c, h in enumerate(headers, 1)]
    records = []
    for r in range(data_start_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if any(v is not None for v in vals):
            records.append(dict(zip(headers, vals)))
    return records


def parse_preharvest_excel(file_bytes):
    """Parse a PreHarvest_GrowingAndSelling_Template.xlsx and return list of batch dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    batches = {}  # batch_number -> aggregated data

    # Tab 1: Aggregate — 6 sections at rows 4, 29, 54, 79, 104, 129 linked by row position
    if "1_PreHarvest_Aggregate" in wb.sheetnames:
        ws = wb["1_PreHarvest_Aggregate"]
        sections = [_read_section(ws, r) for r in [4, 29, 54, 79, 104, 129]]
        max_len = max((len(s) for s in sections), default=0)
        for i in range(max_len):
            merged = {}
            for s in sections:
                if i < len(s):
                    merged.update(s[i])
            bn = str(merged.get("Batch Number") or merged.get("Farm Name") or f"Batch_{i+1}")
            batches.setdefault(bn, {})["aggregate"] = merged
            batches[bn]["key"] = bn

    # Tab 2: Individual Plant Data — sections at rows 4, 29
    if "2_Individual_Plant_Data" in wb.sheetnames:
        ws = wb["2_Individual_Plant_Data"]
        plant_rows = _read_section(ws, 4) + _read_section(ws, 29)
        for rec in plant_rows:
            bn = str(rec.get("Batch Number", ""))
            if bn in batches:
                batches[bn].setdefault("plants", []).append(rec)

    # Tabs 3–6: Log-style (header row 4, data from row 6)
    for tab, key in [("3_Inputs_DoYouUse", "inputs"), ("4_EnergyAudit_Indoor", "energy"),
                     ("5_NutrientAmendment_Log", "nutrients"), ("6_PestControl_Log", "pests")]:
        if tab in wb.sheetnames:
            recs = _read_log_sheet(wb[tab], header_row=4, data_start_row=6)
            for rec in recs:
                bn = str(rec.get("Batch Number", ""))
                if bn in batches:
                    batches[bn].setdefault(key, []).append(rec)

    # Tab 7: Yield & Selling (header row 5, data from row 7)
    if "7_YieldAndSelling" in wb.sheetnames:
        recs = _read_log_sheet(wb["7_YieldAndSelling"], header_row=5, data_start_row=7)
        for rec in recs:
            bn = str(rec.get("Batch Number", ""))
            if bn in batches:
                batches[bn]["yield_selling"] = rec

    return list(batches.values())


# ── Excel download builder ─────────────────────────────────────────────────────

def build_preharvest_excel(batches_dict):
    wb = Workbook()

    # Sheet 1: Batch Aggregate
    ws1 = wb.active
    ws1.title = "1_Batch_Aggregate"
    ws1.append([
        "Batch Number", "Farm Name", "Strain Name", "Seed Source", "Flowering Strategy",
        "Season / Cycle", "Cultivated Area (sqM)", "Growing Media", "Soil Type",
        "Field Aspect", "Field Steepness",
        "# Plants Outdoors", "# Plants Hoop House", "# Plants Greenhouse", "# Plants Indoors",
        "Germination Rate (%)", "First Flower Date", "Days in Ground",
        "Disease Presence", "Disease Name", "Disease Severity", "Disease Date First Seen",
        "% White Trichomes", "% Clear Trichomes", "% Amber Trichomes", "Biomass Weight (kg)",
        "THC (%)", "CBD (%)", "Other Cannabinoids", "B-Myrcene (%)", "B-Caryophyllene (%)",
    ])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws1.append([
            bn, d.get("farm_name", ""), d.get("strain_name", ""), d.get("seed_source", ""),
            d.get("flowering_strategy", ""), d.get("season_cycle", ""),
            d.get("cultivated_area_sqm") or "", d.get("growing_media", ""), d.get("soil_type", ""),
            d.get("field_aspect", ""), d.get("field_steepness", ""),
            d.get("n_plants_outdoor") or 0, d.get("n_plants_hoop") or 0,
            d.get("n_plants_greenhouse") or 0, d.get("n_plants_indoor") or 0,
            d.get("germination_rate") or "", d.get("first_flower_date", ""),
            d.get("days_in_ground") or "",
            d.get("disease_presence", ""), d.get("disease_name", ""),
            d.get("disease_severity", ""), d.get("disease_date", ""),
            d.get("pct_white_trichomes") or "", d.get("pct_clear_trichomes") or "",
            d.get("pct_amber_trichomes") or "", d.get("biomass_weight_kg") or "",
            d.get("thc_pct") or "", d.get("cbd_pct") or "",
            d.get("other_cannabinoids", ""),
            d.get("b_myrcene_pct") or "", d.get("b_caryophyllene_pct") or "",
        ])

    # Sheet 2: Individual Plants
    ws2 = wb.create_sheet("2_Individual_Plants")
    plant_header = [
        "Batch Number", "Plant Number", "Plant ID / Tag", "Environment", "Growing Media",
        "Seed Planting Date", "Transplanting Date", "Topped (Y/N)", "Pruned (Y/N)",
        "Trellised (Y/N)", "Height at Harvest (cm)", "Stem Width (mm)", "Node Count", "Bud Size",
    ]
    ws2.append(plant_header)
    for bn, b in batches_dict.items():
        df = b.get("plants_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws2.append([bn] + list(row))

    # Sheet 3: Inputs Survey
    ws3 = wb.create_sheet("3_Inputs_Survey")
    ws3.append([
        "Batch Number", "Mineral Soil", "Blood Meal", "Green Sand", "Feather Meal",
        "Poultry Manure", "Livestock Manure", "Compost/Worm Cast", "Biological",
        "Regular Compost", "Plasticulture", "Mulch/Cover Crop",
    ])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws3.append([
            bn,
            1 if d.get("input_mineral_soil") else 0,
            1 if d.get("input_blood_meal") else 0,
            1 if d.get("input_greensand") else 0,
            1 if d.get("input_feather_meal") else 0,
            1 if d.get("input_poultry_manure") else 0,
            1 if d.get("input_livestock_manure") else 0,
            1 if d.get("input_compost_worm") else 0,
            1 if d.get("input_biological") else 0,
            1 if d.get("input_regular_compost") else 0,
            1 if d.get("input_plasticulture") else 0,
            1 if d.get("input_mulch") else 0,
        ])

    # Sheet 4: Energy Audit
    ws4 = wb.create_sheet("4_Energy_Audit")
    ws4.append([
        "Batch Number", "Grow Room / Zone", "Lighting Type", "Lighting Wattage (W)",
        "Light Hours/Day", "Weeks Veg", "Weeks Flower", "HVAC System Type",
        "HVAC kWh/Month", "CO2 Supplementation", "CO2 Source", "Irrigation System",
        "Est. Lighting kWh/Cycle", "Est. Total kWh/Cycle",
    ])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        if d.get("grow_room") or d.get("lighting_type"):
            w = d.get("lighting_wattage_w") or 0
            h = d.get("light_hours_per_day") or 0
            days = ((d.get("weeks_veg") or 0) + (d.get("weeks_flower") or 0)) * 7
            l_kwh = round(w * h * days / 1000, 1)
            t_kwh = round(l_kwh + (d.get("hvac_kwh_month") or 0) * days / 30, 1)
            ws4.append([
                bn, d.get("grow_room", ""), d.get("lighting_type", ""),
                d.get("lighting_wattage_w") or "", d.get("light_hours_per_day") or "",
                d.get("weeks_veg") or "", d.get("weeks_flower") or "",
                d.get("hvac_type", ""), d.get("hvac_kwh_month") or "",
                d.get("co2_supplementation", ""), d.get("co2_source", ""),
                d.get("irrigation_system", ""), l_kwh, t_kwh,
            ])

    # Sheet 5: Nutrient Log
    ws5 = wb.create_sheet("5_Nutrient_Log")
    ws5.append([
        "Batch Number", "Application Date", "Crop Stage", "Product / Amendment Name",
        "Type", "NPK / Analysis", "Rate Applied", "Units", "Application Method", "Notes",
    ])
    for bn, b in batches_dict.items():
        df = b.get("nutrients_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws5.append([bn] + list(row))

    # Sheet 6: Pest Control Log
    ws6 = wb.create_sheet("6_Pest_Control_Log")
    ws6.append([
        "Batch Number", "Date", "Event Type", "Target Pest / Disease",
        "Product / Method Name", "Type", "EPA Reg. # / Active Ingredient",
        "Rate Applied", "Application Method", "PHI (days)", "Applied By", "Notes",
    ])
    for bn, b in batches_dict.items():
        df = b.get("pests_df")
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                ws6.append([bn] + list(row))

    # Sheet 7: Yield & Selling
    ws7 = wb.create_sheet("7_Yield_Selling")
    ws7.append([
        "Batch Number", "Strain Name", "Season / Cycle",
        "Flower Sold (g)", "Flower Sale Date", "Flower Price ($/g)", "Flower Revenue ($)",
        "Pre-Rolls Sold (units)", "Pre-Roll Sale Date", "Pre-Roll Price ($/unit)", "Pre-Roll Revenue ($)",
        "Biomass Sold (g)", "Biomass Sale Date", "Biomass Price ($/kg)", "Biomass Revenue ($)",
    ])
    for bn, b in batches_dict.items():
        d = b.get("data", {})
        ws7.append([
            bn, d.get("strain_name", ""), d.get("season_cycle", ""),
            d.get("flower_sold_g") or "", d.get("flower_sale_date", ""),
            d.get("flower_price_per_g") or "", d.get("flower_revenue") or "",
            d.get("preroll_sold_units") or "", d.get("preroll_sale_date", ""),
            d.get("preroll_price_per_unit") or "", d.get("preroll_revenue") or "",
            d.get("biomass_sold_g") or "", d.get("biomass_sale_date", ""),
            d.get("biomass_price_per_kg") or "", d.get("biomass_revenue") or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Page layout ────────────────────────────────────────────────────────────────

st.markdown("""
<div class="disclaimer-box">
<b>Record-keeping tool only.</b> Not a compliance filing system. Always maintain records consistent
with NYS OCM requirements. Data is stored in your browser session — download before closing.
</div>
""", unsafe_allow_html=True)

st.title("🌱 Pre-Harvest Data Collection")
st.caption("Batch identity, growing environment, inputs, pest control, and yield/selling — NYS cannabis & hemp")
st.divider()

# ── Upload ─────────────────────────────────────────────────────────────────────
with st.expander("📤 Upload Existing PreHarvest Excel Template", expanded=False):
    st.markdown(
        "Upload a filled `PreHarvest_GrowingAndSelling_Template.xlsx` to import existing data. "
        "Imported batches are added to the current session."
    )
    if not HAS_OPENPYXL:
        st.error("openpyxl is not installed. Run `pip install openpyxl` to enable Excel upload/download.")
    else:
        uploaded = st.file_uploader("Choose file", type=["xlsx"], key="preharvest_upload")
        if uploaded:
            with st.spinner("Parsing…"):
                try:
                    parsed = parse_preharvest_excel(uploaded.read())
                    imported = 0
                    for b in parsed:
                        k = b.get("key", "")
                        if k and k not in ("None", ""):
                            agg = b.get("aggregate", {})
                            plants = b.get("plants", [])
                            nutrients = b.get("nutrients", [])
                            pests = b.get("pests", [])
                            st.session_state.preharvest_batches[k] = {
                                "data": {
                                    "farm_name": agg.get("Farm Name", ""),
                                    "strain_name": agg.get("Strain Name", ""),
                                    "seed_source": agg.get("Seed Source", ""),
                                    "flowering_strategy": agg.get("Flowering Strategy", "Photoperiod"),
                                    "cultivated_area_sqm": agg.get("Cultivated Area (sqM)"),
                                    "growing_media": agg.get("Growing Media", ""),
                                    "soil_type": agg.get("Soil Type", ""),
                                    "germination_rate": agg.get("Germination Rate (%)"),
                                    "thc_pct": agg.get("THC (%)"),
                                    "cbd_pct": agg.get("CBD (%)"),
                                },
                                "plants_df": pd.DataFrame(plants) if plants else None,
                                "nutrients_df": pd.DataFrame(nutrients) if nutrients else None,
                                "pests_df": pd.DataFrame(pests) if pests else None,
                            }
                            imported += 1
                    if imported:
                        st.success(f"Imported {imported} batch(es).")
                        st.rerun()
                    else:
                        st.warning("No data rows found. Make sure the template has been filled in.")
                except Exception as e:
                    st.error(f"Could not parse file: {e}")

# ── Batch management ───────────────────────────────────────────────────────────
st.markdown("### Batch Management")

col_sel, col_new_id, col_btn = st.columns([3, 2, 1])
with col_sel:
    batch_list = list(st.session_state.preharvest_batches.keys())
    selected_batch = st.selectbox(
        "Select batch to view / edit",
        batch_list if batch_list else ["(none)"],
        key="preharvest_selected",
        disabled=not batch_list,
    )
    if not batch_list:
        selected_batch = None

with col_new_id:
    new_bid = st.text_input("New batch ID", placeholder="e.g. BATCH-2025-001",
                             key="ph_new_bid", label_visibility="collapsed")
with col_btn:
    st.write("")
    if st.button("➕ Create", use_container_width=True):
        bid = new_bid.strip()
        if not bid:
            st.error("Enter a batch ID.")
        elif bid in st.session_state.preharvest_batches:
            st.error(f"'{bid}' already exists.")
        else:
            st.session_state.preharvest_batches[bid] = {
                "data": {}, "plants_df": None, "nutrients_df": None, "pests_df": None,
            }
            st.session_state.preharvest_selected = bid
            st.rerun()

if not selected_batch or selected_batch == "(none)":
    st.info("Create a new batch or upload an existing template to get started.")
    st.stop()

batch = st.session_state.preharvest_batches[selected_batch]
bd = batch.setdefault("data", {})

col_lbl, col_del = st.columns([8, 1])
with col_lbl:
    st.markdown(f"Editing: <span class='batch-badge'>🌿 {selected_batch}</span>", unsafe_allow_html=True)
with col_del:
    if st.button("🗑 Delete", use_container_width=True, help="Remove this batch from the session"):
        del st.session_state.preharvest_batches[selected_batch]
        st.rerun()

st.write("")

# ── Data entry tabs ────────────────────────────────────────────────────────────
tabs = st.tabs([
    "🌱 Batch & Growing",
    "🌿 Individual Plants",
    "📋 Inputs Survey",
    "⚡ Energy Audit",
    "🧪 Nutrient Log",
    "🐛 Pest Control",
    "💰 Yield & Selling",
])

# ── TAB 1: Batch Identity & Growing Environment ────────────────────────────────
with tabs[0]:
    st.markdown("#### Batch Identity")
    c1, c2, c3 = st.columns(3)
    with c1:
        bd["farm_name"] = st.text_input("Farm Name", value=bd.get("farm_name", ""),
                                         key=f"ph_{selected_batch}_farm")
    with c2:
        bd["strain_name"] = st.text_input("Strain Name", value=bd.get("strain_name", ""),
                                           key=f"ph_{selected_batch}_strain")
    with c3:
        bd["seed_source"] = st.text_input("Seed Source", value=bd.get("seed_source", ""),
                                           placeholder="Breeder, clone source…",
                                           key=f"ph_{selected_batch}_seed")
    c1, c2 = st.columns(2)
    with c1:
        bd["flowering_strategy"] = st.selectbox(
            "Flowering Strategy", ["Photoperiod", "Autoflower"],
            index=0 if bd.get("flowering_strategy", "Photoperiod") == "Photoperiod" else 1,
            key=f"ph_{selected_batch}_flowering",
        )
    with c2:
        bd["season_cycle"] = st.text_input("Season / Cycle",
                                            value=bd.get("season_cycle", ""),
                                            placeholder="e.g. Spring 2025, Cycle 2",
                                            key=f"ph_{selected_batch}_season")

    st.divider()
    st.markdown("#### Growing Environment")
    c1, c2, c3 = st.columns(3)
    with c1:
        bd["cultivated_area_sqm"] = st.number_input(
            "Cultivated Area (sqM)", min_value=0.0, step=10.0,
            value=float(bd.get("cultivated_area_sqm") or 0.0),
            key=f"ph_{selected_batch}_area",
        )
    with c2:
        bd["growing_media"] = st.text_input("Growing Media", value=bd.get("growing_media", ""),
                                             placeholder="Soil, Coco, Hydro…",
                                             key=f"ph_{selected_batch}_media")
    with c3:
        bd["soil_type"] = st.text_input("Soil Type", value=bd.get("soil_type", ""),
                                         placeholder="Loam, Sandy loam…",
                                         key=f"ph_{selected_batch}_soil_type")
    c1, c2 = st.columns(2)
    with c1:
        aspects = ["—", "N", "NE", "E", "SE", "S", "SW", "W", "NW", "Flat"]
        cur_asp = bd.get("field_aspect", "—")
        bd["field_aspect"] = st.selectbox("Field Aspect", aspects,
                                           index=aspects.index(cur_asp) if cur_asp in aspects else 0,
                                           key=f"ph_{selected_batch}_aspect")
    with c2:
        steeps = ["—", "Flat", "Gentle slope", "Moderate slope", "Steep"]
        cur_st = bd.get("field_steepness", "—")
        bd["field_steepness"] = st.selectbox("Field Steepness", steeps,
                                              index=steeps.index(cur_st) if cur_st in steeps else 0,
                                              key=f"ph_{selected_batch}_steepness")

    st.markdown("**Plant Count by Environment**")
    pc1, pc2, pc3, pc4 = st.columns(4)
    with pc1:
        bd["n_plants_outdoor"] = st.number_input("# Outdoors", min_value=0, step=1,
            value=int(bd.get("n_plants_outdoor") or 0), key=f"ph_{selected_batch}_nout")
    with pc2:
        bd["n_plants_hoop"] = st.number_input("# Hoop House", min_value=0, step=1,
            value=int(bd.get("n_plants_hoop") or 0), key=f"ph_{selected_batch}_nhoop")
    with pc3:
        bd["n_plants_greenhouse"] = st.number_input("# Greenhouse", min_value=0, step=1,
            value=int(bd.get("n_plants_greenhouse") or 0), key=f"ph_{selected_batch}_ngh")
    with pc4:
        bd["n_plants_indoor"] = st.number_input("# Indoors", min_value=0, step=1,
            value=int(bd.get("n_plants_indoor") or 0), key=f"ph_{selected_batch}_nin")
    total_plants = sum(int(bd.get(k) or 0)
                       for k in ["n_plants_outdoor", "n_plants_hoop", "n_plants_greenhouse", "n_plants_indoor"])
    st.metric("Total Plants This Batch", total_plants)

    st.divider()
    st.markdown("#### Germination & Flowering")
    g1, g2, g3 = st.columns(3)
    with g1:
        bd["germination_rate"] = st.number_input("Germination Rate (%)", 0.0, 100.0, step=0.5,
            value=float(bd.get("germination_rate") or 0.0), key=f"ph_{selected_batch}_germ")
    with g2:
        bd["first_flower_date"] = st.text_input("Approx. First Flower Date (mm/dd/yyyy)",
            value=bd.get("first_flower_date", ""), key=f"ph_{selected_batch}_ff_date")
    with g3:
        bd["days_in_ground"] = st.number_input("Days in Ground", min_value=0, step=1,
            value=int(bd.get("days_in_ground") or 0), key=f"ph_{selected_batch}_days")

    st.divider()
    st.markdown("#### Disease")
    d1, d2, d3, d4 = st.columns(4)
    with d1:
        dp_opts = ["No (0)", "Yes (1)"]
        bd["disease_presence"] = st.selectbox("Disease Present?", dp_opts,
            index=1 if bd.get("disease_presence") == "Yes (1)" else 0,
            key=f"ph_{selected_batch}_dis_yn")
    with d2:
        bd["disease_name"] = st.text_input("Disease Name", value=bd.get("disease_name", ""),
            placeholder="Botrytis, PM, Fusarium…", key=f"ph_{selected_batch}_dis_name")
    with d3:
        sev_opts = ["—", "Mild", "Moderate", "Severe"]
        cur_sev = bd.get("disease_severity", "—")
        bd["disease_severity"] = st.selectbox("Severity", sev_opts,
            index=sev_opts.index(cur_sev) if cur_sev in sev_opts else 0,
            key=f"ph_{selected_batch}_dis_sev")
    with d4:
        bd["disease_date"] = st.text_input("Date First Seen (mm/dd/yy)",
            value=bd.get("disease_date", ""), key=f"ph_{selected_batch}_dis_date")

    st.divider()
    st.markdown("#### Harvest Quality — Trichome Assessment")
    hq1, hq2, hq3, hq4 = st.columns(4)
    with hq1:
        bd["pct_white_trichomes"] = st.number_input("% White (Cloudy)", 0.0, 100.0, step=1.0,
            value=float(bd.get("pct_white_trichomes") or 0.0), key=f"ph_{selected_batch}_wh_tri")
    with hq2:
        bd["pct_clear_trichomes"] = st.number_input("% Clear", 0.0, 100.0, step=1.0,
            value=float(bd.get("pct_clear_trichomes") or 0.0), key=f"ph_{selected_batch}_cl_tri")
    with hq3:
        bd["pct_amber_trichomes"] = st.number_input("% Amber", 0.0, 100.0, step=1.0,
            value=float(bd.get("pct_amber_trichomes") or 0.0), key=f"ph_{selected_batch}_am_tri")
    with hq4:
        bd["biomass_weight_kg"] = st.number_input("Total Biomass Weight (kg)", 0.0, step=0.1,
            value=float(bd.get("biomass_weight_kg") or 0.0), key=f"ph_{selected_batch}_biomass")

    st.divider()
    st.markdown("#### Lab Results — Cannabinoids & Terpenes")
    st.caption("Enter values from COA. Full COA attachment goes in Post-Harvest → Quality Testing tab.")
    lr1, lr2, lr3 = st.columns(3)
    with lr1:
        bd["thc_pct"] = st.number_input("THC (%)", 0.0, 100.0, step=0.1,
            value=float(bd.get("thc_pct") or 0.0), key=f"ph_{selected_batch}_thc")
    with lr2:
        bd["cbd_pct"] = st.number_input("CBD (%)", 0.0, 100.0, step=0.1,
            value=float(bd.get("cbd_pct") or 0.0), key=f"ph_{selected_batch}_cbd")
    with lr3:
        bd["other_cannabinoids"] = st.text_input("Other Cannabinoids (notes)",
            value=bd.get("other_cannabinoids", ""), placeholder="CBG: 0.5%, CBN: 0.2%…",
            key=f"ph_{selected_batch}_other_cann")
    lr4, lr5 = st.columns(2)
    with lr4:
        bd["b_myrcene_pct"] = st.number_input("β-Myrcene (%)", 0.0, 100.0, step=0.001,
            value=float(bd.get("b_myrcene_pct") or 0.0),
            key=f"ph_{selected_batch}_myrcene", format="%.3f")
    with lr5:
        bd["b_caryophyllene_pct"] = st.number_input("β-Caryophyllene (%)", 0.0, 100.0, step=0.001,
            value=float(bd.get("b_caryophyllene_pct") or 0.0),
            key=f"ph_{selected_batch}_caryoph", format="%.3f")

# ── TAB 2: Individual Plant Data ───────────────────────────────────────────────
with tabs[1]:
    st.markdown("#### Individual Plant Data")
    st.caption("One row per plant. Batch Number is inherited from the selected batch.")

    _plant_schema = {
        "Plant Number": pd.Series(dtype="Int64"),
        "Plant ID / Tag": pd.Series(dtype="str"),
        "Environment": pd.Series(dtype="str"),
        "Growing Media": pd.Series(dtype="str"),
        "Seed Planting Date": pd.Series(dtype="str"),
        "Transplanting Date": pd.Series(dtype="str"),
        "Topped (Y/N)": pd.Series(dtype="str"),
        "Pruned (Y/N)": pd.Series(dtype="str"),
        "Trellised (Y/N)": pd.Series(dtype="str"),
        "Height at Harvest (cm)": pd.Series(dtype="Float64"),
        "Stem Width (mm)": pd.Series(dtype="Float64"),
        "Node Count": pd.Series(dtype="Int64"),
        "Bud Size": pd.Series(dtype="str"),
    }
    existing = batch.get("plants_df")
    default_df = pd.DataFrame(_plant_schema) if (existing is None or existing.empty) else existing

    edited = st.data_editor(
        default_df,
        num_rows="dynamic",
        use_container_width=True,
        key=f"plants_editor_{selected_batch}",
        column_config={
            "Plant Number": st.column_config.NumberColumn("Plant #", min_value=1, step=1),
            "Environment": st.column_config.SelectboxColumn("Environment",
                options=["Outdoors", "Hoop House", "Greenhouse", "Indoors", "Aquaponics", "Other"]),
            "Topped (Y/N)": st.column_config.SelectboxColumn("Topped", options=["Y", "N"]),
            "Pruned (Y/N)": st.column_config.SelectboxColumn("Pruned", options=["Y", "N"]),
            "Trellised (Y/N)": st.column_config.SelectboxColumn("Trellised", options=["Y", "N"]),
            "Height at Harvest (cm)": st.column_config.NumberColumn("Height (cm)", min_value=0.0, format="%.1f"),
            "Stem Width (mm)": st.column_config.NumberColumn("Stem (mm)", min_value=0.0, format="%.1f"),
            "Node Count": st.column_config.NumberColumn("Nodes", min_value=0, step=1),
            "Bud Size": st.column_config.SelectboxColumn("Bud Size", options=["S", "M", "L", "XL"]),
        },
    )
    st.session_state.preharvest_batches[selected_batch]["plants_df"] = edited
    if not edited.empty:
        st.caption(f"{len(edited)} plant records")

# ── TAB 3: Inputs Survey ───────────────────────────────────────────────────────
with tabs[2]:
    st.markdown("#### Inputs Survey — Do You Use?")
    st.caption("Check all inputs used this batch / season.")

    inputs_list = [
        ("Mineral / Conventional Soil Amendments", "input_mineral_soil"),
        ("Blood Meal", "input_blood_meal"),
        ("Green Sand (potassium / mineral source)", "input_greensand"),
        ("Feather Meal (slow-release N)", "input_feather_meal"),
        ("Poultry Manure / Chicken Litter", "input_poultry_manure"),
        ("Livestock Manure (cattle, horse, other)", "input_livestock_manure"),
        ("Compost / Worm Castings (Vermicompost)", "input_compost_worm"),
        ("Biological Inoculants (Mycorrhizae, PGPR, etc.)", "input_biological"),
        ("Regular Compost (not vermicompost)", "input_regular_compost"),
        ("Plasticulture (plastic mulch, row covers)", "input_plasticulture"),
        ("Organic Mulch / Cover Crops", "input_mulch"),
    ]
    c1, c2, c3 = st.columns(3)
    for idx, (label, key) in enumerate(inputs_list):
        with [c1, c2, c3][idx % 3]:
            bd[key] = st.checkbox(label, value=bool(bd.get(key, False)),
                                  key=f"ph_{selected_batch}_{key}")

# ── TAB 4: Energy Audit (Indoor) ───────────────────────────────────────────────
with tabs[3]:
    n_indoor = int(bd.get("n_plants_indoor") or 0)
    if n_indoor == 0:
        st.info("No indoor plants entered for this batch. Energy Audit applies to indoor grows only. "
                "Update the plant count in the **Batch & Growing** tab if needed.")

    st.markdown("#### Energy Audit — Indoor Operations")
    st.caption("Complete for each indoor grow zone. Leave blank if fully outdoor or hoop house.")

    ea1, ea2, ea3 = st.columns(3)
    with ea1:
        bd["grow_room"] = st.text_input("Grow Room / Zone", value=bd.get("grow_room", ""),
            placeholder="e.g. Veg Room A", key=f"ph_{selected_batch}_grow_room")
    with ea2:
        lt_opts = ["—", "LED", "HPS", "CMH", "Fluorescent", "Other"]
        cur_lt = bd.get("lighting_type", "—")
        bd["lighting_type"] = st.selectbox("Lighting Type", lt_opts,
            index=lt_opts.index(cur_lt) if cur_lt in lt_opts else 0,
            key=f"ph_{selected_batch}_light_type")
    with ea3:
        bd["lighting_wattage_w"] = st.number_input("Lighting Wattage (W)", min_value=0.0, step=100.0,
            value=float(bd.get("lighting_wattage_w") or 0.0), key=f"ph_{selected_batch}_wattage")

    ea4, ea5, ea6 = st.columns(3)
    with ea4:
        bd["light_hours_per_day"] = st.number_input("Light Hours / Day", 0.0, 24.0, step=0.5,
            value=float(bd.get("light_hours_per_day") or 18.0), key=f"ph_{selected_batch}_l_hrs")
    with ea5:
        bd["weeks_veg"] = st.number_input("Weeks in Veg", min_value=0, step=1,
            value=int(bd.get("weeks_veg") or 0), key=f"ph_{selected_batch}_wks_veg")
    with ea6:
        bd["weeks_flower"] = st.number_input("Weeks in Flower", min_value=0, step=1,
            value=int(bd.get("weeks_flower") or 0), key=f"ph_{selected_batch}_wks_fl")

    ea7, ea8 = st.columns(2)
    with ea7:
        bd["hvac_type"] = st.text_input("HVAC System Type", value=bd.get("hvac_type", ""),
            placeholder="Mini-split, chiller, swamp cooler…", key=f"ph_{selected_batch}_hvac_type")
    with ea8:
        bd["hvac_kwh_month"] = st.number_input("HVAC Est. kWh / Month", min_value=0.0, step=10.0,
            value=float(bd.get("hvac_kwh_month") or 0.0), key=f"ph_{selected_batch}_hvac_kwh")

    ea9, ea10, ea11 = st.columns(3)
    with ea9:
        co2_opts = ["No", "Yes"]
        bd["co2_supplementation"] = st.selectbox("CO2 Supplementation", co2_opts,
            index=1 if bd.get("co2_supplementation") == "Yes" else 0,
            key=f"ph_{selected_batch}_co2")
    with ea10:
        bd["co2_source"] = st.text_input("CO2 Source", value=bd.get("co2_source", ""),
            placeholder="Tank, generator, fermentation…",
            key=f"ph_{selected_batch}_co2_src",
            disabled=(bd.get("co2_supplementation") != "Yes"))
    with ea11:
        irr_opts = ["—", "Drip", "Hand water", "Flood / drain", "Ebb and flow", "Aeroponics", "Other"]
        cur_irr = bd.get("irrigation_system", "—")
        bd["irrigation_system"] = st.selectbox("Irrigation System", irr_opts,
            index=irr_opts.index(cur_irr) if cur_irr in irr_opts else 0,
            key=f"ph_{selected_batch}_irr")

    # Calculated metrics
    w = bd.get("lighting_wattage_w") or 0
    h = bd.get("light_hours_per_day") or 0
    days = ((bd.get("weeks_veg") or 0) + (bd.get("weeks_flower") or 0)) * 7
    l_kwh = round(w * h * days / 1000, 1) if days > 0 else 0
    hvac_kwh = round((bd.get("hvac_kwh_month") or 0) * days / 30, 1) if days > 0 else 0
    if l_kwh > 0 or hvac_kwh > 0:
        st.divider()
        ek1, ek2, ek3 = st.columns(3)
        ek1.metric("Lighting kWh / Cycle", f"{l_kwh:,.0f}")
        ek2.metric("HVAC kWh / Cycle", f"{hvac_kwh:,.0f}")
        ek3.metric("Total Est. kWh / Cycle", f"{l_kwh + hvac_kwh:,.0f}")

# ── TAB 5: Nutrient Amendment Log ─────────────────────────────────────────────
with tabs[4]:
    st.markdown("#### Nutrient Amendment Log")
    st.caption("One row per application event. Record every fertilizer, compost, or pH amendment.")

    _nut_schema = {
        "Application Date (mm/dd/yy)": pd.Series(dtype="str"),
        "Crop Stage": pd.Series(dtype="str"),
        "Product / Amendment Name": pd.Series(dtype="str"),
        "Type": pd.Series(dtype="str"),
        "NPK / Analysis": pd.Series(dtype="str"),
        "Rate Applied": pd.Series(dtype="Float64"),
        "Units": pd.Series(dtype="str"),
        "Application Method": pd.Series(dtype="str"),
        "Notes": pd.Series(dtype="str"),
    }
    existing_n = batch.get("nutrients_df")
    default_n = pd.DataFrame(_nut_schema) if (existing_n is None or existing_n.empty) else existing_n

    edited_n = st.data_editor(
        default_n,
        num_rows="dynamic",
        use_container_width=True,
        key=f"nut_editor_{selected_batch}",
        column_config={
            "Crop Stage": st.column_config.SelectboxColumn("Crop Stage",
                options=["Seedling", "Veg", "Early Flower", "Mid Flower", "Late Flower", "Flush", "Other"]),
            "Type": st.column_config.SelectboxColumn("Type",
                options=["Organic", "Synthetic", "Biological", "pH Amendment", "Other"]),
            "Application Method": st.column_config.SelectboxColumn("Method",
                options=["Top dress", "Drench", "Foliar", "Fertigation", "Broadcast", "Other"]),
            "Units": st.column_config.SelectboxColumn("Units",
                options=["g/plant", "kg/acre", "mL/L", "oz/gallon", "lbs/acre", "Other"]),
            "Rate Applied": st.column_config.NumberColumn("Rate", min_value=0.0, format="%.2f"),
        },
    )
    st.session_state.preharvest_batches[selected_batch]["nutrients_df"] = edited_n
    if not edited_n.empty:
        st.caption(f"{len(edited_n)} amendment event(s) recorded")

# ── TAB 6: Pest Control Log ────────────────────────────────────────────────────
with tabs[5]:
    st.markdown("#### Pest Control Application Log")
    st.caption("One row per scouting or application event. Required for NYS licensing compliance recordkeeping.")

    _pest_schema = {
        "Date (mm/dd/yy)": pd.Series(dtype="str"),
        "Event Type": pd.Series(dtype="str"),
        "Target Pest / Disease": pd.Series(dtype="str"),
        "Product / Method Name": pd.Series(dtype="str"),
        "Type": pd.Series(dtype="str"),
        "EPA Reg. # / Active Ingredient": pd.Series(dtype="str"),
        "Rate Applied": pd.Series(dtype="str"),
        "Application Method": pd.Series(dtype="str"),
        "PHI (days)": pd.Series(dtype="Float64"),
        "Applied By": pd.Series(dtype="str"),
        "Notes": pd.Series(dtype="str"),
    }
    existing_p = batch.get("pests_df")
    default_p = pd.DataFrame(_pest_schema) if (existing_p is None or existing_p.empty) else existing_p

    edited_p = st.data_editor(
        default_p,
        num_rows="dynamic",
        use_container_width=True,
        key=f"pest_editor_{selected_batch}",
        column_config={
            "Event Type": st.column_config.SelectboxColumn("Event Type",
                options=["Scouting", "Preventative Application", "Curative Application"]),
            "Type": st.column_config.SelectboxColumn("Type",
                options=["Pesticide (OMRI listed)", "Pesticide (synthetic)",
                         "Biocontrol", "Physical/Mechanical", "None (scouting)"]),
            "Application Method": st.column_config.SelectboxColumn("Method",
                options=["Foliar spray", "Soil drench", "Biocontrol release", "Trap placement", "Other"]),
            "PHI (days)": st.column_config.NumberColumn("PHI (days)", min_value=0, step=1),
        },
    )
    st.session_state.preharvest_batches[selected_batch]["pests_df"] = edited_p
    if not edited_p.empty:
        st.caption(f"{len(edited_p)} event(s) recorded")

# ── TAB 7: Yield & Selling ─────────────────────────────────────────────────────
with tabs[6]:
    st.markdown("#### Yield & Selling Information")
    st.caption("Record quantities sold and prices. Processing weights (bucked, dried, trimmed) go in Post-Harvest.")

    y1, y2 = st.columns(2)
    with y1:
        st.markdown("**Flower**")
        bd["flower_sold_g"] = st.number_input("Total Sold (g)", min_value=0.0, step=1.0,
            value=float(bd.get("flower_sold_g") or 0.0), key=f"ph_{selected_batch}_fl_sold")
        bd["flower_price_per_g"] = st.number_input("Price ($/g)", min_value=0.0, step=0.01,
            value=float(bd.get("flower_price_per_g") or 0.0),
            key=f"ph_{selected_batch}_fl_price", format="%.2f")
        bd["flower_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
            value=bd.get("flower_sale_date", ""), key=f"ph_{selected_batch}_fl_date")
        fl_rev = round((bd.get("flower_sold_g") or 0) * (bd.get("flower_price_per_g") or 0), 2)
        bd["flower_revenue"] = fl_rev
        if fl_rev > 0:
            st.metric("Flower Revenue", f"${fl_rev:,.2f}")

    with y2:
        st.markdown("**Pre-Rolls**")
        bd["preroll_sold_units"] = st.number_input("Total Sold (units)", min_value=0, step=1,
            value=int(bd.get("preroll_sold_units") or 0), key=f"ph_{selected_batch}_pr_sold")
        bd["preroll_price_per_unit"] = st.number_input("Price ($/unit)", min_value=0.0, step=0.01,
            value=float(bd.get("preroll_price_per_unit") or 0.0),
            key=f"ph_{selected_batch}_pr_price", format="%.2f")
        bd["preroll_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
            value=bd.get("preroll_sale_date", ""), key=f"ph_{selected_batch}_pr_date")
        pr_rev = round((bd.get("preroll_sold_units") or 0) * (bd.get("preroll_price_per_unit") or 0), 2)
        bd["preroll_revenue"] = pr_rev
        if pr_rev > 0:
            st.metric("Pre-Roll Revenue", f"${pr_rev:,.2f}")

    st.divider()
    st.markdown("**Biomass**")
    bm1, bm2, bm3 = st.columns(3)
    with bm1:
        bd["biomass_sold_g"] = st.number_input("Total Sold (g)", min_value=0.0, step=10.0,
            value=float(bd.get("biomass_sold_g") or 0.0), key=f"ph_{selected_batch}_bm_sold")
    with bm2:
        bd["biomass_price_per_kg"] = st.number_input("Price ($/kg)", min_value=0.0, step=1.0,
            value=float(bd.get("biomass_price_per_kg") or 0.0), key=f"ph_{selected_batch}_bm_price")
    with bm3:
        bd["biomass_sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
            value=bd.get("biomass_sale_date", ""), key=f"ph_{selected_batch}_bm_date")

    bm_rev = round((bd.get("biomass_sold_g") or 0) / 1000 * (bd.get("biomass_price_per_kg") or 0), 2)
    bd["biomass_revenue"] = bm_rev

    total_rev = (bd.get("flower_revenue") or 0) + (bd.get("preroll_revenue") or 0) + bm_rev
    if total_rev > 0:
        st.divider()
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Flower", f"${bd.get('flower_revenue', 0):,.2f}")
        r2.metric("Pre-Rolls", f"${bd.get('preroll_revenue', 0):,.2f}")
        r3.metric("Biomass", f"${bm_rev:,.2f}")
        r4.metric("Total Revenue", f"${total_rev:,.2f}")

# ── Download ───────────────────────────────────────────────────────────────────
st.divider()
st.markdown("### Download Data")

dl1, dl2, dl3 = st.columns(3)

with dl1:
    if st.session_state.preharvest_batches:
        rows = []
        for bn, b in st.session_state.preharvest_batches.items():
            row = {"Batch Number": bn}
            row.update(b.get("data", {}))
            rows.append(row)
        csv_bytes = pd.DataFrame(rows).to_csv(index=False).encode()
        st.download_button("⬇ Batch Summary (CSV)", data=csv_bytes,
                           file_name="preharvest_batch_summary.csv", mime="text/csv",
                           use_container_width=True)

with dl2:
    if HAS_OPENPYXL and st.session_state.preharvest_batches:
        xl = build_preharvest_excel(st.session_state.preharvest_batches)
        st.download_button("⬇ Full Data (Excel)", data=xl,
                           file_name="preharvest_data.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

with dl3:
    all_logs = []
    for bn, b in st.session_state.preharvest_batches.items():
        for df_key in ["nutrients_df", "pests_df"]:
            df = b.get(df_key)
            if df is not None and not df.empty:
                tmp = df.copy()
                tmp.insert(0, "Log Type", "Nutrient" if df_key == "nutrients_df" else "Pest")
                tmp.insert(0, "Batch Number", bn)
                all_logs.append(tmp)
    if all_logs:
        combined = pd.concat(all_logs, ignore_index=True)
        st.download_button("⬇ All Logs (CSV)", data=combined.to_csv(index=False).encode(),
                           file_name="preharvest_logs.csv", mime="text/csv",
                           use_container_width=True)

st.caption("Session data is not persisted after the browser tab closes. Download your data to keep it.")

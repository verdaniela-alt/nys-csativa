"""
5_Batch_Overview.py — Batch Dashboard
Links Pre-Harvest and Post-Harvest data by Batch Number / Harvest Batch ID.
Shows key metrics, completeness indicators, weight flow, and revenue summary.
"""

import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd

try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

st.set_page_config(
    page_title="Batch Overview | NYS Cannabis Tool",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
.status-complete { color: #155724; font-weight: bold; }
.status-partial  { color: #856404; font-weight: bold; }
.status-missing  { color: #721c24; font-weight: bold; }
.section-header  {
    font-size: 1.05rem; font-weight: 600; color: #1565C0;
    border-bottom: 2px solid #1565C0; padding-bottom: 4px; margin-bottom: 12px;
}
.batch-card {
    background: #f8f9fa; border: 1px solid #dee2e6;
    border-radius: 10px; padding: 20px; height: 100%;
}
</style>
""", unsafe_allow_html=True)

# ── Helpers ────────────────────────────────────────────────────────────────────

def _has_data(d, *keys):
    return any(d.get(k) not in (None, "", 0, 0.0) for k in keys)


def _completeness(label, d, required_keys, optional_keys=None):
    """Return (label, status_str, pct_complete)."""
    total = len(required_keys) + len(optional_keys or [])
    filled = sum(1 for k in (required_keys + (optional_keys or [])) if d.get(k) not in (None, "", 0))
    pct = round(filled / total * 100) if total else 0
    if pct == 100:
        status = "complete"
    elif pct == 0:
        status = "missing"
    else:
        status = "partial"
    return label, status, pct


def _build_combined_report(batch_id, ph, pos):
    """Build a combined Excel report for one batch."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ph_d = ph.get("data", {}) if ph else {}
    pos_d = pos.get("data", {}) if pos else {}

    ws.append(["Field", "Value", "Source"])
    ws.append(["Batch Number / ID", batch_id, ""])
    ws.append([])

    ws.append(["--- PRE-HARVEST ---", "", ""])
    for k, v in ph_d.items():
        ws.append([k, v, "Pre-Harvest"])

    ws.append([])
    ws.append(["--- POST-HARVEST ---", "", ""])
    for k, v in pos_d.items():
        ws.append([k, v, "Post-Harvest"])

    # Plants sheet
    if ph and ph.get("plants_df") is not None and not ph["plants_df"].empty:
        ws2 = wb.create_sheet("Individual_Plants")
        df = ph["plants_df"]
        ws2.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows():
            ws2.append([batch_id] + list(row))

    # Nutrient log
    if ph and ph.get("nutrients_df") is not None and not ph["nutrients_df"].empty:
        ws3 = wb.create_sheet("Nutrient_Log")
        df = ph["nutrients_df"]
        ws3.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows():
            ws3.append([batch_id] + list(row))

    # Pest log
    if ph and ph.get("pests_df") is not None and not ph["pests_df"].empty:
        ws4 = wb.create_sheet("Pest_Log")
        df = ph["pests_df"]
        ws4.append(["Batch Number"] + list(df.columns))
        for _, row in df.iterrows():
            ws4.append([batch_id] + list(row))

    # Curing log
    if pos and pos.get("curing_log_df") is not None and not pos["curing_log_df"].empty:
        ws5 = wb.create_sheet("Curing_Log")
        df = pos["curing_log_df"]
        ws5.append(["Harvest Batch ID"] + list(df.columns))
        for _, row in df.iterrows():
            ws5.append([batch_id] + list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Page layout ────────────────────────────────────────────────────────────────

st.title("📊 Batch Overview Dashboard")
st.caption("Links Pre-Harvest and Post-Harvest data by Batch Number. "
           "Enter data in the Pre-Harvest and Post-Harvest pages first.")
st.divider()

ph_batches = st.session_state.get("preharvest_batches", {})
pos_batches = st.session_state.get("postharvest_batches", {})

all_ids = sorted(set(list(ph_batches.keys()) + list(pos_batches.keys())))

if not all_ids:
    st.info("No batch data found in this session. "
            "Go to the **Pre-Harvest** or **Post-Harvest** pages to enter or upload data, "
            "then return here.")
    st.stop()

# ── Batch selector ─────────────────────────────────────────────────────────────
col_sel, col_all = st.columns([3, 1])
with col_sel:
    selected = st.selectbox("Select Batch", all_ids, key="overview_batch")
with col_all:
    st.write("")
    show_all = st.checkbox("Show all batches summary", value=False)

ph = ph_batches.get(selected)
pos = pos_batches.get(selected)
ph_d = ph.get("data", {}) if ph else {}
pos_d = pos.get("data", {}) if pos else {}

# ── Completeness row ───────────────────────────────────────────────────────────
st.markdown("#### Data Completeness")

checks = []
if ph:
    checks += [
        _completeness("Batch Identity", ph_d,
                      ["farm_name", "strain_name", "flowering_strategy"],
                      ["seed_source", "season_cycle"]),
        _completeness("Growing Env.", ph_d,
                      ["cultivated_area_sqm"],
                      ["growing_media", "n_plants_outdoor", "n_plants_greenhouse", "n_plants_indoor"]),
        _completeness("Lab Results", ph_d,
                      ["thc_pct", "cbd_pct"], ["b_myrcene_pct"]),
        _completeness("Yield / Selling", ph_d,
                      [], ["flower_sold_g", "flower_price_per_g"]),
    ]
else:
    checks.append(("Pre-Harvest", "missing", 0))

if pos:
    checks += [
        _completeness("Harvest Log", pos_d,
                      ["harvest_date", "harvested_plant_count"], ["flower_room"]),
        _completeness("Processing Wts", pos_d,
                      ["total_wet_weight_g", "dried_weight_g"], ["trimmed_flower_g"]),
        _completeness("Quality / COA", pos_d,
                      ["test_result"], ["lab_name", "coa_link"]),
        _completeness("Packaging", pos_d,
                      [], ["flower_available_for_packaging_g", "pkg_1g", "pkg_eighth"]),
    ]
else:
    checks.append(("Post-Harvest", "missing", 0))

icon_map = {"complete": "✅", "partial": "⚠️", "missing": "❌"}
cols = st.columns(len(checks))
for col, (label, status, pct) in zip(cols, checks):
    with col:
        st.markdown(f"**{label}**")
        st.markdown(f"{icon_map[status]} {pct}%",
                    help=f"Status: {status} ({pct}% of fields filled)")

st.divider()

# ── Key metrics ────────────────────────────────────────────────────────────────
st.markdown("#### Key Metrics")

km_cols = st.columns(4)

with km_cols[0]:
    st.markdown('<p class="section-header">Strain & Grow</p>', unsafe_allow_html=True)
    strain = ph_d.get("strain_name") or "—"
    farm = ph_d.get("farm_name") or "—"
    strat = ph_d.get("flowering_strategy") or "—"
    season = ph_d.get("season_cycle") or "—"
    st.markdown(f"**Strain:** {strain}")
    st.markdown(f"**Farm:** {farm}")
    st.markdown(f"**Type:** {strat}")
    st.markdown(f"**Season:** {season}")

with km_cols[1]:
    st.markdown('<p class="section-header">Plants & Environment</p>', unsafe_allow_html=True)
    total_plants = sum(int(ph_d.get(k) or 0)
                       for k in ["n_plants_outdoor", "n_plants_hoop",
                                 "n_plants_greenhouse", "n_plants_indoor"])
    harvested = int(pos_d.get("harvested_plant_count") or 0)
    area = ph_d.get("cultivated_area_sqm")
    st.markdown(f"**Total Plants:** {total_plants or '—'}")
    st.markdown(f"**Harvested:** {harvested or '—'}")
    if total_plants and harvested:
        loss_pct = round((total_plants - harvested) / total_plants * 100, 1)
        st.markdown(f"**Plant Loss:** {total_plants - harvested} ({loss_pct}%)")
    st.markdown(f"**Cultivated Area:** {f'{area} sqM' if area else '—'}")

with km_cols[2]:
    st.markdown('<p class="section-header">Harvest Date & COA</p>', unsafe_allow_html=True)
    hdate = pos_d.get("harvest_date") or "—"
    result = pos_d.get("test_result") or "—"
    thc = ph_d.get("thc_pct")
    cbd = ph_d.get("cbd_pct")
    st.markdown(f"**Harvest Date:** {hdate}")
    result_fmt = f"✅ {result}" if "Pass" in str(result) else (f"❌ {result}" if "Fail" in str(result) else result)
    st.markdown(f"**COA Result:** {result_fmt}")
    st.markdown(f"**THC:** {f'{thc}%' if thc else '—'}")
    st.markdown(f"**CBD:** {f'{cbd}%' if cbd else '—'}")

with km_cols[3]:
    st.markdown('<p class="section-header">Revenue</p>', unsafe_allow_html=True)
    fl_rev = ph_d.get("flower_revenue") or 0
    pr_rev = ph_d.get("preroll_revenue") or 0
    bm_rev = ph_d.get("biomass_revenue") or 0
    trim_rev = pos_d.get("trim_revenue") or 0
    total_rev = fl_rev + pr_rev + bm_rev + trim_rev
    if total_rev > 0:
        st.markdown(f"**Flower:** ${fl_rev:,.2f}")
        st.markdown(f"**Pre-Rolls:** ${pr_rev:,.2f}")
        st.markdown(f"**Biomass:** ${bm_rev:,.2f}")
        st.markdown(f"**Trim:** ${trim_rev:,.2f}")
        st.markdown(f"**Total: ${total_rev:,.2f}**")
    else:
        st.markdown("No revenue data entered.")

st.divider()

# ── Weight flow ────────────────────────────────────────────────────────────────
wet = pos_d.get("total_wet_weight_g") or 0
dried = pos_d.get("dried_weight_g") or 0
trimmed = pos_d.get("trimmed_flower_g") or 0
trim_byp = pos_d.get("trim_weight_g") or 0
packaged = (
    (pos_d.get("pkg_1g") or 0) * 1 +
    (pos_d.get("pkg_eighth") or 0) * 3.5 +
    (pos_d.get("pkg_quarter") or 0) * 7 +
    (pos_d.get("pkg_half") or 0) * 14 +
    (pos_d.get("pkg_oz") or 0) * 28
)

if wet > 0:
    st.markdown("#### Weight Flow")

    if HAS_PLOTLY and dried > 0:
        stages = ["Wet Weight", "Dried Weight", "Trimmed Flower", "Packaged"]
        values = [wet, dried, trimmed if trimmed else dried, packaged if packaged else (trimmed or dried)]
        colors = ["#4CAF50", "#2196F3", "#FF9800", "#9C27B0"]

        fig = go.Figure(go.Funnel(
            y=stages,
            x=values,
            textinfo="value+percent initial",
            marker_color=colors,
            connector={"line": {"color": "royalblue", "dash": "solid", "width": 2}},
        ))
        fig.update_layout(height=300, margin=dict(l=20, r=20, t=20, b=20))
        st.plotly_chart(fig, use_container_width=True)
    else:
        wf_cols = st.columns(4)
        for col, (label, val) in zip(wf_cols, [
            ("Wet Weight (g)", wet), ("Dried (g)", dried),
            ("Trimmed Flower (g)", trimmed), ("Packaged (g)", packaged),
        ]):
            col.metric(label, f"{val:,.0f}" if val else "—")

    if wet and dried:
        moisture = round((1 - dried / wet) * 100, 1)
        notes = [f"Moisture loss: **{moisture}%**"]
        if trimmed and dried:
            notes.append(f"Flower recovery from dry: **{trimmed/dried*100:.1f}%**")
        if trim_byp:
            notes.append(f"Trim byproduct: **{trim_byp:,.0f} g**")
        st.caption(" | ".join(notes))

    st.divider()

# ── Revenue breakdown ──────────────────────────────────────────────────────────
if total_rev > 0 and HAS_PLOTLY:
    st.markdown("#### Revenue Breakdown")
    rev_labels, rev_values = [], []
    for label, val in [("Flower", fl_rev), ("Pre-Rolls", pr_rev),
                        ("Biomass", bm_rev), ("Trim", trim_rev)]:
        if val > 0:
            rev_labels.append(label)
            rev_values.append(val)

    if rev_labels:
        fig_rev = px.pie(values=rev_values, names=rev_labels,
                         title=f"Total Revenue: ${total_rev:,.2f}",
                         hole=0.4,
                         color_discrete_sequence=["#4CAF50", "#FF9800", "#9C27B0", "#2196F3"])
        fig_rev.update_layout(height=320, legend=dict(orientation="h", y=-0.15))
        st.plotly_chart(fig_rev, use_container_width=True)
        st.divider()

# ── Terpene & cannabinoid summary ─────────────────────────────────────────────
if _has_data(ph_d, "thc_pct", "cbd_pct", "b_myrcene_pct", "b_caryophyllene_pct"):
    st.markdown("#### Cannabinoids & Terpenes")
    cn_cols = st.columns(5)
    for col, (label, key, fmt) in zip(cn_cols, [
        ("THC (%)", "thc_pct", "{:.1f}%"),
        ("CBD (%)", "cbd_pct", "{:.1f}%"),
        ("Other Cannabinoids", "other_cannabinoids", "{}"),
        ("β-Myrcene (%)", "b_myrcene_pct", "{:.3f}%"),
        ("β-Caryophyllene (%)", "b_caryophyllene_pct", "{:.3f}%"),
    ]):
        val = ph_d.get(key)
        if val:
            try:
                col.metric(label, fmt.format(val))
            except (TypeError, ValueError):
                col.metric(label, str(val))
    st.divider()

# ── Pest control summary ───────────────────────────────────────────────────────
if ph and ph.get("pests_df") is not None and not ph["pests_df"].empty:
    pest_df = ph["pests_df"]
    st.markdown("#### Pest Control Log Summary")
    n_events = len(pest_df)
    n_apps = len(pest_df[pest_df.get("Event Type", pd.Series()).str.contains("Application", na=False)])
    pc1, pc2 = st.columns(2)
    pc1.metric("Total Events", n_events)
    pc2.metric("Application Events", n_apps)
    with st.expander("View pest control log"):
        st.dataframe(pest_df, use_container_width=True, hide_index=True)
    st.divider()

# ── Nutrient log summary ───────────────────────────────────────────────────────
if ph and ph.get("nutrients_df") is not None and not ph["nutrients_df"].empty:
    nut_df = ph["nutrients_df"]
    st.markdown("#### Nutrient Amendment Log Summary")
    st.metric("Total Amendment Events", len(nut_df))
    with st.expander("View nutrient log"):
        st.dataframe(nut_df, use_container_width=True, hide_index=True)
    st.divider()

# ── All-batches summary table ──────────────────────────────────────────────────
if show_all and all_ids:
    st.markdown("#### All Batches Summary")
    rows = []
    for bid in all_ids:
        ph_b = ph_batches.get(bid, {})
        pos_b = pos_batches.get(bid, {})
        ph_dd = ph_b.get("data", {})
        pos_dd = pos_b.get("data", {})
        total_pl = sum(int(ph_dd.get(k) or 0)
                       for k in ["n_plants_outdoor", "n_plants_hoop", "n_plants_greenhouse", "n_plants_indoor"])
        wet_w = pos_dd.get("total_wet_weight_g") or 0
        dry_w = pos_dd.get("dried_weight_g") or 0
        rev = sum([
            ph_dd.get("flower_revenue") or 0,
            ph_dd.get("preroll_revenue") or 0,
            ph_dd.get("biomass_revenue") or 0,
            pos_dd.get("trim_revenue") or 0,
        ])
        rows.append({
            "Batch ID": bid,
            "Strain": ph_dd.get("strain_name") or "—",
            "Season": ph_dd.get("season_cycle") or "—",
            "Harvest Date": pos_dd.get("harvest_date") or "—",
            "Total Plants": total_pl or "—",
            "Wet Wt (g)": f"{wet_w:,.0f}" if wet_w else "—",
            "Dry Wt (g)": f"{dry_w:,.0f}" if dry_w else "—",
            "Moisture Loss": f"{round((1-dry_w/wet_w)*100,1)}%" if wet_w and dry_w else "—",
            "COA Result": pos_dd.get("test_result") or "—",
            "THC (%)": ph_dd.get("thc_pct") or "—",
            "Total Revenue": f"${rev:,.2f}" if rev else "—",
            "Pre-Harvest": "✅" if bid in ph_batches else "❌",
            "Post-Harvest": "✅" if bid in pos_batches else "❌",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Combined CSV
    st.download_button(
        "⬇ Download All-Batches Summary (CSV)",
        data=pd.DataFrame(rows).to_csv(index=False).encode(),
        file_name="all_batches_summary.csv",
        mime="text/csv",
    )
    st.divider()

# ── Per-batch download ─────────────────────────────────────────────────────────
st.markdown("### Download — Selected Batch")
dd1, dd2 = st.columns(2)

with dd1:
    # CSV: all fields for selected batch
    row_ph = {"Batch Number": selected}
    row_ph.update(ph_d)
    row_pos = {"Harvest Batch ID": selected}
    row_pos.update(pos_d)
    combined_df = pd.DataFrame([row_ph, row_pos])
    st.download_button(
        f"⬇ {selected} — Combined Summary (CSV)",
        data=combined_df.to_csv(index=False).encode(),
        file_name=f"{selected}_combined_summary.csv",
        mime="text/csv",
        use_container_width=True,
    )

with dd2:
    if HAS_OPENPYXL and (ph or pos):
        xl = _build_combined_report(selected, ph, pos)
        st.download_button(
            f"⬇ {selected} — Full Report (Excel)",
            data=xl,
            file_name=f"{selected}_full_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.caption("All data is session-based. Use the download buttons above to save your records.")

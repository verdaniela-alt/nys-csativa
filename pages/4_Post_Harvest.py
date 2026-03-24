"""
4_Post_Harvest.py — NY Cannabis/Hemp Post-Harvest Data Collection Tool
Upload existing PostHarvest Excel template OR enter data manually.
Download results as CSV or Excel.
"""

import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import streamlit as st
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

st.set_page_config(
    page_title="Post-Harvest | NYS Cannabis Tool",
    page_icon="🌾",
    layout="wide",
)

st.markdown("""
<style>
.disclaimer-box {
    background: #fff3cd; border: 2px solid #e0a800;
    border-radius: 8px; padding: 16px 20px; margin-bottom: 20px; font-size: 0.92rem;
}
.batch-badge {
    display: inline-block; padding: 3px 12px; border-radius: 12px;
    background: #e8f4fd; color: #0c5460; font-weight: bold; font-size: 0.9rem;
}
.weight-card {
    background: #f8f9fa; border-left: 4px solid #2e7d32;
    padding: 10px 14px; border-radius: 6px; margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
if "postharvest_batches" not in st.session_state:
    st.session_state.postharvest_batches = {}

# ── Upload parser ──────────────────────────────────────────────────────────────

def _read_section(ws, section_row):
    """Read column headers (section_row+1) and data rows (section_row+3 to section_row+24)."""
    header_r = section_row + 1
    data_start = section_row + 3
    data_end = min(section_row + 24, ws.max_row)
    headers = [ws.cell(header_r, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h is not None else f"_col{c}" for c, h in enumerate(headers, 1)]
    records = []
    for r in range(data_start, data_end + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if any(v is not None for v in vals):
            records.append(dict(zip(headers, vals)))
    return records


def parse_postharvest_excel(file_bytes):
    """Parse a PostHarvest_DataCollection_Template.xlsx and return list of batch dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    batches = {}

    # Tab 1: HarvestBatchLog — sections at rows 4 (ID) and 29 (plant count)
    if "1_HarvestBatch_Log" in wb.sheetnames:
        ws = wb["1_HarvestBatch_Log"]
        id_recs = _read_section(ws, 4)
        cnt_recs = _read_section(ws, 29)
        for i, rec in enumerate(id_recs):
            bid = str(rec.get("Harvest Batch Name / ID", f"Batch_{i+1}"))
            batches.setdefault(bid, {})["key"] = bid
            batches[bid]["harvest_date"] = rec.get("Harvest Date (mm/dd/yyyy)", "")
            batches[bid]["flower_room"] = rec.get("Flower Room / Zone", "")
            if i < len(cnt_recs):
                batches[bid]["initial_plant_count"] = cnt_recs[i].get("Initial Plant Count")
                batches[bid]["harvested_plant_count"] = cnt_recs[i].get("Harvested Plant Count")
                batches[bid]["plant_loss_notes"] = cnt_recs[i].get("Plant Loss Notes", "")

    # Tab 2: ProcessingWeights — sections at rows 4 (harvest), 29 (bucking), 54 (drying), 79 (trimming)
    if "2_Processing_Weights" in wb.sheetnames:
        ws = wb["2_Processing_Weights"]
        sections = {4: "harvest", 29: "bucking", 54: "drying", 79: "trimming"}
        for sec_row, label in sections.items():
            recs = _read_section(ws, sec_row)
            for i, rec in enumerate(recs):
                bid = str(rec.get("Harvest Batch Name / ID", ""))
                if bid in batches:
                    batches[bid][f"processing_{label}"] = rec

    # Tab 3: Curing
    if "3_Curing" in wb.sheetnames:
        ws = wb["3_Curing"]
        entry_recs = _read_section(ws, 4)
        log_recs = _read_section(ws, 29)
        for rec in entry_recs:
            bid = str(rec.get("Harvest Batch Name / ID", ""))
            if bid in batches:
                batches[bid]["curing_entry"] = rec
        for rec in log_recs:
            bid = str(rec.get("Harvest Batch Name / ID", ""))
            if bid in batches:
                batches[bid].setdefault("curing_log", []).append(rec)

    # Tab 4: Quality Testing
    if "4_Quality_Testing" in wb.sheetnames:
        ws = wb["4_Quality_Testing"]
        sub_recs = _read_section(ws, 4)
        res_recs = _read_section(ws, 29)
        for i, rec in enumerate(sub_recs):
            bid = str(rec.get("Harvest Batch Name / ID", ""))
            if bid in batches:
                batches[bid]["quality_submission"] = rec
                if i < len(res_recs):
                    batches[bid]["quality_results"] = res_recs[i]

    # Tab 5: PackagingInventory
    if "5_Packaging_Inventory" in wb.sheetnames:
        ws = wb["5_Packaging_Inventory"]
        run_recs = _read_section(ws, 4)
        sku_recs = _read_section(ws, 29)
        for i, rec in enumerate(run_recs):
            bid = str(rec.get("Harvest Batch Name / ID", ""))
            if bid in batches:
                batches[bid]["packaging_run"] = rec
                if i < len(sku_recs):
                    batches[bid]["packaging_sku"] = sku_recs[i]

    # Tab 6: Sales Tracking
    if "6_Sales_Tracking" in wb.sheetnames:
        ws = wb["6_Sales_Tracking"]
        sale_recs = _read_section(ws, 4)
        buyer_recs = _read_section(ws, 29)
        for i, rec in enumerate(sale_recs):
            bid = str(rec.get("Harvest Batch Name / ID", ""))
            if bid in batches:
                batches[bid]["sales"] = rec
                if i < len(buyer_recs):
                    batches[bid]["sales_buyer"] = buyer_recs[i]

    return [{"key": k, **v} for k, v in batches.items()]


def _map_imported_batch(raw):
    """Flatten imported parsed data into our session state data dict."""
    d = {}
    d["flower_room"] = raw.get("flower_room", "")
    d["harvest_date"] = str(raw.get("harvest_date", ""))
    d["initial_plant_count"] = raw.get("initial_plant_count")
    d["harvested_plant_count"] = raw.get("harvested_plant_count")
    d["plant_loss_notes"] = raw.get("plant_loss_notes", "")

    ph = raw.get("processing_harvest", {})
    d["total_wet_weight_g"] = ph.get("TOTAL Wet Weight (g)")
    d["harvest_waste_g"] = ph.get("Total Waste From Harvest (g)")

    pb = raw.get("processing_bucking", {})
    d["bucking_date"] = str(pb.get("Date Ran", ""))
    d["bucked_weight_g"] = pb.get("Bucked Weight (g)")
    d["bucking_waste_g"] = pb.get("Weight of Waste from Bucking (g)")

    pd_ = raw.get("processing_drying", {})
    d["drying_date"] = str(pd_.get("Date Ran", ""))
    d["dried_weight_g"] = pd_.get("Dried Weight (g)")
    d["drying_waste_g"] = pd_.get("Weight of Waste from Drying (g)")
    d["dry_time_days"] = pd_.get("Dry Time (days)")

    pt = raw.get("processing_trimming", {})
    d["trim_date"] = str(pt.get("Date Ran", ""))
    d["trimmed_flower_g"] = pt.get("Trimmed Flower Weight (g)")
    d["trim_weight_g"] = pt.get("Trim Weight (g)")
    d["trim_waste_g"] = pt.get("Trim Waste (g)")

    ce = raw.get("curing_entry", {})
    d["cure_label"] = str(ce.get("Strain / Harvest Batch Label", ""))
    d["cure_entry_date"] = str(ce.get("Date Brought Into Cure Room", ""))
    d["cure_flower_entry_g"] = ce.get("Flower Weight at Cure Entry (g)")
    d["cure_trim_entry_g"] = ce.get("Trim Weight at Cure Entry (g)")

    qs = raw.get("quality_submission", {})
    qr = raw.get("quality_results", {})
    d["lab_date_submitted"] = str(qs.get("Date Submitted to Lab", ""))
    d["lab_name"] = str(qs.get("Lab Name", ""))
    d["sample_id"] = str(qs.get("Sample ID / Tracking #", ""))
    d["date_tested"] = str(qr.get("Date Tested", ""))
    d["date_received"] = str(qr.get("Date Results Received", ""))
    d["test_result"] = str(qr.get("Result (Pass / Fail / Retest Pass / Retest Fail)", ""))
    d["fail_reason"] = str(qr.get("If Failed — Why?", ""))
    d["coa_link"] = str(qr.get("Attach COA (file path or link)", ""))

    pr = raw.get("packaging_run", {})
    ps = raw.get("packaging_sku", {})
    d["packaging_date"] = str(pr.get("Date Entered Into Packaging", ""))
    d["flower_available_for_packaging_g"] = pr.get("Total Flower Available for Packaging (g)")
    d["pkg_1g"] = ps.get("1 Gram (units)")
    d["pkg_eighth"] = ps.get("Eighths / 3.5g (units)")
    d["pkg_quarter"] = ps.get("Quarters / 7g (units)")
    d["pkg_half"] = ps.get("Half Oz / 14g (units)")
    d["pkg_oz"] = ps.get("1 Oz / 28g (units)")
    d["pkg_prerolls"] = ps.get("Pre-Rolls (units)")
    d["pkg_other"] = str(ps.get("Other SKU — describe in notes", ""))

    sl = raw.get("sales", {})
    sb = raw.get("sales_buyer", {})
    d["trim_sold_g"] = sl.get("Trim Total Sold (g)")
    d["trim_price_per_g"] = sl.get("Trim Price ($/g)")
    d["trim_revenue"] = sl.get("Trim Revenue ($)")
    d["popcorn_sold_g"] = sl.get("Popcorn / Smalls Sold (g)")
    d["popcorn_price_per_g"] = sl.get("Popcorn / Smalls Price ($/g)")
    d["other_byproduct"] = str(sl.get("Other Byproduct Sold — describe", ""))
    d["buyer_name"] = str(sb.get("Buyer / Dispensary Name", ""))
    d["sale_date"] = str(sb.get("Sale Date", ""))
    d["invoice_num"] = str(sb.get("Invoice / PO #", ""))
    d["sale_notes"] = str(sb.get("Notes", ""))

    curing_log = raw.get("curing_log", [])
    return d, pd.DataFrame(curing_log) if curing_log else None


# ── Excel download builder ─────────────────────────────────────────────────────

def build_postharvest_excel(batches_dict):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "1_Harvest_Batch_Log"
    ws1.append([
        "Harvest Batch ID", "Flower Room / Zone", "Harvest Date",
        "Initial Plant Count", "Harvested Plant Count", "Plant Loss (#)", "Plant Loss Notes",
    ])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        initial = d.get("initial_plant_count") or 0
        harvested = d.get("harvested_plant_count") or 0
        ws1.append([
            bid, d.get("flower_room", ""), d.get("harvest_date", ""),
            initial or "", harvested or "",
            (initial - harvested) if (initial and harvested) else "",
            d.get("plant_loss_notes", ""),
        ])

    ws2 = wb.create_sheet("2_Processing_Weights")
    ws2.append([
        "Harvest Batch ID",
        "Total Wet Weight (g)", "Harvest Waste (g)",
        "Bucking Date", "Bucked Weight (g)", "Bucking Waste (g)",
        "Drying Date", "Dried Weight (g)", "Drying Waste (g)", "Dry Time (days)",
        "Trim Date", "Trimmed Flower (g)", "Trim Weight (g)", "Trim Waste (g)",
        "Moisture Loss (%)",
    ])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        wet = d.get("total_wet_weight_g") or 0
        dry = d.get("dried_weight_g") or 0
        moisture_loss = round((1 - dry / wet) * 100, 1) if wet and dry else ""
        ws2.append([
            bid,
            wet or "", d.get("harvest_waste_g") or "",
            d.get("bucking_date", ""), d.get("bucked_weight_g") or "", d.get("bucking_waste_g") or "",
            d.get("drying_date", ""), dry or "", d.get("drying_waste_g") or "", d.get("dry_time_days") or "",
            d.get("trim_date", ""), d.get("trimmed_flower_g") or "", d.get("trim_weight_g") or "",
            d.get("trim_waste_g") or "", moisture_loss,
        ])

    ws3 = wb.create_sheet("3_Curing")
    ws3.append(["Harvest Batch ID", "Label", "Date Into Cure Room",
                "Flower Weight Entry (g)", "Trim Weight Entry (g)"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws3.append([bid, d.get("cure_label", ""), d.get("cure_entry_date", ""),
                    d.get("cure_flower_entry_g") or "", d.get("cure_trim_entry_g") or ""])
    ws3.append([])
    ws3.append(["Harvest Batch ID", "Check Date", "Days Since Start",
                "Flower Weight (g)", "Trim Weight (g)", "RH %", "Action Taken"])
    for bid, b in batches_dict.items():
        cdf = b.get("curing_log_df")
        if cdf is not None and not cdf.empty:
            for _, row in cdf.iterrows():
                ws3.append([bid] + list(row))

    ws4 = wb.create_sheet("4_Quality_Testing")
    ws4.append(["Harvest Batch ID", "Date Submitted", "Lab Name", "Sample ID / Tracking #",
                "Date Tested", "Date Results Received", "Result", "If Failed — Why?", "COA Link"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws4.append([bid, d.get("lab_date_submitted", ""), d.get("lab_name", ""), d.get("sample_id", ""),
                    d.get("date_tested", ""), d.get("date_received", ""), d.get("test_result", ""),
                    d.get("fail_reason", ""), d.get("coa_link", "")])

    ws5 = wb.create_sheet("5_Packaging_Inventory")
    ws5.append(["Harvest Batch ID", "Date Packaged", "Flower Available (g)",
                "1g Units", "Eighths (3.5g)", "Quarters (7g)", "Half Oz (14g)", "1 Oz (28g)",
                "Pre-Rolls (units)", "Other SKU", "End of Stock Date"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws5.append([bid, d.get("packaging_date", ""), d.get("flower_available_for_packaging_g") or "",
                    d.get("pkg_1g") or "", d.get("pkg_eighth") or "", d.get("pkg_quarter") or "",
                    d.get("pkg_half") or "", d.get("pkg_oz") or "", d.get("pkg_prerolls") or "",
                    d.get("pkg_other", ""), d.get("end_stock_date", "")])

    ws6 = wb.create_sheet("6_Sales_Tracking")
    ws6.append(["Harvest Batch ID", "Trim Sold (g)", "Trim Price ($/g)", "Trim Revenue ($)",
                "Popcorn Sold (g)", "Popcorn Price ($/g)", "Other Byproduct",
                "Buyer / Dispensary", "Sale Date", "Invoice / PO #", "Notes"])
    for bid, b in batches_dict.items():
        d = b.get("data", {})
        ws6.append([bid, d.get("trim_sold_g") or "", d.get("trim_price_per_g") or "",
                    d.get("trim_revenue") or "", d.get("popcorn_sold_g") or "",
                    d.get("popcorn_price_per_g") or "", d.get("other_byproduct", ""),
                    d.get("buyer_name", ""), d.get("sale_date", ""),
                    d.get("invoice_num", ""), d.get("sale_notes", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Page layout ────────────────────────────────────────────────────────────────

st.markdown("""
<div class="disclaimer-box">
<b>Record-keeping tool only.</b> Not a compliance filing system. Always maintain records consistent
with NYS OCM requirements. Data is stored in your browser session — download before closing.
</div>
""", unsafe_allow_html=True)

st.title("🌾 Post-Harvest Data Collection")
st.caption("Harvest log, processing weights, curing, quality testing, packaging, and sales — NYS cannabis & hemp")
st.divider()

# ── Upload ─────────────────────────────────────────────────────────────────────
with st.expander("📤 Upload Existing PostHarvest Excel Template", expanded=False):
    st.markdown(
        "Upload a filled `PostHarvest_DataCollection_Template.xlsx` to import existing data."
    )
    if not HAS_OPENPYXL:
        st.error("openpyxl is not installed. Run `pip install openpyxl` to enable Excel upload/download.")
    else:
        uploaded = st.file_uploader("Choose file", type=["xlsx"], key="postharvest_upload")
        if uploaded:
            with st.spinner("Parsing…"):
                try:
                    parsed = parse_postharvest_excel(uploaded.read())
                    imported = 0
                    for raw in parsed:
                        k = raw.get("key", "")
                        if k and k not in ("None", ""):
                            mapped_data, curing_df = _map_imported_batch(raw)
                            st.session_state.postharvest_batches[k] = {
                                "data": mapped_data,
                                "curing_log_df": curing_df,
                            }
                            imported += 1
                    if imported:
                        st.success(f"Imported {imported} batch(es).")
                        st.rerun()
                    else:
                        st.warning("No data rows found. Make sure the template has been filled in.")
                except Exception as e:
                    st.error(f"Could not parse file: {e}")

# ── Batch management ───────────────────────────────────────────────────────────
st.markdown("### Batch Management")
st.caption("Harvest Batch Name / ID must match the Batch Number used in the Pre-Harvest tool to link records.")

col_sel, col_new_id, col_btn = st.columns([3, 2, 1])
with col_sel:
    batch_list = list(st.session_state.postharvest_batches.keys())
    # Also suggest unlinked pre-harvest batch IDs
    ph_ids = [k for k in st.session_state.get("preharvest_batches", {}).keys()
              if k not in st.session_state.postharvest_batches]
    if ph_ids:
        st.caption(f"Pre-harvest batches not yet linked: **{', '.join(ph_ids)}**")

    selected_batch = st.selectbox(
        "Select batch to view / edit",
        batch_list if batch_list else ["(none)"],
        key="postharvest_selected",
        disabled=not batch_list,
    )
    if not batch_list:
        selected_batch = None

with col_new_id:
    new_bid = st.text_input("New batch ID", placeholder="Must match Pre-Harvest Batch Number",
                             key="ph_new_bid_post", label_visibility="collapsed")
with col_btn:
    st.write("")
    if st.button("➕ Create", use_container_width=True):
        bid = new_bid.strip()
        if not bid:
            st.error("Enter a batch ID.")
        elif bid in st.session_state.postharvest_batches:
            st.error(f"'{bid}' already exists.")
        else:
            st.session_state.postharvest_batches[bid] = {"data": {}, "curing_log_df": None}
            st.session_state.postharvest_selected = bid
            st.rerun()

if not selected_batch or selected_batch == "(none)":
    st.info("Create a new batch or upload an existing template to get started.")
    st.stop()

batch = st.session_state.postharvest_batches[selected_batch]
bd = batch.setdefault("data", {})

col_lbl, col_del = st.columns([8, 1])
with col_lbl:
    st.markdown(f"Editing: <span class='batch-badge'>🌾 {selected_batch}</span>", unsafe_allow_html=True)
with col_del:
    if st.button("🗑 Delete", use_container_width=True, help="Remove this batch from the session"):
        del st.session_state.postharvest_batches[selected_batch]
        st.rerun()

st.write("")

# ── Data entry tabs ────────────────────────────────────────────────────────────
tabs = st.tabs([
    "📦 Harvest Log",
    "⚖️ Processing Weights",
    "🫙 Curing",
    "🔬 Quality Testing",
    "📦 Packaging",
    "💵 Sales Tracking",
])

# ── TAB 1: Harvest Batch Log ───────────────────────────────────────────────────
with tabs[0]:
    st.markdown("#### Harvest Batch Log")
    st.caption("Record batch identity and plant counts at time of harvest.")

    c1, c2 = st.columns(2)
    with c1:
        bd["flower_room"] = st.text_input("Flower Room / Zone", value=bd.get("flower_room", ""),
            placeholder="e.g. Greenhouse A, Field 2", key=f"pos_{selected_batch}_room")
        bd["harvest_date"] = st.text_input("Harvest Date (mm/dd/yyyy)",
            value=bd.get("harvest_date", ""), key=f"pos_{selected_batch}_hdate")
    with c2:
        bd["initial_plant_count"] = st.number_input("Initial Plant Count", min_value=0, step=1,
            value=int(bd.get("initial_plant_count") or 0), key=f"pos_{selected_batch}_init_cnt")
        bd["harvested_plant_count"] = st.number_input("Harvested Plant Count", min_value=0, step=1,
            value=int(bd.get("harvested_plant_count") or 0), key=f"pos_{selected_batch}_harv_cnt")

    initial = int(bd.get("initial_plant_count") or 0)
    harvested = int(bd.get("harvested_plant_count") or 0)
    plant_loss = initial - harvested if initial >= harvested else 0
    if initial > 0:
        pc1, pc2, pc3 = st.columns(3)
        pc1.metric("Initial Count", initial)
        pc2.metric("Harvested", harvested)
        pc3.metric("Plant Loss", plant_loss,
                   delta=f"{-plant_loss}" if plant_loss > 0 else "None",
                   delta_color="inverse" if plant_loss > 0 else "off")

    bd["plant_loss_notes"] = st.text_area("Plant Loss Notes", value=bd.get("plant_loss_notes", ""),
        placeholder="Reason for any plant loss (disease, hermie, compliance, etc.)",
        key=f"pos_{selected_batch}_loss_notes", height=80)

# ── TAB 2: Processing Weights ──────────────────────────────────────────────────
with tabs[1]:
    st.markdown("#### Processing Weights")
    st.caption("Track weight at every step from harvest through finished trim. All weights in grams (g).")

    st.markdown("**Harvest Weights**")
    hw1, hw2 = st.columns(2)
    with hw1:
        bd["total_wet_weight_g"] = st.number_input("Total Wet Weight (g)", min_value=0.0, step=10.0,
            value=float(bd.get("total_wet_weight_g") or 0.0), key=f"pos_{selected_batch}_wet_wt")
    with hw2:
        bd["harvest_waste_g"] = st.number_input("Waste From Harvest (g)", min_value=0.0, step=1.0,
            help="Stems, fan leaves, debris discarded at harvest",
            value=float(bd.get("harvest_waste_g") or 0.0), key=f"pos_{selected_batch}_harv_waste")

    st.divider()
    st.markdown("**Bucking**")
    bk1, bk2, bk3 = st.columns(3)
    with bk1:
        bd["bucking_date"] = st.text_input("Bucking Date (mm/dd/yyyy)",
            value=bd.get("bucking_date", ""), key=f"pos_{selected_batch}_buck_date")
    with bk2:
        bd["bucked_weight_g"] = st.number_input("Bucked Weight (g)", min_value=0.0, step=10.0,
            value=float(bd.get("bucked_weight_g") or 0.0), key=f"pos_{selected_batch}_buck_wt")
    with bk3:
        bd["bucking_waste_g"] = st.number_input("Waste from Bucking (g)", min_value=0.0, step=1.0,
            value=float(bd.get("bucking_waste_g") or 0.0), key=f"pos_{selected_batch}_buck_waste")

    st.divider()
    st.markdown("**Drying**")
    dr1, dr2, dr3, dr4 = st.columns(4)
    with dr1:
        bd["drying_date"] = st.text_input("Drying Date (mm/dd/yyyy)",
            value=bd.get("drying_date", ""), key=f"pos_{selected_batch}_dry_date")
    with dr2:
        bd["dried_weight_g"] = st.number_input("Dried Weight (g)", min_value=0.0, step=10.0,
            value=float(bd.get("dried_weight_g") or 0.0), key=f"pos_{selected_batch}_dry_wt")
    with dr3:
        bd["drying_waste_g"] = st.number_input("Waste from Drying (g)", min_value=0.0, step=1.0,
            value=float(bd.get("drying_waste_g") or 0.0), key=f"pos_{selected_batch}_dry_waste")
    with dr4:
        bd["dry_time_days"] = st.number_input("Dry Time (days)", min_value=0, step=1,
            value=int(bd.get("dry_time_days") or 0), key=f"pos_{selected_batch}_dry_days")

    st.divider()
    st.markdown("**Trimming**")
    tr1, tr2, tr3, tr4 = st.columns(4)
    with tr1:
        bd["trim_date"] = st.text_input("Trim Date (mm/dd/yyyy)",
            value=bd.get("trim_date", ""), key=f"pos_{selected_batch}_trim_date")
    with tr2:
        bd["trimmed_flower_g"] = st.number_input("Trimmed Flower (g)", min_value=0.0, step=10.0,
            value=float(bd.get("trimmed_flower_g") or 0.0), key=f"pos_{selected_batch}_trim_fl")
    with tr3:
        bd["trim_weight_g"] = st.number_input("Trim Weight (g)", min_value=0.0, step=1.0,
            value=float(bd.get("trim_weight_g") or 0.0), key=f"pos_{selected_batch}_trim_wt")
    with tr4:
        bd["trim_waste_g"] = st.number_input("Trim Waste (g)", min_value=0.0, step=1.0,
            value=float(bd.get("trim_waste_g") or 0.0), key=f"pos_{selected_batch}_trim_waste")

    # Weight flow summary
    wet = bd.get("total_wet_weight_g") or 0
    dried = bd.get("dried_weight_g") or 0
    trimmed = bd.get("trimmed_flower_g") or 0
    trim_byproduct = bd.get("trim_weight_g") or 0
    if wet > 0:
        st.divider()
        st.markdown("**Weight Flow Summary**")
        wf1, wf2, wf3, wf4 = st.columns(4)
        wf1.metric("Wet Weight", f"{wet:,.0f} g")
        wf2.metric("Dried Weight", f"{dried:,.0f} g",
                   delta=f"{((dried/wet)-1)*100:.1f}% vs wet" if wet else None, delta_color="off")
        wf3.metric("Trimmed Flower", f"{trimmed:,.0f} g")
        wf4.metric("Trim / Byproduct", f"{trim_byproduct:,.0f} g")
        if wet and dried:
            moisture_loss = round((1 - dried / wet) * 100, 1)
            st.caption(f"Moisture loss: **{moisture_loss}%** | "
                       f"Dry/wet ratio: **{dried/wet:.3f}** | "
                       f"Flower recovery from dry: **{trimmed/dried*100:.1f}%**" if dried else "")

# ── TAB 3: Curing ──────────────────────────────────────────────────────────────
with tabs[2]:
    st.markdown("#### Curing")

    st.markdown("**Cure Room Entry**")
    ce1, ce2 = st.columns(2)
    with ce1:
        bd["cure_label"] = st.text_input("Strain / Batch Label on Container",
            value=bd.get("cure_label", ""), key=f"pos_{selected_batch}_cure_label")
        bd["cure_entry_date"] = st.text_input("Date Brought Into Cure Room (mm/dd/yyyy)",
            value=bd.get("cure_entry_date", ""), key=f"pos_{selected_batch}_cure_date")
    with ce2:
        bd["cure_flower_entry_g"] = st.number_input("Flower Weight at Cure Entry (g)",
            min_value=0.0, step=1.0,
            value=float(bd.get("cure_flower_entry_g") or 0.0), key=f"pos_{selected_batch}_cure_fl")
        bd["cure_trim_entry_g"] = st.number_input("Trim Weight at Cure Entry (g)",
            min_value=0.0, step=1.0,
            value=float(bd.get("cure_trim_entry_g") or 0.0), key=f"pos_{selected_batch}_cure_tr")

    st.divider()
    st.markdown("**Cure Monitoring Log**")
    st.caption("Add a row each time you check or burp the cure containers.")

    _cure_schema = {
        "Check Date (mm/dd/yy)": pd.Series(dtype="str"),
        "Days Since Cure Start": pd.Series(dtype="Int64"),
        "Flower Weight (g)": pd.Series(dtype="Float64"),
        "Trim Weight (g)": pd.Series(dtype="Float64"),
        "RH % Inside Container": pd.Series(dtype="Float64"),
        "Action Taken": pd.Series(dtype="str"),
    }
    existing_c = batch.get("curing_log_df")
    default_c = pd.DataFrame(_cure_schema) if (existing_c is None or existing_c.empty) else existing_c

    edited_c = st.data_editor(
        default_c,
        num_rows="dynamic",
        use_container_width=True,
        key=f"cure_editor_{selected_batch}",
        column_config={
            "Days Since Cure Start": st.column_config.NumberColumn("Days", min_value=0, step=1),
            "Flower Weight (g)": st.column_config.NumberColumn("Flower (g)", min_value=0.0, format="%.1f"),
            "Trim Weight (g)": st.column_config.NumberColumn("Trim (g)", min_value=0.0, format="%.1f"),
            "RH % Inside Container": st.column_config.NumberColumn("RH %", min_value=0.0, max_value=100.0, format="%.1f"),
            "Action Taken": st.column_config.SelectboxColumn("Action",
                options=["Burp", "Seal", "Note only", "Removed from cure", "Other"]),
        },
    )
    st.session_state.postharvest_batches[selected_batch]["curing_log_df"] = edited_c
    if not edited_c.empty:
        st.caption(f"{len(edited_c)} cure check(s) logged")

# ── TAB 4: Quality Testing (COA) ──────────────────────────────────────────────
with tabs[3]:
    st.markdown("#### Quality Testing — Certificate of Analysis (COA)")

    st.markdown("**Lab Submission**")
    qt1, qt2 = st.columns(2)
    with qt1:
        bd["lab_date_submitted"] = st.text_input("Date Submitted to Lab (mm/dd/yyyy)",
            value=bd.get("lab_date_submitted", ""), key=f"pos_{selected_batch}_lab_sub")
        bd["lab_name"] = st.text_input("Lab Name", value=bd.get("lab_name", ""),
            key=f"pos_{selected_batch}_lab_name")
    with qt2:
        bd["sample_id"] = st.text_input("Sample ID / Tracking #", value=bd.get("sample_id", ""),
            key=f"pos_{selected_batch}_sample_id")

    st.divider()
    st.markdown("**Test Results**")
    st.caption("THC% and CBD% go in Pre-Harvest → Lab Results tab.")

    tr1, tr2 = st.columns(2)
    with tr1:
        bd["date_tested"] = st.text_input("Date Tested (mm/dd/yyyy)",
            value=bd.get("date_tested", ""), key=f"pos_{selected_batch}_dt_tested")
        bd["date_received"] = st.text_input("Date Results Received (mm/dd/yyyy)",
            value=bd.get("date_received", ""), key=f"pos_{selected_batch}_dt_received")
    with tr2:
        result_opts = ["—", "Pass", "Fail", "Retest Pass", "Retest Fail", "Pending"]
        cur_res = bd.get("test_result", "—")
        bd["test_result"] = st.selectbox("Result", result_opts,
            index=result_opts.index(cur_res) if cur_res in result_opts else 0,
            key=f"pos_{selected_batch}_result")

    if bd.get("test_result") in ("Fail", "Retest Fail"):
        bd["fail_reason"] = st.text_area("If Failed — Why?",
            value=bd.get("fail_reason", ""),
            placeholder="e.g. Pesticides detected, mold count exceeded limit, potency out of range…",
            key=f"pos_{selected_batch}_fail_reason", height=80)
    else:
        bd["fail_reason"] = bd.get("fail_reason", "")

    bd["coa_link"] = st.text_input("COA File Path or Link",
        value=bd.get("coa_link", ""),
        placeholder="e.g. /files/coa_batch001.pdf  or  https://lab.example.com/coa/…",
        key=f"pos_{selected_batch}_coa")

    # Result badge
    result = bd.get("test_result", "—")
    if result == "Pass" or result == "Retest Pass":
        st.success(f"✅ COA Result: **{result}**")
    elif result in ("Fail", "Retest Fail"):
        st.error(f"❌ COA Result: **{result}** — {bd.get('fail_reason', '')}")
    elif result == "Pending":
        st.warning("⏳ Results pending")

# ── TAB 5: Packaging & Inventory ──────────────────────────────────────────────
with tabs[4]:
    st.markdown("#### Packaging & Inventory")

    st.markdown("**Packaging Run**")
    pk1, pk2 = st.columns(2)
    with pk1:
        bd["packaging_date"] = st.text_input("Date Entered Into Packaging (mm/dd/yyyy)",
            value=bd.get("packaging_date", ""), key=f"pos_{selected_batch}_pkg_date")
    with pk2:
        bd["flower_available_for_packaging_g"] = st.number_input(
            "Total Flower Available for Packaging (g)", min_value=0.0, step=1.0,
            value=float(bd.get("flower_available_for_packaging_g") or 0.0),
            key=f"pos_{selected_batch}_fl_avail",
            help="Cured flower weight entering packaging — from Curing tab")

    st.divider()
    st.markdown("**Packaged Units by SKU**")
    sk1, sk2, sk3 = st.columns(3)
    with sk1:
        bd["pkg_1g"] = st.number_input("1 Gram (units)", min_value=0, step=1,
            value=int(bd.get("pkg_1g") or 0), key=f"pos_{selected_batch}_pkg_1g")
        bd["pkg_eighth"] = st.number_input("Eighths / 3.5g (units)", min_value=0, step=1,
            value=int(bd.get("pkg_eighth") or 0), key=f"pos_{selected_batch}_pkg_8th")
    with sk2:
        bd["pkg_quarter"] = st.number_input("Quarters / 7g (units)", min_value=0, step=1,
            value=int(bd.get("pkg_quarter") or 0), key=f"pos_{selected_batch}_pkg_qtr")
        bd["pkg_half"] = st.number_input("Half Oz / 14g (units)", min_value=0, step=1,
            value=int(bd.get("pkg_half") or 0), key=f"pos_{selected_batch}_pkg_half")
    with sk3:
        bd["pkg_oz"] = st.number_input("1 Oz / 28g (units)", min_value=0, step=1,
            value=int(bd.get("pkg_oz") or 0), key=f"pos_{selected_batch}_pkg_oz")
        bd["pkg_prerolls"] = st.number_input("Pre-Rolls (units)", min_value=0, step=1,
            value=int(bd.get("pkg_prerolls") or 0), key=f"pos_{selected_batch}_pkg_pr")

    bd["pkg_other"] = st.text_input("Other SKU — describe",
        value=bd.get("pkg_other", ""), key=f"pos_{selected_batch}_pkg_other")
    bd["end_stock_date"] = st.text_input("End of Stock Date (mm/dd/yyyy)",
        value=bd.get("end_stock_date", ""), key=f"pos_{selected_batch}_end_stock")

    # Total packaged weight
    total_packaged_g = (
        (bd.get("pkg_1g") or 0) * 1 +
        (bd.get("pkg_eighth") or 0) * 3.5 +
        (bd.get("pkg_quarter") or 0) * 7 +
        (bd.get("pkg_half") or 0) * 14 +
        (bd.get("pkg_oz") or 0) * 28
    )
    available = bd.get("flower_available_for_packaging_g") or 0
    if total_packaged_g > 0:
        st.divider()
        pg1, pg2, pg3 = st.columns(3)
        pg1.metric("Total Packaged (g)", f"{total_packaged_g:,.1f}")
        if available:
            unpackaged = available - total_packaged_g
            pg2.metric("Available (g)", f"{available:,.1f}")
            pg3.metric("Remaining / Variance (g)", f"{unpackaged:+,.1f}",
                       delta_color="normal" if unpackaged >= 0 else "inverse")

# ── TAB 6: Sales Tracking ──────────────────────────────────────────────────────
with tabs[5]:
    st.markdown("#### Sales Tracking — Trim & Byproducts")
    st.caption("Flower, pre-roll, and biomass sales go in Pre-Harvest → Yield & Selling tab. "
               "Track trim and other byproducts here.")

    st.markdown("**Trim & Byproduct Sales**")
    s1, s2 = st.columns(2)
    with s1:
        bd["trim_sold_g"] = st.number_input("Trim Total Sold (g)", min_value=0.0, step=1.0,
            value=float(bd.get("trim_sold_g") or 0.0), key=f"pos_{selected_batch}_trim_sold")
        bd["trim_price_per_g"] = st.number_input("Trim Price ($/g)", min_value=0.0, step=0.01,
            value=float(bd.get("trim_price_per_g") or 0.0),
            key=f"pos_{selected_batch}_trim_price", format="%.3f")
        trim_rev = round((bd.get("trim_sold_g") or 0) * (bd.get("trim_price_per_g") or 0), 2)
        bd["trim_revenue"] = trim_rev
        if trim_rev > 0:
            st.metric("Trim Revenue", f"${trim_rev:,.2f}")
    with s2:
        bd["popcorn_sold_g"] = st.number_input("Popcorn / Smalls Sold (g)", min_value=0.0, step=1.0,
            value=float(bd.get("popcorn_sold_g") or 0.0), key=f"pos_{selected_batch}_pop_sold")
        bd["popcorn_price_per_g"] = st.number_input("Popcorn / Smalls Price ($/g)", min_value=0.0, step=0.01,
            value=float(bd.get("popcorn_price_per_g") or 0.0),
            key=f"pos_{selected_batch}_pop_price", format="%.3f")
        pop_rev = round((bd.get("popcorn_sold_g") or 0) * (bd.get("popcorn_price_per_g") or 0), 2)
        if pop_rev > 0:
            st.metric("Popcorn Revenue", f"${pop_rev:,.2f}")

    bd["other_byproduct"] = st.text_input("Other Byproduct — describe",
        value=bd.get("other_byproduct", ""),
        placeholder="e.g. extract input material, 500g @ $2.00/g",
        key=f"pos_{selected_batch}_other_byp")

    st.divider()
    st.markdown("**Buyer / Channel (optional)**")
    b1, b2, b3 = st.columns(3)
    with b1:
        bd["buyer_name"] = st.text_input("Buyer / Dispensary Name",
            value=bd.get("buyer_name", ""), key=f"pos_{selected_batch}_buyer")
    with b2:
        bd["sale_date"] = st.text_input("Sale Date (mm/dd/yyyy)",
            value=bd.get("sale_date", ""), key=f"pos_{selected_batch}_sale_date")
    with b3:
        bd["invoice_num"] = st.text_input("Invoice / PO #",
            value=bd.get("invoice_num", ""), key=f"pos_{selected_batch}_invoice")
    bd["sale_notes"] = st.text_area("Notes", value=bd.get("sale_notes", ""),
        key=f"pos_{selected_batch}_sale_notes", height=60)

# ── Download ───────────────────────────────────────────────────────────────────
st.divider()
st.markdown("### Download Data")

dl1, dl2, dl3 = st.columns(3)

with dl1:
    if st.session_state.postharvest_batches:
        rows = []
        for bn, b in st.session_state.postharvest_batches.items():
            row = {"Harvest Batch ID": bn}
            row.update(b.get("data", {}))
            rows.append(row)
        csv_bytes = pd.DataFrame(rows).to_csv(index=False).encode()
        st.download_button("⬇ Batch Summary (CSV)", data=csv_bytes,
                           file_name="postharvest_batch_summary.csv", mime="text/csv",
                           use_container_width=True)

with dl2:
    if HAS_OPENPYXL and st.session_state.postharvest_batches:
        xl = build_postharvest_excel(st.session_state.postharvest_batches)
        st.download_button("⬇ Full Data (Excel)", data=xl,
                           file_name="postharvest_data.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

with dl3:
    all_curing = []
    for bn, b in st.session_state.postharvest_batches.items():
        cdf = b.get("curing_log_df")
        if cdf is not None and not cdf.empty:
            tmp = cdf.copy()
            tmp.insert(0, "Harvest Batch ID", bn)
            all_curing.append(tmp)
    if all_curing:
        combined = pd.concat(all_curing, ignore_index=True)
        st.download_button("⬇ Curing Log (CSV)", data=combined.to_csv(index=False).encode(),
                           file_name="curing_log.csv", mime="text/csv",
                           use_container_width=True)

st.caption("Session data is not persisted after the browser tab closes. Download your data to keep it.")
